"""
annual.py
Laporan keuangan TAHUNAN: Ringkasan Eksekutif, Laba Rugi, Neraca (+ Saldo
Bulanan per Rekening), Arus Kas, Roster Gaji 12 Bulan, dan Analisis & Tren
- dibangun dari 12 file hasil rekonsiliasi bulanan yang SUDAH completed
(bulan harus berurutan, biasanya Januari-Desember).

Laporan ini murni RINGKASAN/KESIMPULAN setahun - tidak ada sheet rekening
mentah maupun detail transaksi, KECUALI satu sheet "Saldo Bulanan per
Rekening" yang menampilkan saldo awal & akhir tiap rekening per bulan
(diminta khusus - berguna untuk audit trail tanpa harus buka 12 file
terpisah, tapi tetap di level ringkasan saldo, bukan detail transaksi).

Reuse besar-besaran dari quarterly.py (fungsi-fungsi di sana sudah generik
terhadap jumlah bulan via parameter period_word) - modul ini menambahkan
validasi 12 bulan, sheet Saldo Bulanan per Rekening, dan Ringkasan
Eksekutif ala laporan tahunan UMKM restoran.

Catatan soal Desember: gaji Desember sering baru dibayar/dicatat Januari
tahun berikutnya (accrual). Roster Gaji (reuse dari quarterly.py) sudah
menangani ini - kolom Desember boleh kosong dengan catatan otomatis,
tidak dianggap error.
"""

import re
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import reconcile as rc
import quarterly as q

REPORT_SHEET_NAMES = q.REPORT_SHEET_NAMES


class AnnualInputError(Exception):
    """Dilempar kalau input tidak valid (bukan 12 bulan, bulan tidak
    berurutan, file rusak, dll) - pesannya dirancang untuk langsung
    ditunjukkan ke user di bot."""


# ---------------------------------------------------------------------------
# Baca & validasi 12 file bulanan
# ---------------------------------------------------------------------------

def load_month(path):
    return q.load_month(path)


def validate_consecutive_12(months):
    """months: list of dict dari load_month(), URUTAN SESUAI UPLOAD (belum
    tentu urut kronologis). Return list yang sama tapi sudah disortir
    kronologis. Raise AnnualInputError kalau bukan 12 bulan berturut-turut."""
    if len(months) != 12:
        raise AnnualInputError(f"Butuh tepat 12 file (satu tahun penuh), yang diterima {len(months)}.")
    ordered = sorted(months, key=lambda m: (m["year"], m["month"]))
    labels = [m["label"] for m in ordered]
    for i in range(1, 12):
        y0, m0 = ordered[i - 1]["year"], ordered[i - 1]["month"]
        y1, m1 = ordered[i]["year"], ordered[i]["month"]
        exp_m = 1 if m0 == 12 else m0 + 1
        exp_y = y0 + 1 if m0 == 12 else y0
        if (y1, m1) != (exp_y, exp_m):
            raise AnnualInputError(
                "Bulan yang diupload tidak berurutan: " + ", ".join(labels) +
                f". Setelah {rc.MONTHS_ID[m0]} {y0} seharusnya "
                f"{rc.MONTHS_ID[exp_m]} {exp_y}, bukan {rc.MONTHS_ID[m1]} {y1}."
            )
    return ordered


# ---------------------------------------------------------------------------
# Saldo Bulanan per Rekening - satu-satunya "detail" yang tetap ditampilkan
# di laporan tahunan ini (diminta khusus), di luar itu laporan murni
# kesimpulan/ringkasan. Nama sheet rekening beda tiap bulan (ada nama bulan
# & tahun nempel, mis. "BCA-887 Januari 2023" vs "BCA-887 Februari 2023"),
# jadi perlu "identitas dasar" rekening yang stabil buat menggabungkan
# rekening yang sama lintas 12 bulan.
# ---------------------------------------------------------------------------

_MONTH_NAMES_LOWER = {m.lower() for m in rc.MONTHS_ID[1:]}
_MONTH_ABBR_LOWER = {m[:3].lower() for m in rc.MONTHS_ID[1:]}


def account_base_name(sheet_title):
    """'BCA-887 Januari 2023' -> 'BCA-887'; 'BCA 0561864887 Jan 2025' ->
    'BCA 0561864887'; 'Kas-Buku April 2023' -> 'Kas-Buku'. Buang token
    terakhir kalau itu tahun 4 digit, lalu buang token terakhir lagi kalau
    itu nama/singkatan bulan."""
    tokens = sheet_title.split()
    if tokens and re.fullmatch(r"(19|20)\d{2}", tokens[-1]):
        tokens = tokens[:-1]
    if tokens and tokens[-1].lower() in (_MONTH_NAMES_LOWER | _MONTH_ABBR_LOWER):
        tokens = tokens[:-1]
    return " ".join(tokens).strip() or sheet_title


def load_month_accounts(path):
    """Baca satu file bulanan, return list of dict per sheet rekening:
    {'base_name', 'sheet_title', 'saldo_awal', 'saldo_akhir'}."""
    wb = openpyxl.load_workbook(path)
    account_sheets = [s for s in wb.sheetnames if s not in REPORT_SHEET_NAMES]
    result = []
    for sname in account_sheets:
        ws = wb[sname]
        txns, _ = rc.read_account_sheet(ws)
        opening, closing = q.compute_sheet_k(txns)
        result.append({
            "base_name": account_base_name(sname),
            "sheet_title": sname,
            "saldo_awal": opening,
            "saldo_akhir": closing,
        })
    return result


def write_saldo_bulanan_per_rekening(wb, months, months_accounts):
    """months_accounts: list (selaras urutan `months`) of list-of-dict dari
    load_month_accounts(). Tabel panjang: satu baris per (rekening, bulan),
    dikelompokkan per rekening supaya gampang lihat tren satu rekening."""
    name = "Saldo Bulanan per Rekening"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    labels = [m["label"] for m in months]
    ws["A1"] = f"SALDO BULANAN PER REKENING - {labels[0].upper()} S.D. {labels[-1].upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Satu-satunya rincian per rekening di laporan tahunan ini (di luar ini murni "
                "ringkasan) - saldo awal & akhir tiap rekening, per bulan, sepanjang tahun.")
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    # kumpulkan semua identitas rekening unik, urut sesuai kemunculan pertama
    all_bases = []
    seen = set()
    for month_accs in months_accounts:
        for acc in month_accs:
            if acc["base_name"] not in seen:
                seen.add(acc["base_name"])
                all_bases.append(acc["base_name"])

    r = 4
    headers = ["Rekening", "Bulan", "Saldo Awal", "Saldo Akhir", "Perubahan"]
    hdr_row = r
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hdr_row, column=i, value=h)
    rc.style_header(ws, hdr_row, len(headers))
    r += 1

    for base in all_bases:
        for i, label in enumerate(labels):
            acc = next((a for a in months_accounts[i] if a["base_name"] == base), None)
            ws.cell(row=r, column=1, value=base)
            ws.cell(row=r, column=2, value=label)
            if acc is not None:
                ws.cell(row=r, column=3, value=acc["saldo_awal"])
                ws.cell(row=r, column=4, value=acc["saldo_akhir"])
                ws.cell(row=r, column=5, value=round(acc["saldo_akhir"] - acc["saldo_awal"], 2))
                for c in (3, 4, 5):
                    ws.cell(row=r, column=c).number_format = rc.NUMBER_FORMAT
            else:
                ws.cell(row=r, column=3, value="(rekening tidak ada di bulan ini)")
                ws.cell(row=r, column=3).font = Font(italic=True, color="6B7280")
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = rc.BORDER
            r += 1
        r += 1  # baris kosong pemisah antar rekening

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    for col in "CDE":
        ws.column_dimensions[col].width = 20
    ws.freeze_panes = "A5"
    return ws


# ---------------------------------------------------------------------------
# Ringkasan Eksekutif - halaman pembuka ala laporan tahunan UMKM restoran:
# KPI utama, posisi kas, info pegawai, dan pembacaan tren singkat. Sebagian
# besar rumus merujuk ke sheet Laba Rugi/Neraca Tahunan (bukan dihitung
# ulang di sini) supaya satu sumber angka yang sama dipakai di semua sheet.
# ---------------------------------------------------------------------------

def write_ringkasan_eksekutif(wb, months, income_ref, balance_ref, roster_summary, assets, business_name=None):
    name = "Ringkasan Eksekutif"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    labels = [m["label"] for m in months]
    tahun = months[0]["year"] if months[0]["year"] == months[-1]["year"] else f"{months[0]['year']}-{months[-1]['year']}"

    r = 1
    if business_name:
        ws.cell(row=r, column=1, value=business_name)
        ws.cell(row=r, column=1).font = Font(bold=True, size=18)
        r += 1
    ws.cell(row=r, column=1, value=f"LAPORAN KEUANGAN TAHUNAN {tahun}")
    ws.cell(row=r, column=1).font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value=f"Periode: {labels[0]} - {labels[-1]}")
    ws.cell(row=r, column=1).font = Font(italic=True, size=10, color="6B7280")
    r += 1
    ws.cell(row=r, column=1, value=f"Disusun otomatis dari {len(labels)} laporan rekonsiliasi bulanan")
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="6B7280")
    r += 2

    income_sheet = income_ref["sheet"]
    balance_sheet = balance_ref["sheet"]
    income_total_col = rc.col_letter(income_ref["total_col"])
    balance_total_col = rc.col_letter(balance_ref["total_col"])
    balance_col_b = rc.col_letter(2)

    def kpi_row(label, formula, bold=True, pct=False, note=None):
        nonlocal r
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=formula)
        cell.number_format = "0.0%" if pct else rc.NUMBER_FORMAT
        if bold:
            ws.cell(row=r, column=1).font = Font(bold=True)
            cell.font = Font(bold=True, size=12)
        if note:
            ws.cell(row=r, column=3, value=note)
            ws.cell(row=r, column=3).font = Font(italic=True, size=9, color="6B7280")
        r += 1
        return r - 1

    ws.cell(row=r, column=1, value="KINERJA KEUANGAN")
    ws.cell(row=r, column=1).font = rc.SECTION_FONT
    ws.cell(row=r, column=1).fill = rc.SECTION_FILL
    r += 1
    rev_row = kpi_row("Total Pendapatan", f"='{income_sheet}'!{income_total_col}{income_ref['total_rev']}")
    exp_row = kpi_row("Total Beban", f"='{income_sheet}'!{income_total_col}{income_ref['total_exp']}")
    net_row = kpi_row("LABA / RUGI BERSIH TAHUN INI", f"='{income_sheet}'!{income_total_col}{income_ref['net']}")
    kpi_row("Margin Laba Bersih", f"=B{net_row}/B{rev_row}", bold=False, pct=True,
            note="Laba bersih dibagi total pendapatan")
    r += 1

    ws.cell(row=r, column=1, value="POSISI KAS")
    ws.cell(row=r, column=1).font = rc.SECTION_FONT
    ws.cell(row=r, column=1).fill = rc.SECTION_FILL
    r += 1
    kas_awal_row = kpi_row("Kas Awal Tahun", f"='{balance_sheet}'!{balance_col_b}{balance_ref['saldo_awal']}", bold=False)
    kas_akhir_row = kpi_row("Kas Akhir Tahun", f"='{balance_sheet}'!{balance_total_col}{balance_ref['total_asset']}", bold=False)
    growth_row = kpi_row("Pertumbuhan Kas (Rp)", f"=B{kas_akhir_row}-B{kas_awal_row}", bold=False)
    kpi_row(
        "Pertumbuhan Kas (%)",
        f'=IF(B{kas_awal_row}=0,"(kas awal Rp0)",(B{kas_akhir_row}-B{kas_awal_row})/ABS(B{kas_awal_row}))',
        bold=False, pct=True,
    )
    r += 1

    ws.cell(row=r, column=1, value="PEGAWAI")
    ws.cell(row=r, column=1).font = rc.SECTION_FONT
    ws.cell(row=r, column=1).fill = rc.SECTION_FILL
    r += 1
    ws.cell(row=r, column=1, value="Jumlah Pegawai (tercatat gaji tahun ini)")
    ws.cell(row=r, column=2, value=roster_summary["n_pegawai"])
    r += 1
    if roster_summary["n_belum_dibayar_bulan_terakhir"] > 0:
        ws.cell(row=r, column=1,
                value=(f"{roster_summary['n_belum_dibayar_bulan_terakhir']} pegawai belum tercatat dibayar "
                       f"di {labels[-1]} - kemungkinan accrual, dibayar bulan berikutnya "
                       "(lihat sheet Roster Gaji)."))
        ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="B45309")
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 32
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="RINGKASAN TREN")
    ws.cell(row=r, column=1).font = rc.SECTION_FONT
    ws.cell(row=r, column=1).fill = rc.SECTION_FILL
    r += 1
    summaries = [
        q.compute_month_summary(m["all_txns"], q.depreciation_for_month_idx(assets, i))
        for i, m in enumerate(months)
    ]
    revenues = [s[0] for s in summaries]
    expenses = [s[1] for s in summaries]
    nets = [s[3] for s in summaries]
    kas = [m["total_aset"] for m in months]
    all_txns_year = [t for m in months for t in m["all_txns"]]
    expense_cats = [c for c in rc.INCOME_CATEGORIES_EXPENSE if c.strip().lower() != q.ASSET_CATEGORY_TEXT]
    expense_cats = list(expense_cats) + ["Marketing & RnD", "Gaji Pegawai", "Biaya Admin Bank"]
    expense_vals = []
    for cat in rc.INCOME_CATEGORIES_EXPENSE:
        if cat.strip().lower() == q.ASSET_CATEGORY_TEXT:
            continue
        expense_vals.append(abs(q.sum_category(all_txns_year, cat)))
    expense_vals.append(abs(q.sum_category_multi(all_txns_year, rc.MARKETING_RND_CATEGORY_TEXTS)))
    expense_vals.append(abs(q.sum_category_prefix(all_txns_year, "gaji")))
    expense_vals.append(abs(q.sum_category_multi(all_txns_year, rc.BANK_FEE_CATEGORY_TEXTS)))
    if assets:
        total_depresiasi = sum(q.depreciation_for_month_idx(assets, i) for i in range(len(months)))
        expense_cats.append("Beban Penyusutan (Aset Tetap)")
        expense_vals.append(round(total_depresiasi, 2))
    kalimat = q.build_narasi_tren(labels, revenues, expenses, nets, kas, expense_cats, expense_vals)
    for line in kalimat:
        ws.cell(row=r, column=1, value=f"- {line}")
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 28
        r += 1
    r += 1
    ws.cell(row=r, column=1,
            value="Detail bulanan: sheet Laba Rugi Tahunan, Neraca Tahunan, Arus Kas Tahunan, "
                  "Roster Gaji 12 Bulan, Saldo Bulanan per Rekening, dan Analisis & Tren.")
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="6B7280")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30
    return ws


# ---------------------------------------------------------------------------
# Orkestrasi utama
# ---------------------------------------------------------------------------

def run_annual_report(paths, output_path, business_name="Stoa Space", carry_forward_paths=None):
    """paths: list of 12 path file bulanan (urutan upload bebas, akan
    disortir kronologis). carry_forward_paths: opsional, list path laporan
    kuartalan/tahunan SEBELUMNYA (yang punya sheet 'Buku Aset Tetap') untuk
    menyambung penyusutan aset yang dibeli sebelum periode laporan ini.
    Return dict ringkasan. Output 8 sheet: Ringkasan Eksekutif, Laba Rugi
    Tahunan, Neraca Tahunan, Saldo Bulanan per Rekening, Arus Kas Tahunan,
    Roster Gaji 12 Bulan, Analisis & Tren, Buku Aset Tetap - tidak ada
    sheet rekening/transaksi mentah selain saldo per rekening."""
    months = [load_month(p) for p in paths]
    accounts_raw = [load_month_accounts(p) for p in paths]
    # urutkan months DAN months_accounts BERSAMAAN sejak awal (kunci sort
    # yang sama, dipasangkan sebelum diurutkan) - sebelumnya months di-
    # overwrite jadi terurut lebih dulu lalu di-zip dengan months_accounts
    # yang masih ikut urutan upload, jadi bulan yang sudah terurut malah
    # kepasang dengan data rekening bulan LAIN kalau file diupload tidak
    # berurutan kronologis (mis. dari Desember mundur ke Januari)
    combined = sorted(zip(months, accounts_raw), key=lambda pair: (pair[0]["year"], pair[0]["month"]))
    months = [c[0] for c in combined]
    months_accounts = [c[1] for c in combined]
    months = validate_consecutive_12(months)  # sudah terurut - ini tinggal validasi berurutan/tidaknya

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    carried = []
    carried_hutang = []
    previous_closing = None
    for p in (carry_forward_paths or []):
        carried.extend(q.load_asset_ledger(p))
        carried_hutang.extend(q.load_hutang_ledger(p))
        if previous_closing is None:
            previous_closing = q.load_previous_period_closing(p)
    scanned = q.scan_assets_from_months(months)
    all_assets = q.merge_asset_lists(carried, scanned)
    rel_assets = q.assets_with_relative_idx(all_assets, months)
    all_hutang = q.merge_hutang_lists(carried_hutang)

    income_ref = q.write_quarterly_income_statement(out_wb, months, rel_assets, period_word="Tahunan")
    balance_ref = q.write_quarterly_balance_sheet(
        out_wb, months, income_ref, rel_assets, period_word="Tahunan", previous_closing=previous_closing
    )
    q.write_quarterly_cash_flow(out_wb, months, income_ref, balance_ref, period_word="Tahunan")
    write_saldo_bulanan_per_rekening(out_wb, months, months_accounts)
    roster_summary = q.write_roster_gaji(out_wb, months, period_word="tahun")
    q.write_analysis_sheet(out_wb, months, rel_assets, period_word="Tahunan")
    write_ringkasan_eksekutif(out_wb, months, income_ref, balance_ref, roster_summary, rel_assets, business_name)
    q.write_buku_aset_tetap(out_wb, all_assets, months)
    q.write_buku_hutang(out_wb, all_hutang)

    order = ["Ringkasan Eksekutif", "Laba Rugi Tahunan", "Neraca Tahunan",
             "Saldo Bulanan per Rekening", "Arus Kas Tahunan",
             "Roster Gaji 12 Bulan", "Analisis & Tren", "Buku Aset Tetap", "Buku Hutang"]
    out_wb._sheets = [out_wb[s] for s in order]

    out_wb.save(output_path)

    return {
        "periode": [m["label"] for m in months],
        "n_pegawai": roster_summary["n_pegawai"],
        "n_belum_dibayar_bulan_terakhir": roster_summary["n_belum_dibayar_bulan_terakhir"],
        "n_aset_tetap": len(all_assets),
        "n_aset_carry_forward": len(carried),
        "n_hutang": len(all_hutang),
    }


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 14:
        print("Usage: python annual.py bulan1.xlsx ... bulan12.xlsx output.xlsx "
              "[--carry-forward laporan_lama1.xlsx laporan_lama2.xlsx ...]")
        sys.exit(1)
    paths = sys.argv[1:13]
    out = sys.argv[13]
    carry_forward = []
    if "--carry-forward" in sys.argv:
        idx = sys.argv.index("--carry-forward")
        carry_forward = sys.argv[idx + 1:]
    try:
        s = run_annual_report(paths, out, carry_forward_paths=carry_forward)
        print(s)
    except AnnualInputError as e:
        print("Error:", e)
        sys.exit(1)
