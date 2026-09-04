"""
bot.py
Bot Telegram Rekonsiliasi Stoa Space (bot baru, terpisah dari stoabot &
bank-statement-bot).

Fungsi UTAMA (selalu jalan tiap ada file masuk): rekonsiliasi antar rekening
saja - pencocokan transfer, deteksi transfer terpecah/tergabung, dan
indikasi minus/selisih kas yang perlu verifikasi manual.

Fungsi laporan keuangan (Laba Rugi / Neraca / Arus Kas) DIBUAT OPSIONAL,
hanya keluar kalau ditrigger:
1. Kirim file dengan caption mengandung kata "laporan" (mis. "laporan
   lengkap"), ATAU
2. Kirim file dulu (dapat hasil recon-only), lalu kirim /laporan untuk
   generate ulang file yang sama + 3 sheet laporan keuangan.

Fungsi TAMBAHAN: laporan keuangan KUARTALAN dan TAHUNAN. Trigger /kuartal
atau /tahunan dulu, baru upload file-file hasil rekonsiliasi bulanan yang
SUDAH completed (bulan harus berurutan) DITAMBAH file laporan {kuartalan/
tahunan} SEBELUMNYA (untuk kontinuitas aset & data) - total 4 file untuk
/kuartal (3 bulanan + 1 laporan lama), 13 file untuk /tahunan (12 bulanan
+ 1 laporan lama). Bot otomatis membedakan mana file bulanan dan mana
laporan lama lewat sheet "Buku Aset Tetap" (urutan upload bebas). Kalau
ini laporan pertama kalinya (belum ada laporan sebelumnya), upload file
bulanan saja lalu kirim /selesai. Outputnya laporan ringkasan/kesimpulan
(bukan rekonsiliasi baru - itu sudah beres di masing-masing file bulanan).

Kontinuitas yang dijaga lewat file laporan sebelumnya itu ada DUA: (1)
penyusutan Aset Tetap yang dibeli sebelum periode laporan ini (sheet
"Buku Aset Tetap"), dan (2) Saldo Awal bulan pertama laporan ini
dibandingkan dengan Saldo Akhir bulan terakhir laporan sebelumnya (baris
"Cek Kontinuitas" di sheet Neraca).

Environment variable yang dibutuhkan:
- BOT_TOKEN : token bot dari BotFather
"""

import logging
import os
import shutil
import tempfile
import time

import openpyxl
from telegram import Update
from telegram.constants import ParseMode
from telegram.ext import Application, CommandHandler, MessageHandler, ContextTypes, filters
from telegram.error import Conflict, NetworkError

from reconcile import run_reconciliation
import reconcile as rc
from quarterly import run_quarterly_report, QuarterlyInputError, add_roster_to_monthly_report, check_continuity_between_months
from kasir_audit import run_kasir_audit, parse_pos_sales, KasirAuditError, _looks_like_account_sheet
from annual import run_annual_report, AnnualInputError
import shared_rules

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.environ.get("BOT_TOKEN")

# cache file terakhir yang diupload tiap user, supaya /laporan tidak perlu
# upload ulang. Disimpan di disk (bukan cuma di memori) karena worker bisa
# restart; dibersihkan otomatis kalau lebih tua dari CACHE_TTL.
CACHE_DIR = os.path.join(tempfile.gettempdir(), "reconbot_cache")
CACHE_TTL = 6 * 3600  # 6 jam
os.makedirs(CACHE_DIR, exist_ok=True)

# staging file buat mode /kuartal & /tahunan (N file per user, ditumpuk
# sampai lengkap). Satu direktori generik dipakai untuk kedua mode,
# dipisah per-user sekaligus per-mode di path-nya.
MULTI_DIR = os.path.join(tempfile.gettempdir(), "reconbot_multi")
os.makedirs(MULTI_DIR, exist_ok=True)

# konfigurasi tiap mode multi-file: jumlah file BULANAN yang dibutuhkan,
# jumlah file carry-forward (laporan sebelumnya) yang WAJIB disertakan
# untuk menjaga kontinuitas aset & data, fungsi generator laporan,
# exception khusus mode itu, dan teks-teks tampilan
MULTI_MODE_CONFIG = {
    "kuartal": {
        "n_months": 3,
        "n_carry_forward": 1,
        "run_fn": run_quarterly_report,
        "error_cls": QuarterlyInputError,
        "label": "kuartalan",
        "output_prefix": "Laporan Kuartalan",
    },
    "tahunan": {
        "n_months": 12,
        "n_carry_forward": 1,
        "run_fn": run_annual_report,
        "error_cls": AnnualInputError,
        "label": "tahunan",
        "output_prefix": "Laporan Tahunan",
    },
}

LAPORAN_TRIGGER_WORDS = ["laporan", "lengkap", "statement", "financial"]

WELCOME = (
    "Halo! Kirim satu file *.xlsx* gabungan berisi sheet per rekening "
    "(BRI, BCA, Jago, Kasir, dst - format kolom: Tanggal, Keterangan, "
    "Kategori, Debit, Kredit, Saldo Kumulatif, Subjek, Objek, Keterangan "
    "Tambahan).\n\n"
    "Default: bot cuma melakukan *rekonsiliasi antar rekening* - "
    "mencocokkan transfer, deteksi transfer terpecah/tergabung, dan "
    "menandai indikasi minus/selisih kas.\n\n"
    "Kalau butuh Laporan Laba Rugi, Neraca, dan Arus Kas juga (rumus "
    "Excel beralamat absolut), ada 2 cara:\n"
    "1. Tambahkan kata *laporan* di caption waktu upload file, atau\n"
    "2. Kirim /laporan setelah upload (pakai file yang sama, tanpa upload ulang)\n\n"
    "Butuh laporan KUARTALAN (3 bulan + 1 laporan kuartal sebelumnya)? Kirim /kuartal.\n"
    "Butuh laporan TAHUNAN (12 bulan + 1 laporan tahunan sebelumnya)? Kirim /tahunan."
)


def _cache_path(user_id):
    return os.path.join(CACHE_DIR, f"{user_id}_last_upload.xlsx")


def _cleanup_cache():
    now = time.time()
    for fname in os.listdir(CACHE_DIR):
        fpath = os.path.join(CACHE_DIR, fname)
        if os.path.isfile(fpath) and now - os.path.getmtime(fpath) > CACHE_TTL:
            os.remove(fpath)


def _multi_user_dir(user_id, mode):
    return os.path.join(MULTI_DIR, mode, str(user_id))


def _is_carry_forward_file(path):
    """Deteksi apakah file yang diupload adalah laporan kuartalan/tahunan
    LAMA (punya sheet 'Buku Aset Tetap' dan/atau 'Buku Hutang') - kalau
    iya, dipakai untuk menyambung penyusutan aset/catatan hutang, bukan
    dihitung sebagai salah satu file bulanan yang wajib."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        return "Buku Aset Tetap" in wb.sheetnames or "Buku Hutang" in wb.sheetnames
    except Exception:
        return False


CMD_LIST_TEXT = """*Daftar Perintah Bot*

*Rekonsiliasi bulanan (dasar)*
/start — pesan pembuka & cara pakai
Upload file .xlsx — proses rekonsiliasi bulanan langsung (tanpa command). Laporan keuangan lengkap otomatis ikut HANYA kalau tidak ada isu (semua transfer matched, Neraca balanced, tidak ada Kategori Baru).
/laporan — proses ulang file terakhir yang diupload, PAKSA sertakan laporan keuangan meski masih ada isu

*Laporan kuartalan & tahunan*
/kuartal — mulai sesi: upload 3 file bulanan berurutan + (opsional) 1 file carry-forward dari kuartal sebelumnya
/tahunan — mulai sesi: upload 12 file bulanan berurutan + (opsional) 1 file carry-forward dari tahun sebelumnya
/selesai — tutup sesi /kuartal, /tahunan, atau /auditkasir yang sedang berjalan dan mulai proses
/batal — batalkan sesi /kuartal, /tahunan, /kontinuitas, atau /auditkasir yang sedang berjalan

*Audit tambahan*
/kontinuitas — bandingkan Saldo Akhir file bulan lalu dengan Saldo Awal file bulan ini (deteksi selisih di batas antar bulan). Upload 2 file berurutan setelah command ini.
/auditkasir — audit silang mesin kasir (POS) vs rekap keuangan, termasuk asumsi settlement QRIS/kartu H+1. Upload file Detail Penjualan + Rekap (urutan bebas, jenis dideteksi otomatis), lalu /selesai.

*Kelola kategori & alias pegawai (butuh Postgres tersambung)*
/tambahkategori kata1, kata2 => Kategori Tujuan — tambah aturan kategori baru
/tambahalias alias => Nama Lengkap — tambah alias pegawai baru
/lihataturan — lihat semua aturan kategori yang aktif
/lihatalias — lihat semua alias pegawai yang aktif
/hapusaturan <nomor> — hapus satu aturan kategori (nomor sesuai /lihataturan)

/cmd — tampilkan daftar ini lagi"""


async def cmd_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(CMD_LIST_TEXT, parse_mode=ParseMode.MARKDOWN)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(WELCOME, parse_mode=ParseMode.MARKDOWN)


def build_caption(summary):
    lines = [
        "Rekonsiliasi selesai.",
        "",
        f"Transfer cocok: {summary['n_transfer_high']} high, "
        f"{summary['n_transfer_medium']} medium",
        f"Transfer terpecah/tergabung: {summary['n_transfer_split_merge']}",
        f"Transfer belum ketemu pasangan: {summary['n_transfer_unmatched']} "
        "(lihat sheet Rekonsiliasi, bagian 1)",
        f"Indikasi minus/selisih perlu verifikasi: {summary['n_minus_flags']}",
    ]
    if summary.get("n_new_category", 0) > 0:
        lines.append(
            f"Transaksi kategori tidak dikenal: {summary['n_new_category']} "
            "(lihat sheet Rekonsiliasi, bagian 4 - perlu diaudit)"
        )
    if summary.get("n_gaji_telat", 0) > 0:
        lines.append(
            f"Gaji telat/susulan (bukan untuk bulan ini): {summary['n_gaji_telat']} "
            "(lihat sheet 'Roster Gaji Bulan Ini')"
        )
    if summary["with_statements"]:
        lines += [
            "",
            "Laporan Laba Rugi / Neraca / Arus Kas ikut dibuat. Baris "
            "'CEK KESEIMBANGAN' di Neraca menandai kalau masih ada selisih "
            "yang harus ditelusuri manual.",
        ]
    elif summary.get("no_issues"):
        lines += [
            "",
            "Belum termasuk Laporan Laba Rugi/Neraca/Arus Kas. Kirim "
            "/laporan kalau butuh itu juga.",
        ]
    else:
        lines += [
            "",
            "Laporan Laba Rugi/Neraca/Arus Kas BELUM dibuat - masih ada isu "
            "di atas yang perlu dibereskan dulu (transfer belum matched/"
            "indikasi minus/kategori tidak dikenal). Beresi lalu upload "
            "ulang, atau kirim /laporan kalau tetap mau lihat laporan "
            "keuangan meski masih ada isu.",
        ]
    return "\n".join(lines)


def build_output_filename(summary, with_statements):
    period = summary.get("period_label") or "Periode Tidak Diketahui"
    if summary.get("no_issues"):
        base = f"Reconciliation Completed - {period}"
    else:
        base = f"Reconciliation - {period} (perlu verifikasi manual)"
    suffix = " + Laporan Keuangan" if with_statements else ""
    return f"{base}{suffix}.xlsx"


async def _process_and_reply(update, input_path, with_statements):
    status_msg = await update.message.reply_text("Memproses, tunggu sebentar...")
    with tempfile.TemporaryDirectory() as tmp:
        # nama file internal sementara - nama final ditentukan setelah tahu
        # hasil analisis (lihat build_output_filename)
        raw_output_path = os.path.join(tmp, "output.xlsx")
        try:
            summary = run_reconciliation(input_path, raw_output_path, with_statements=with_statements)
        except Exception as e:
            logger.exception("Gagal memproses file")
            await status_msg.edit_text(
                f"Gagal memproses file: {e}\n\n"
                "Cek apakah nama kolom dan urutan sheet sesuai format baku."
            )
            return
        # Roster Gaji Bulan Ini (audit telat/tepat waktu) - ikut aturan
        # yang sama dengan laporan keuangan (with_statements): otomatis
        # HANYA kalau rekonsiliasi bersih, atau tetap disertakan kalau
        # user memaksa lewat /laporan. Kegagalan di sini TIDAK
        # menggagalkan pengiriman file rekonsiliasi utama.
        if summary.get("with_statements"):
            try:
                roster_summary = add_roster_to_monthly_report(raw_output_path)
                summary["n_gaji_telat"] = roster_summary.get("n_telat", 0)
            except Exception:
                logger.exception("Gagal menambahkan Roster Gaji Bulan Ini (dilewati, file utama tetap dikirim)")
        final_filename = build_output_filename(summary, summary["with_statements"])
        await status_msg.delete()
        with open(raw_output_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=final_filename,
                caption=build_caption(summary),
            )


async def kontinuitas_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Bandingkan Saldo Akhir tiap rekening di file BULAN LALU dengan
    Saldo Awal rekening yang sama di file BULAN INI - deteksi selisih
    tepat di batas antar bulan (mis. Saldo Awal diketik manual, tidak
    nyambung dari Saldo Akhir bulan sebelumnya)."""
    context.user_data["kontinuitas_mode"] = True
    context.user_data["kontinuitas_files"] = []
    await update.message.reply_text(
        "Mode cek kontinuitas aktif. Upload 2 file secara berurutan:\n"
        "1. File rekonsiliasi bulan LALU (yang lebih awal)\n"
        "2. File rekonsiliasi bulan INI (yang lebih baru, urutan setelahnya)\n\n"
        "Kirim /batal kalau mau keluar dari mode ini."
    )


async def handle_kontinuitas_document(update: Update, context: ContextTypes.DEFAULT_TYPE, doc):
    files = context.user_data.setdefault("kontinuitas_files", [])
    user_id = update.effective_user.id
    tmp_dir = os.path.join(tempfile.gettempdir(), f"kontinuitas_{user_id}")
    os.makedirs(tmp_dir, exist_ok=True)
    dest = os.path.join(tmp_dir, f"file_{len(files) + 1}.xlsx")
    tg_file = await doc.get_file()
    await tg_file.download_to_drive(dest)
    files.append(dest)

    if len(files) < 2:
        await update.message.reply_text(f"File {len(files)}/2 diterima. Upload 1 file lagi (bulan yang lebih baru).")
        return

    status_msg = await update.message.reply_text("Membandingkan kontinuitas, tunggu sebentar...")
    try:
        result = check_continuity_between_months(files[0], files[1])
    except QuarterlyInputError as e:
        await status_msg.edit_text(f"Gagal: {e}")
        shutil.rmtree(tmp_dir, ignore_errors=True)
        context.user_data["kontinuitas_mode"] = False
        context.user_data["kontinuitas_files"] = []
        return
    except Exception as e:
        logger.exception("Gagal cek kontinuitas")
        await status_msg.edit_text(f"Gagal memproses: {e}")
        shutil.rmtree(tmp_dir, ignore_errors=True)
        context.user_data["kontinuitas_mode"] = False
        context.user_data["kontinuitas_files"] = []
        return
    shutil.rmtree(tmp_dir, ignore_errors=True)
    context.user_data["kontinuitas_mode"] = False
    context.user_data["kontinuitas_files"] = []

    lines = ["Hasil cek kontinuitas:\n"]
    for row in result["rows"]:
        akhir = f"Rp{row['saldo_akhir_lalu']:,.0f}".replace(",", ".") if row["saldo_akhir_lalu"] is not None else "-"
        awal = f"Rp{row['saldo_awal_ini']:,.0f}".replace(",", ".") if row["saldo_awal_ini"] is not None else "-"
        selisih = f"Rp{row['selisih']:,.0f}".replace(",", ".") if row["selisih"] is not None else "-"
        mark = "✅" if row["status"] == "OK" else "⚠️"
        lines.append(f"{mark} {row['rekening']}: {akhir} → {awal} (selisih {selisih}) [{row['status']}]")
    lines.append("")
    lines.append(f"Total selisih: Rp{result['total_selisih']:,.0f}".replace(",", "."))
    if result["n_bermasalah"] == 0:
        lines.append("SEMUA REKENING NYAMBUNG - tidak ada selisih yang perlu ditelusuri.")
    else:
        lines.append(f"{result['n_bermasalah']} rekening ada selisih/tidak match - perlu ditelusuri manual.")
    await status_msg.edit_text("\n".join(lines))


async def auditkasir_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Fitur MANUAL (tidak otomatis, harus di-trigger) - audit silang
    total penjualan mesin kasir (POS) vs rekap keuangan (bank/kas) per
    metode bayar. Terima file 'Detail Penjualan' (export POS) dan file
    Rekap (hasil rekonsiliasi bulanan biasa) dalam urutan BEBAS - jenis
    file dideteksi otomatis dari strukturnya, bukan dari nama file."""
    context.user_data["auditkasir_mode"] = True
    context.user_data["auditkasir_pos_files"] = []
    context.user_data["auditkasir_rekap_files"] = []
    await update.message.reply_text(
        "Mode audit kasir aktif. Upload file 'Detail Penjualan' (export mesin kasir) DAN file "
        "Rekap (hasil rekonsiliasi bulanan) - urutan bebas, boleh diselang-seling, jenis file "
        "dideteksi otomatis. Minimal 1 file masing-masing jenis.\n\n"
        "Kirim /selesai kalau sudah semua file terupload, atau /batal untuk keluar dari mode ini."
    )


async def handle_auditkasir_document(update: Update, context: ContextTypes.DEFAULT_TYPE, doc):
    user_id = update.effective_user.id
    tmp_dir = os.path.join(tempfile.gettempdir(), f"auditkasir_{user_id}")
    os.makedirs(tmp_dir, exist_ok=True)
    pos_files = context.user_data.setdefault("auditkasir_pos_files", [])
    rekap_files = context.user_data.setdefault("auditkasir_rekap_files", [])
    dest = os.path.join(tmp_dir, f"file_{len(pos_files) + len(rekap_files) + 1}.xlsx")
    tg_file = await doc.get_file()
    await tg_file.download_to_drive(dest)

    # deteksi jenis file dari strukturnya, bukan nama file
    try:
        parse_pos_sales(dest)
        pos_files.append(dest)
        jenis = "Detail Penjualan (POS)"
    except KasirAuditError:
        import openpyxl
        wb = openpyxl.load_workbook(dest)
        if any(_looks_like_account_sheet(wb[s]) for s in wb.sheetnames):
            rekap_files.append(dest)
            jenis = "Rekap keuangan"
        else:
            os.remove(dest)
            await update.message.reply_text(
                f"File '{doc.file_name}' tidak dikenali sebagai Detail Penjualan POS maupun Rekap "
                "keuangan - dilewati. Pastikan format filenya sesuai export POS atau hasil rekonsiliasi bulanan."
            )
            return
    await update.message.reply_text(
        f"Diterima sebagai {jenis} ({len(pos_files)} POS, {len(rekap_files)} Rekap terkumpul). "
        "Lanjut upload atau kirim /selesai."
    )


async def _process_auditkasir(update: Update, context: ContextTypes.DEFAULT_TYPE):
    pos_files = context.user_data.get("auditkasir_pos_files", [])
    rekap_files = context.user_data.get("auditkasir_rekap_files", [])
    user_id = update.effective_user.id
    tmp_dir = os.path.join(tempfile.gettempdir(), f"auditkasir_{user_id}")
    status_msg = await update.message.reply_text("Menjalankan audit kasir, tunggu sebentar...")
    output_path = os.path.join(tmp_dir, "Audit_Kasir.xlsx")
    try:
        summary = run_kasir_audit(pos_files, rekap_files, output_path)
    except KasirAuditError as e:
        await status_msg.edit_text(f"Gagal: {e}")
        shutil.rmtree(tmp_dir, ignore_errors=True)
        context.user_data["auditkasir_mode"] = False
        return
    except Exception as e:
        logger.exception("Gagal menjalankan audit kasir")
        await status_msg.edit_text(f"Gagal memproses: {e}")
        shutil.rmtree(tmp_dir, ignore_errors=True)
        context.user_data["auditkasir_mode"] = False
        return
    context.user_data["auditkasir_mode"] = False
    context.user_data["auditkasir_pos_files"] = []
    context.user_data["auditkasir_rekap_files"] = []
    caption = (
        f"Audit kasir selesai - {summary['n_bulan']} bulan, {summary['n_transaksi_pos']} transaksi POS.\n"
        f"Refund: {summary['n_refund']} transaksi, total Rp{summary['total_refund']:,.0f}.\n"
        f"Belum lunas (dikecualikan dari audit): {summary['n_belum_lunas']} transaksi.\n\n"
        "Cek sheet 'Ringkasan Audit Kasir' - baris merah/kuning perlu ditelusuri manual."
    ).replace(",", ".")
    await status_msg.delete()
    with open(output_path, "rb") as f:
        await update.message.reply_document(document=f, filename="Audit_Kasir.xlsx", caption=caption)
    shutil.rmtree(tmp_dir, ignore_errors=True)


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc or not doc.file_name.lower().endswith(".xlsx"):
        await update.message.reply_text("Kirim file .xlsx ya, format lain belum didukung.")
        return

    if context.user_data.get("auditkasir_mode"):
        await handle_auditkasir_document(update, context, doc)
        return

    if context.user_data.get("kontinuitas_mode"):
        await handle_kontinuitas_document(update, context, doc)
        return

    active_mode = context.user_data.get("multi_mode")
    if active_mode:
        await handle_multi_document(update, context, doc, active_mode)
        return

    _cleanup_cache()
    user_id = update.effective_user.id
    cache_path = _cache_path(user_id)

    tg_file = await doc.get_file()
    await tg_file.download_to_drive(cache_path)  # simpan buat trigger /laporan nanti

    # Laporan keuangan (Laba Rugi/Neraca/Arus Kas) OTOMATIS disertakan
    # HANYA kalau rekonsiliasi ini bersih (tidak ada isu) - kalau masih
    # ada transfer belum matched/indikasi minus/Neraca tidak balanced/
    # Kategori Baru, cuma sheet rekening + Rekonsiliasi yang dibuat.
    # Kirim /laporan kalau tetap mau lihat laporan keuangan meski masih
    # ada isu yang belum beres.
    await _process_and_reply(update, cache_path, with_statements=None)


async def laporan_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    cache_path = _cache_path(user_id)
    if not os.path.exists(cache_path):
        await update.message.reply_text(
            "Belum ada file yang diupload. Kirim file .xlsx dulu, baru /laporan."
        )
        return
    await _process_and_reply(update, cache_path, with_statements=True)


# ---------------------------------------------------------------------------
# Mode multi-file (/kuartal, /tahunan): trigger dulu, baru upload N file
# bulanan yang sudah completed (bulan harus berurutan)
# ---------------------------------------------------------------------------

async def _start_multi_mode(update: Update, context: ContextTypes.DEFAULT_TYPE, mode: str):
    cfg = MULTI_MODE_CONFIG[mode]
    n_total = cfg["n_months"] + cfg["n_carry_forward"]
    user_id = update.effective_user.id
    user_dir = _multi_user_dir(user_id, mode)
    shutil.rmtree(user_dir, ignore_errors=True)
    os.makedirs(user_dir, exist_ok=True)
    context.user_data["multi_mode"] = mode
    context.user_data["multi_files"] = []
    context.user_data["multi_carry_forward"] = []
    await update.message.reply_text(
        f"Mode laporan {cfg['label']} aktif. Upload total {n_total} file (urutan bebas, nanti "
        f"diurutkan/dipilah otomatis):\n"
        f"- {cfg['n_months']} file hasil rekonsiliasi bulanan yang sudah *completed* (bulan "
        "harus berurutan)\n"
        f"- {cfg['n_carry_forward']} file laporan {cfg['label']} SEBELUMNYA (untuk menjaga "
        "kontinuitas aset tetap & data antar periode) - bot otomatis mengenalinya lewat sheet "
        "'Buku Aset Tetap'\n\n"
        f"Kalau ini laporan {cfg['label']} PERTAMA kalinya (belum ada laporan sebelumnya), "
        f"upload {cfg['n_months']} file bulanan saja lalu kirim /selesai.\n\n"
        "Kirim /batal kalau mau keluar dari mode ini.",
        parse_mode=ParseMode.MARKDOWN,
    )


async def kuartal_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await _start_multi_mode(update, context, "kuartal")


async def tahunan_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await _start_multi_mode(update, context, "tahunan")


async def batal_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get("auditkasir_mode"):
        context.user_data["auditkasir_mode"] = False
        context.user_data["auditkasir_pos_files"] = []
        context.user_data["auditkasir_rekap_files"] = []
        await update.message.reply_text("Mode audit kasir dibatalkan.")
        return
    if context.user_data.get("kontinuitas_mode"):
        context.user_data["kontinuitas_mode"] = False
        context.user_data["kontinuitas_files"] = []
        await update.message.reply_text("Mode cek kontinuitas dibatalkan.")
        return
    mode = context.user_data.get("multi_mode")
    if not mode:
        await update.message.reply_text("Tidak ada proses yang bisa dibatalkan.")
        return
    user_id = update.effective_user.id
    shutil.rmtree(_multi_user_dir(user_id, mode), ignore_errors=True)
    context.user_data["multi_mode"] = None
    context.user_data["multi_files"] = []
    context.user_data["multi_carry_forward"] = []
    cfg = MULTI_MODE_CONFIG[mode]
    await update.message.reply_text(f"Mode laporan {cfg['label']} dibatalkan.")


async def selesai_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Buat kasus laporan PERTAMA kalinya - tidak ada laporan sebelumnya
    untuk di-carry-forward, jadi trigger manual dengan file bulanan yang
    sudah ada tanpa menunggu file carry-forward yang memang tidak akan
    pernah datang. Juga dipakai untuk menutup sesi /auditkasir (jumlah
    file variabel, tidak ada carry-forward tetap)."""
    if context.user_data.get("auditkasir_mode"):
        pos_files = context.user_data.get("auditkasir_pos_files", [])
        rekap_files = context.user_data.get("auditkasir_rekap_files", [])
        if not pos_files or not rekap_files:
            await update.message.reply_text(
                f"Baru ada {len(pos_files)} file POS dan {len(rekap_files)} file Rekap - butuh "
                "minimal 1 dari masing-masing jenis sebelum kirim /selesai."
            )
            return
        await _process_auditkasir(update, context)
        return
    mode = context.user_data.get("multi_mode")
    if not mode:
        await update.message.reply_text("Tidak ada proses laporan kuartalan/tahunan yang sedang berjalan.")
        return
    cfg = MULTI_MODE_CONFIG[mode]
    files = context.user_data.get("multi_files", [])
    if len(files) != cfg["n_months"]:
        await update.message.reply_text(
            f"Baru ada {len(files)}/{cfg['n_months']} file bulanan - upload dulu sampai lengkap "
            "sebelum kirim /selesai."
        )
        return
    await _process_multi_report(update, context, mode)


def _reset_multi_mode(context):
    context.user_data["multi_mode"] = None
    context.user_data["multi_files"] = []
    context.user_data["multi_carry_forward"] = []


async def handle_multi_document(update: Update, context: ContextTypes.DEFAULT_TYPE, doc, mode):
    cfg = MULTI_MODE_CONFIG[mode]
    n_months = cfg["n_months"]
    n_carry_forward = cfg["n_carry_forward"]
    user_id = update.effective_user.id
    user_dir = _multi_user_dir(user_id, mode)
    os.makedirs(user_dir, exist_ok=True)
    files = context.user_data.setdefault("multi_files", [])
    carry_forward = context.user_data.setdefault("multi_carry_forward", [])

    idx = len(files) + len(carry_forward) + 1
    dest_path = os.path.join(user_dir, f"upload_{idx}.xlsx")
    tg_file = await doc.get_file()
    await tg_file.download_to_drive(dest_path)

    if _is_carry_forward_file(dest_path):
        carry_forward.append(dest_path)
        await update.message.reply_text(
            f"File ini dikenali sebagai laporan {cfg['label']} sebelumnya (ada sheet 'Buku Aset "
            f"Tetap') - akan dipakai untuk menyambung kontinuitas aset & data.\n\n"
            f"Progres: {len(files)}/{n_months} bulan, {len(carry_forward)}/{n_carry_forward} "
            "file carry-forward."
        )
    else:
        files.append(dest_path)
        if len(files) < n_months:
            await update.message.reply_text(
                f"Diterima ({len(files)}/{n_months} bulan, {len(carry_forward)}/{n_carry_forward} "
                f"carry-forward). Upload {n_months - len(files)} file bulanan lagi."
            )
            return

    if len(files) >= n_months and len(carry_forward) >= n_carry_forward:
        await _process_multi_report(update, context, mode)
        return

    if len(files) >= n_months:
        await update.message.reply_text(
            f"{n_months} file bulanan sudah lengkap. Upload {n_carry_forward - len(carry_forward)} "
            f"file laporan {cfg['label']} sebelumnya lagi untuk kontinuitas, atau kirim /selesai "
            "kalau ini laporan pertama kali (belum ada laporan sebelumnya)."
        )


async def _process_multi_report(update: Update, context: ContextTypes.DEFAULT_TYPE, mode: str):
    cfg = MULTI_MODE_CONFIG[mode]
    n_months = cfg["n_months"]
    user_id = update.effective_user.id
    user_dir = _multi_user_dir(user_id, mode)
    files = context.user_data.get("multi_files", [])
    carry_forward = context.user_data.get("multi_carry_forward", [])

    status_msg = await update.message.reply_text(f"Memproses laporan {cfg['label']}...")
    with tempfile.TemporaryDirectory() as tmp:
        output_path = os.path.join(tmp, "laporan.xlsx")
        try:
            summary = cfg["run_fn"](files, output_path, carry_forward_paths=carry_forward)
        except cfg["error_cls"] as e:
            await status_msg.edit_text(f"Gagal: {e}\n\nKirim /{mode} lagi untuk coba ulang.")
            _reset_multi_mode(context)
            shutil.rmtree(user_dir, ignore_errors=True)
            return
        except Exception as e:
            logger.exception("Gagal memproses laporan %s", cfg["label"])
            await status_msg.edit_text(
                f"Gagal memproses: {e}\n\nCek apakah semua {n_months} file bulanan itu benar hasil "
                f"rekonsiliasi bulanan yang sudah completed. Kirim /{mode} lagi untuk coba ulang."
            )
            _reset_multi_mode(context)
            shutil.rmtree(user_dir, ignore_errors=True)
            return

        periode = summary["periode"]
        final_filename = f"{cfg['output_prefix']} - {periode[0]} s.d. {periode[-1]}.xlsx"
        caption_lines = [
            f"Laporan {cfg['label']} {periode[0]} - {periode[-1]} selesai.",
            "",
            f"Jumlah pegawai di roster gaji: {summary['n_pegawai']}",
        ]
        if not carry_forward:
            caption_lines.append(
                "Tidak ada laporan sebelumnya yang di-carry-forward - dianggap laporan pertama "
                "kalinya, kontinuitas cuma dicek antar bulan di dalam laporan ini saja."
            )
        if summary["n_belum_dibayar_bulan_terakhir"] > 0:
            caption_lines.append(
                f"{summary['n_belum_dibayar_bulan_terakhir']} pegawai belum tercatat "
                f"dibayar di bulan terakhir ({periode[-1]}) - lihat kolom Catatan di "
                "sheet Roster Gaji, kemungkinan dibebankan sebagai accrual bulan berikutnya."
            )
        if summary.get("n_aset_tetap", 0) > 0:
            carry_note = (
                f" ({summary['n_aset_carry_forward']} di antaranya disambung dari laporan lama.)"
                if summary.get("n_aset_carry_forward", 0) > 0 else ""
            )
            caption_lines.append(
                f"Aset tetap tercatat: {summary['n_aset_tetap']}{carry_note} - lihat sheet "
                "'Buku Aset Tetap' untuk detail jadwal penyusutan. Simpan file ini kalau nanti "
                "mau membuat laporan berikutnya, supaya kontinuitasnya bisa disambung lagi."
            )
        if summary.get("n_hutang", 0) > 0:
            caption_lines.append(
                f"Catatan hutang tercatat: {summary['n_hutang']} - lihat sheet 'Buku Hutang' "
                "(diisi/update manual, ikut tersalin ke laporan berikutnya kalau file ini "
                "dipakai sebagai carry-forward)."
            )
        await status_msg.delete()
        with open(output_path, "rb") as f:
            await update.message.reply_document(
                document=f, filename=final_filename, caption="\n".join(caption_lines)
            )

    _reset_multi_mode(context)
    shutil.rmtree(user_dir, ignore_errors=True)


_VALID_CATEGORIES = set(
    rc.INCOME_CATEGORIES_EXPENSE + rc.INCOME_CATEGORIES_REVENUE +
    rc.MARKETING_RND_CATEGORY_TEXTS + rc.BANK_FEE_CATEGORY_TEXTS +
    rc.OTHER_CATEGORIES + rc.TRANSFER_CATEGORY_TEXTS + ["Modal & Setoran Pemilik"]
)


async def tambahkategori_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Format: /tambahkategori kata1, kata2 => Kategori Tujuan
    Opsional tambahkan " | sheet:nama" di akhir untuk batasi ke sheet
    tertentu (mis. cuma berlaku di rekening Jago)."""
    text = " ".join(context.args)
    if "=>" not in text:
        await update.message.reply_text(
            "Format: /tambahkategori kata1, kata2 => Kategori Tujuan\n"
            "Contoh: /tambahkategori dana bos, dana bantuan => Penjualan\n"
            "Opsional batasi ke satu rekening: tambahkan | sheet:jago di akhir\n\n"
            f"Kategori yang dikenal: {', '.join(sorted(_VALID_CATEGORIES))}"
        )
        return
    left, category = text.split("=>", 1)
    sheet_contains = None
    if "|" in category:
        category, modifier = category.split("|", 1)
        modifier = modifier.strip()
        if modifier.lower().startswith("sheet:"):
            sheet_contains = modifier.split(":", 1)[1].strip().lower()
    category = category.strip()
    keywords = [k.strip().lower() for k in left.split(",") if k.strip()]
    if not keywords:
        await update.message.reply_text("Tidak ada kata kunci yang diberikan.")
        return
    try:
        rule = shared_rules.add_category_rule(keywords, category, _VALID_CATEGORIES, sheet_contains)
    except ValueError as e:
        await update.message.reply_text(f"Gagal: {e}")
        return
    except Exception as e:
        logger.error("Gagal tulis ke shared_rules", exc_info=e)
        await update.message.reply_text(f"Gagal terhubung ke database bersama: {shared_rules.diagnose_connection_error(e)}")
        return
    sheet_note = f" (khusus rekening mengandung '{sheet_contains}')" if sheet_contains else ""
    await update.message.reply_text(
        f"Tersimpan: kata kunci {rule['any']} -> kategori '{category}'{sheet_note}.\n"
        "Berlaku untuk reconbot, bank-statement-bot, dan bot lain yang baca shared_rules yang sama."
    )


async def tambahalias_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Format: /tambahalias nama_pendek => Nama Lengkap"""
    text = " ".join(context.args)
    if "=>" not in text:
        await update.message.reply_text(
            "Format: /tambahalias nama_pendek => Nama Lengkap\n"
            "Contoh: /tambahalias budi => Budi Santoso"
        )
        return
    short_name, full_name = text.split("=>", 1)
    short_name, full_name = short_name.strip(), full_name.strip()
    if not short_name or not full_name:
        await update.message.reply_text("Nama pendek dan nama lengkap harus diisi.")
        return
    try:
        shared_rules.add_employee_alias(short_name, full_name)
    except Exception as e:
        logger.error("Gagal tulis ke shared_rules", exc_info=e)
        await update.message.reply_text(f"Gagal terhubung ke database bersama: {shared_rules.diagnose_connection_error(e)}")
        return
    await update.message.reply_text(f"Tersimpan: '{short_name}' -> '{full_name}'.")


async def lihataturan_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        rules = shared_rules.list_category_rules()
    except Exception as e:
        logger.error("Gagal baca shared_rules", exc_info=e)
        await update.message.reply_text(f"Gagal terhubung ke database bersama: {shared_rules.diagnose_connection_error(e)}")
        return
    if not rules:
        await update.message.reply_text("Belum ada aturan kategori tersimpan.")
        return
    lines = ["Aturan kategori saat ini (urutan = prioritas, yang duluan menang):"]
    for i, r in enumerate(rules):
        kw = r.get("any") or r.get("all") or []
        joiner = " ATAU " if "any" in r else " DAN "
        sheet = f" [khusus: {r['sheet_contains']}]" if r.get("sheet_contains") else ""
        lines.append(f"{i}. {joiner.join(kw)} -> {r.get('category')}{sheet}")
    text = "\n".join(lines)
    for i in range(0, len(text), 3500):
        await update.message.reply_text(text[i:i + 3500])


async def lihatalias_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        aliases = shared_rules.list_employee_aliases()
    except Exception as e:
        logger.error("Gagal baca shared_rules", exc_info=e)
        await update.message.reply_text(f"Gagal terhubung ke database bersama: {shared_rules.diagnose_connection_error(e)}")
        return
    if not aliases:
        await update.message.reply_text("Belum ada alias pegawai tersimpan.")
        return
    lines = ["Alias pegawai saat ini:"] + [f"'{k}' -> '{v}'" for k, v in aliases.items()]
    text = "\n".join(lines)
    for i in range(0, len(text), 3500):
        await update.message.reply_text(text[i:i + 3500])


async def hapusaturan_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Format: /hapusaturan <index> - index dilihat dari /lihataturan."""
    if not context.args or not context.args[0].isdigit():
        await update.message.reply_text("Format: /hapusaturan <index> (lihat index lewat /lihataturan)")
        return
    idx = int(context.args[0])
    try:
        removed = shared_rules.remove_category_rule(idx)
    except IndexError as e:
        await update.message.reply_text(f"Gagal: {e}")
        return
    except Exception as e:
        logger.error("Gagal tulis ke shared_rules", exc_info=e)
        await update.message.reply_text(f"Gagal terhubung ke database bersama: {shared_rules.diagnose_connection_error(e)}")
        return
    await update.message.reply_text(f"Dihapus: {removed}")


async def error_handler(update, context: ContextTypes.DEFAULT_TYPE):
    """Handler error global - PTB akan spam traceback mentah ke log kalau
    tidak ada ini terdaftar. Conflict (dua instance bot rebutan getUpdates,
    biasanya sisa deploy lama yang belum mati) ditangani beda dari error
    lain: cukup dicatat singkat, PTB sendiri sudah auto-retry sampai
    instance lama berhenti - tidak perlu tindakan lebih lanjut di sini.
    Error lain dicatat lengkap supaya tetap bisa ditelusuri."""
    err = context.error
    if isinstance(err, Conflict):
        logger.warning(
            "409 Conflict: kemungkinan ada instance bot lain yang masih jalan "
            "(sisa deploy lama). PTB akan retry otomatis sampai instance lama mati."
        )
        return
    if isinstance(err, NetworkError):
        logger.warning("Network error sementara, PTB akan retry otomatis: %s", err)
        return
    logger.error("Unhandled exception saat proses update", exc_info=err)


def main():
    if not BOT_TOKEN:
        raise RuntimeError("Env var BOT_TOKEN belum di-set")
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("laporan", laporan_command))
    app.add_handler(CommandHandler("kuartal", kuartal_command))
    app.add_handler(CommandHandler("tahunan", tahunan_command))
    app.add_handler(CommandHandler("selesai", selesai_command))
    app.add_handler(CommandHandler("batal", batal_command))
    app.add_handler(CommandHandler("tambahkategori", tambahkategori_command))
    app.add_handler(CommandHandler("tambahalias", tambahalias_command))
    app.add_handler(CommandHandler("lihataturan", lihataturan_command))
    app.add_handler(CommandHandler("lihatalias", lihatalias_command))
    app.add_handler(CommandHandler("hapusaturan", hapusaturan_command))
    app.add_handler(CommandHandler("kontinuitas", kontinuitas_command))
    app.add_handler(CommandHandler("auditkasir", auditkasir_command))
    app.add_handler(CommandHandler("cmd", cmd_command))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_error_handler(error_handler)
    logger.info("Bot rekonsiliasi jalan...")
    # drop_pending_updates: buang antrean update lama saat start - supaya
    # kalau ada command yang "hilang" waktu jendela konflik 2 instance
    # (mis. saat redeploy), instance baru tidak balas telat/dobel ke pesan
    # basi, dan user tinggal kirim ulang command-nya
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
