"""
reconcile.py
Mesin rekonsiliasi antar rekening untuk Stoa Space.

Input : workbook xlsx berisi 1 sheet per rekening (BRI, BCA, Jago, Kasir, dst)
        dengan kolom baku: Tanggal | Keterangan Transaksi | Kategori Transaksi |
        Debit | Kredit | Saldo Kumulatif | Subjek Transaksi | Objek Transaksi |
        Keterangan Tambahan

Output: workbook baru dengan:
        - Setiap sheet rekening asli disalin apa adanya, plus kolom bantu
          "Nominal Bersih" (formula, bukan angka mati)
        - Sheet "Rekonsiliasi" berisi hasil pencocokan transfer antar rekening
          dan daftar minus/selisih yang perlu verifikasi
        - Sheet "Laporan Laba Rugi" (Income Statement)
        - Sheet "Neraca" (Balance Sheet)
        - Sheet "Laporan Arus Kas" (Cash Flow Statement)

Semua angka di sheet laporan dibuat dengan rumus Excel beralamat absolut
($Kolom$Baris), merujuk langsung ke sheet rekening. Tidak ada angka hasil
kalkulasi Python yang ditulis sebagai nilai mati kecuali memang tidak
mungkin direpresentasikan sebagai rumus (contoh: catatan naratif audit).
"""

import re
import shared_rules
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dataclasses import dataclass, field

# ---------------------------------------------------------------------------
# Konfigurasi
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="E5E7EB")
SECTION_FONT = Font(bold=True)
HIGH_FILL = PatternFill("solid", fgColor="C6EFCE")
MED_FILL = PatternFill("solid", fgColor="FFEB9C")
LOW_FILL = PatternFill("solid", fgColor="FFC7CE")
DATE_FORMAT = "d-mmm-yy"
NUMBER_FORMAT = "#,##0"
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Kategori yang menandakan perpindahan uang ANTAR rekening sendiri
# (bukan pendapatan/beban riil) -> tidak masuk Laba Rugi, harus saling
# menutup nol di Rekonsiliasi.
TRANSFER_KEYWORDS = shared_rules.get("transfer_keywords", [
    "pindah rekening internal",
    "pindang rekening internal",
    "transfer internal",
    "transfer lainnya",
    "transaksi internal",
])

# Kata kunci di KETERANGAN (bukan kategori) yang juga menandakan transfer
# antar rekening sendiri, buat jaga-jaga kalau kategorinya salah/tidak
# konsisten diisi di sumber data (mis. "Setoran Tunai" kadang tidak diberi
# kategori "Pindah/Transfer Internal" padahal itu kasir->bank sendiri).
# Tidak dipakai kalau kategorinya sudah eksplisit "Modal & Setoran Pemilik"
# (setoran modal dari luar, bukan pindah antar rekening sendiri).
DESC_TRANSFER_KEYWORDS = shared_rules.get("desc_transfer_keywords", [
    "setoran tunai",
    "setor tunai",
    "setoran via cdm",
    "pindah rekening",
    "transfer internal",
])

# Kategori setoran/penarikan modal pemilik -> masuk Neraca (ekuitas),
# bukan Laba Rugi.
CAPITAL_KEYWORDS = shared_rules.get("capital_keywords", [
    "modal & setoran pemilik",
    "modal dan setoran pemilik",
    "laba ditahan bulanan",
])

# "Transfer Masuk" sendirian terlalu umum untuk langsung dianggap Modal
# (transfer masuk dari pelanggan/pihak luar seharusnya Penjualan, bukan
# Modal) - jadi HANYA dianggap setara Modal & Setoran Pemilik kalau
# keterangannya eksplisit bilang "dari rekening sendiri" (uang milik
# owner sendiri yang dipindah antar rekening, mis. pencairan investasi
# pribadi yang disetor ke rekening bisnis), sesuai kasus nyata yang
# ditemukan: "BI-Fast Transfer Masuk ... (dari rekening sendiri)".
CAPITAL_SELF_TRANSFER_KEYWORDS = shared_rules.get("capital_self_transfer_keywords", ["dari rekening sendiri", "rekening sendiri"])

# Aturan override kategori berdasarkan kata kunci di Keterangan/Keterangan
# Tambahan/Objek, ditegaskan langsung oleh user berdasarkan pengalaman
# nyata bisnisnya - dicek SEBELUM aturan-aturan lain (is_transfer, dst),
# dan MENGGANTIKAN Kategori asli untuk semua perhitungan (lihat
# Txn.effective_kategori). Urutan penting: yang pertama cocok menang.
# "any": salah satu kata kunci cukup. "all": semua kata kunci harus ada
# (tidak harus berdekatan). "sheet_contains": opsional, cuma berlaku
# kalau nama sheet mengandung teks ini (case-insensitive).
_DEFAULT_CATEGORY_OVERRIDE_RULES = [
    {"all": ["briva", "tokopedia"], "amount_min": 900000, "amount_max": 1100000,
     "category": "Belanja Operasional", "sheet_contains": None},
    {"any": ["tokopedia"], "category": "Belanja Bahan", "sheet_contains": None},
    {"all": ["cashback", "qris"], "category": "Biaya Admin Bank", "sheet_contains": None},
    {"any": ["cashback mdr"], "category": "Biaya Admin Bank", "sheet_contains": None},
    {"any": ["cashback jago"], "category": "Biaya Admin Bank", "sheet_contains": "jago"},
    {"any": ["dr koreksi bunga"], "category": "Biaya Admin Bank", "sheet_contains": None},
    {"any": ["layanan"], "category": "Belanja Operasional", "sheet_contains": "jago"},
    {"any": ["fb", "facebook", "meta ads"], "category": "Marketing", "sheet_contains": None},
    {"any": ["sponsorship"], "category": "Marketing", "sheet_contains": None},
    {"any": ["masuya graha trikencana", "sukanda", "dineta"], "category": "Belanja Bahan", "sheet_contains": None},
    {"any": ["muh yani sh", "muh. yani sh", "muhammad yani sh"], "category": "Pembayaran Hutang", "sheet_contains": None},
    {"any": ["hutang", "pinjaman"], "direction": "masuk", "category": "Modal & Setoran Pemilik", "sheet_contains": None},
    {"any": ["hutang", "pinjaman"], "direction": "keluar", "category": "Pembayaran Hutang", "sheet_contains": None},
    {"any": ["setoran via cdm"], "category": "Transaksi Internal", "sheet_contains": None},
    {"any": ["tarik tunai qris"], "category": "Penjualan", "sheet_contains": None},
    {"any": ["listrik"], "category": "Belanja Operasional", "sheet_contains": None},
    {"any": ["yulia indah pratiwi", "yulia indah pratiw", "anugerah plastik"], "category": "Belanja Operasional", "sheet_contains": None},
    {"any": ["minus", "lebih", "cust", "tip", "tips"], "category": "Tip/Minus/Lebih", "sheet_contains": None},
]
# Dimuat dari shared_rules.json (dipakai bersama reconbot & bank-statement-bot)
# kalau ada; kalau file/kunci tidak ada, pakai daftar default di atas.
CATEGORY_OVERRIDE_RULES = shared_rules.get("category_override_rules", _DEFAULT_CATEGORY_OVERRIDE_RULES)

_PROTECTED_FROM_CATEGORY_OVERRIDE = set(shared_rules.get("protected_from_category_override", [
    "modal & setoran pemilik", "modal dan setoran pemilik", "laba ditahan bulanan",
    "saldo awal", "saldo awal bulan", "modal",
]))


def _override_keyword_found(pattern, text):
    return re.search(r"\b" + re.escape(pattern) + r"\b", text) is not None


# Sebagian bank menulis Keterangan Transaksi sebagai KODE ANGKA PANJANG
# (mis. nomor rekening pengirim diulang) alih-alih teks deskriptif -
# kadang malah rusak jadi notasi ilmiah kalau kolomnya kebetulan
# terbaca sebagai angka oleh Excel/software lain (mis. "1.57e+33").
# User menegaskan pola ini = transfer internal MASUK (kredit).
_LONG_NUMERIC_KETERANGAN_RE = re.compile(r"^\d{10,}$")
_SCI_NOTATION_KETERANGAN_RE = re.compile(r"^\d(\.\d+)?e\+?\d+$", re.IGNORECASE)


def _looks_like_long_numeric_code(desc):
    text = str(desc if desc is not None else "").strip()
    if not text:
        return False
    return bool(_LONG_NUMERIC_KETERANGAN_RE.match(text)) or bool(_SCI_NOTATION_KETERANGAN_RE.match(text))


# Kategori saldo awal -> dipakai untuk saldo awal Neraca, dilewati saat
# menjumlah transaksi berjalan.
OPENING_KEYWORDS = ["saldo awal"]

# Alias yang dipakai di kolom Subjek/Objek untuk merujuk rekening lain.
# Kunci = potongan teks yang mungkin muncul (huruf kecil), nilai = None
# (akan dicocokkan dengan resolve_account_alias terhadap nama sheet asli).
ACCOUNT_HINTS = ["bri-507", "bri-567", "bca-887", "bca-", "jago", "kasir"]

TOLERANCI_HARI = 30  # jendela pencarian pasangan transfer, sesuai prinsip audit
TOLERANSI_NOMINAL_PERSEN = 0.02  # 2% -> untuk toleransi biaya admin/pembulatan
TOLERANSI_NOMINAL_ABS = 5000  # atau selisih absolut di bawah ini dianggap wajar


@dataclass
class Txn:
    sheet: str
    row: int  # nomor baris di sheet asal (1-based, termasuk header)
    date: object
    desc: str
    kategori: str
    debit: float
    kredit: float
    saldo: float
    subjek: str
    objek: str
    ket: str

    @property
    def nominal(self):
        """Nilai transaksi bertanda: negatif jika debit (uang keluar),
        positif jika kredit (uang masuk)."""
        if self.debit:
            return self.debit  # sudah negatif di data sumber
        if self.kredit:
            return self.kredit
        return 0

    @property
    def is_transfer(self):
        k = (self.effective_kategori or "").lower()
        if any(kw in k for kw in TRANSFER_KEYWORDS):
            return True
        # fallback ke keterangan kalau kategori tidak/salah diisi, kecuali
        # sudah eksplisit dikategorikan sebagai modal (setoran dari luar,
        # bukan pindah antar rekening sendiri)
        if any(kw in k for kw in CAPITAL_KEYWORDS):
            return False
        d = (self.desc or "").lower()
        return any(kw in d for kw in DESC_TRANSFER_KEYWORDS)

    @property
    def is_tip_minus_variant(self):
        return (self.effective_kategori or "").strip().lower() == "tip/minus/lebih"

    @property
    def category_override(self):
        """Kategori pengganti berdasarkan CATEGORY_OVERRIDE_RULES (kata
        kunci di Keterangan/Keterangan Tambahan/Objek/Subjek, ditegaskan
        user berdasarkan pengalaman nyata bisnisnya) - None kalau tidak
        ada aturan yang cocok atau kategori aslinya sudah deliberate/tegas
        (modal, laba ditahan, saldo awal, gaji - lihat
        _PROTECTED_FROM_CATEGORY_OVERRIDE) dan tidak boleh ditimpa."""
        k = (self.kategori or "").strip().lower()
        if k in _PROTECTED_FROM_CATEGORY_OVERRIDE or k.startswith("gaji"):
            return None
        if self.nominal > 0 and _looks_like_long_numeric_code(self.desc):
            return "Transaksi Internal"
        text = f"{self.desc or ''} {self.ket or ''} {self.objek or ''} {self.subjek or ''}".lower()
        for rule in CATEGORY_OVERRIDE_RULES:
            sheet_filter = rule.get("sheet_contains")
            if sheet_filter and sheet_filter not in (self.sheet or "").lower():
                continue
            amount_min = rule.get("amount_min")
            amount_max = rule.get("amount_max")
            if amount_min is not None and abs(self.nominal) < amount_min:
                continue
            if amount_max is not None and abs(self.nominal) > amount_max:
                continue
            direction = rule.get("direction")  # "masuk" (kredit) / "keluar" (debit) / None (keduanya)
            if direction == "masuk" and self.nominal <= 0:
                continue
            if direction == "keluar" and self.nominal >= 0:
                continue
            if "any" in rule and not any(_override_keyword_found(kw, text) for kw in rule["any"]):
                continue
            if "all" in rule and not all(_override_keyword_found(kw, text) for kw in rule["all"]):
                continue
            return rule["category"]
        return None

    @property
    def effective_kategori(self):
        """Kategori yang SEBENARNYA dipakai untuk semua perhitungan (Laba
        Rugi, Neraca, pencocokan transfer, dst) - hasil category_override
        kalau ada yang cocok, kalau tidak ya Kategori asli dari data
        sumber apa adanya. SEMUA rumus SUMIF/SUMIFS kategori di laporan
        merujuk ke kolom bantu M (ditulis dari nilai ini), bukan langsung
        ke kolom C (Kategori asli), supaya override konsisten di semua
        laporan tanpa perlu duplikasi logika di tiap rumus."""
        return self.category_override or self.kategori

    @property
    def is_capital(self):
        k = (self.effective_kategori or "").lower()
        if any(kw in k for kw in CAPITAL_KEYWORDS):
            return True
        # kategori pendek "Modal" (tanpa "& Setoran Pemilik") tetap
        # dianggap modal - user kadang menyingkat, mis. Keterangan "Modal
        # Masuk" dengan Kategori cuma "Modal"
        if k.strip() == "modal" or k.strip().startswith("modal "):
            return True
        if "transfer masuk" in k:
            combined = f"{self.desc or ''} {self.ket or ''}".lower()
            if any(kw in combined for kw in CAPITAL_SELF_TRANSFER_KEYWORDS):
                return True
        return False

    @property
    def is_opening(self):
        k = (self.kategori or "").lower()
        return any(kw in k for kw in OPENING_KEYWORDS)

    def cell_ref(self, col):
        return f"'{self.sheet}'!${col}${self.row}"


# ---------------------------------------------------------------------------
# Membaca sheet rekening
# ---------------------------------------------------------------------------

def is_closing_summary_row(tanggal, kategori, keterangan, is_first_row=False):
    """Deteksi baris rekap penutup (Saldo Awal/Saldo Akhir/Total Debit/Total
    Kredit) yang kadang ada di baris-baris akhir sheet rekening sebagai
    ringkasan, BUKAN transaksi. Kalau ikut dimasukkan ke rekonstruksi saldo
    berjalan (kolom J/K/L), nilainya (yang merupakan TOTAL/ringkasan, bukan
    nominal transaksi tunggal) akan merusak saldo kumulatif dan jadi sumber
    selisih di Neraca.

    Syarat UTAMA: baris TIDAK punya tanggal (dan bukan baris data pertama -
    baris Saldo Awal/Saldo Awal Bulan yang legitimate di baris pertama
    kadang memang tidak diisi tanggal, itu bukan blok penutup). Transaksi
    asli SELALU bertanggal, termasuk checkpoint tengah bulan seperti 'Saldo
    akhir sesi' (checkpoint akhir shift kasir, muncul berkali-kali per
    bulan, tapi tetap punya tanggal asli) - kalau syarat tanpa-tanggal ini
    tidak dijadikan gerbang wajib untuk kata kunci juga, baris seperti itu
    akan salah dikira blok penutup HANYA karena mengandung teks 'saldo
    akhir', dan memotong rekonstruksi saldo jauh sebelum akhir bulan
    sesungguhnya."""
    if is_first_row or tanggal is not None:
        return False
    text = f"{kategori or ''} {keterangan or ''}".strip().lower()
    return bool(text)


def read_account_sheet(ws):
    """Baca satu sheet rekening jadi list[Txn], berhenti di baris kosong
    pertama setelah header (baris trailing kosong diabaikan), ATAU begitu
    ketemu baris rekap penutup (Saldo Akhir/Total Debit/Total Kredit) -
    baris itu dan seterusnya tidak dianggap transaksi, tapi nilainya
    ditangkap terpisah sebagai acuan cross-check (lihat closing_info)."""
    txns = []
    closing_info = {}
    seen_first_row = False
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        tanggal, ket, kategori, debit, kredit, saldo, subjek, objek, ket_tambahan = (
            (c.value for c in row[:9])
        )
        if tanggal is None and ket is None and debit is None and kredit is None:
            continue
        is_first = not seen_first_row
        seen_first_row = True
        if is_closing_summary_row(tanggal, kategori, ket, is_first_row=is_first):
            label = f"{kategori or ''} {ket or ''}".strip().lower()
            value = None
            for cand in (debit, kredit, saldo):
                if isinstance(cand, (int, float)):
                    value = cand
                    break
            if "total debit" in label:
                closing_info["total_debit"] = value
            elif "total kredit" in label:
                closing_info["total_kredit"] = value
            elif "saldo akhir" in label:
                closing_info["saldo_akhir"] = value
            continue  # bukan transaksi, jangan dimasukkan ke txns
        txns.append(
            Txn(
                sheet=ws.title,
                row=row[0].row,
                date=tanggal,
                # Semua kolom teks di-str()-kan eksplisit - beberapa file
                # sumber (mis. Numbers/Excel yang salah menebak tipe sel)
                # kadang menyimpan kolom teks (Keterangan dst) sebagai
                # ANGKA MENTAH (bahkan notasi ilmiah kalau angkanya sangat
                # panjang, mis. nomor rekening yang diulang di Keterangan)
                # - tanpa str() eksplisit, downstream code yang
                # mengasumsikan string (mis. .lower()) akan crash.
                desc=str(ket) if ket is not None else "",
                kategori=str(kategori) if kategori is not None else "",
                debit=float(debit) if isinstance(debit, (int, float)) else 0,
                kredit=float(kredit) if isinstance(kredit, (int, float)) else 0,
                saldo=float(saldo) if isinstance(saldo, (int, float)) else None,
                subjek=str(subjek) if subjek is not None else "",
                objek=str(objek) if objek is not None else "",
                ket=str(ket_tambahan) if ket_tambahan is not None else "",
            )
        )
    return txns, closing_info


FLIPTECH_SPLIT_REMAINDER_MAX = 1000  # sisa di bawah ini dianggap biaya
                                       # admin/bunga gabungan, bukan nominal genuine


def split_fliptech_combined_rows(ws):
    """Sebagian transaksi Fliptech tercatat sebagai SATU baris dengan
    nominal gabungan (mis. -131103 = -131000 transfer + -103 biaya admin,
    -67105 = -67000 + -105) - bukan dua baris terpisah seperti konvensi
    normal ("Bagian dari transaksi Fliptech: Biaya Admin" di baris
    nol-nominal terpisah). Kalau dibiarkan satu baris, transfer TIDAK
    AKAN PERNAH cocok dengan pasangannya di rekening lain (yang biasanya
    nominal genap/dibulatkan), karena selisih kecil (103/105/dst) itu di
    luar toleransi pencocokan normal.

    Deteksi pola ini (kategori transfer-like, Objek/Keterangan/Keterangan
    Tambahan menyebut 'fliptech', sisa nominal di bawah Rp1.000 dan bukan
    0) dan PECAH jadi 2 baris: baris asli jadi nominal genap (dibulatkan
    ke kelipatan 1000 terdekat ke arah nol), baris baru (disisipkan tepat
    sesudahnya) untuk sisanya sebagai Biaya Admin Bank (kalau debit) atau
    Bunga Bank (kalau kredit) - konsisten dengan konvensi Fliptech yang
    sudah ada di file lain.

    HARUS dipanggil PALING AWAL, sebelum last_data_row/add_helper_column/
    read_account_sheet - supaya penyisipan baris tidak merusak rumus/
    kolom bantu yang sudah ditulis di baris-baris sesudahnya."""
    row = 2
    while True:
        tanggal = ws.cell(row=row, column=1).value
        keterangan = ws.cell(row=row, column=2).value
        kategori = ws.cell(row=row, column=3).value
        if tanggal is None and keterangan is None and kategori is None:
            break  # sudah lewat baris data terakhir
        objek = ws.cell(row=row, column=8).value
        ket_tambahan = ws.cell(row=row, column=9).value
        text = f"{keterangan or ''} {objek or ''} {ket_tambahan or ''}".lower()
        is_transfer_like = isinstance(kategori, str) and any(kw in kategori.lower() for kw in TRANSFER_KEYWORDS)
        if is_transfer_like and "fliptech" in text:
            debit = ws.cell(row=row, column=4).value
            kredit = ws.cell(row=row, column=5).value
            is_debit = isinstance(debit, (int, float)) and debit
            nominal = debit if is_debit else (kredit if isinstance(kredit, (int, float)) else None)
            if isinstance(nominal, (int, float)) and nominal != 0:
                remainder = round(abs(nominal) % 1000, 2)
                if 0 < remainder < FLIPTECH_SPLIT_REMAINDER_MAX:
                    sign = 1 if nominal > 0 else -1
                    main_amount = sign * round(abs(nominal) - remainder, 2)
                    fee_amount = sign * remainder
                    subjek = ws.cell(row=row, column=7).value
                    saldo_tercatat = ws.cell(row=row, column=6).value
                    # baris asli jadi nominal genap; Saldo Kumulatif
                    # dikosongkan (bukan snapshot resmi bank lagi, cuma
                    # posisi ANTARA - baris baru di bawah yang bawa
                    # Saldo Kumulatif resmi hasil rekaman bank)
                    if is_debit:
                        ws.cell(row=row, column=4, value=main_amount)
                    else:
                        ws.cell(row=row, column=5, value=main_amount)
                    ws.cell(row=row, column=6, value=None)

                    ws.insert_rows(row + 1)
                    fee_label = "Biaya Admin Bank" if sign < 0 else "Bunga Bank"
                    note = (f"Bagian dari transaksi Fliptech: {fee_label} "
                            "(dipisah otomatis dari nominal gabungan oleh reconcile.py)")
                    ws.cell(row=row + 1, column=1, value=tanggal)
                    ws.cell(row=row + 1, column=2, value=note)
                    ws.cell(row=row + 1, column=3, value="Biaya Admin Bank")
                    if sign < 0:
                        ws.cell(row=row + 1, column=4, value=fee_amount)
                    else:
                        ws.cell(row=row + 1, column=5, value=fee_amount)
                    ws.cell(row=row + 1, column=6, value=saldo_tercatat)
                    ws.cell(row=row + 1, column=7, value="-")
                    ws.cell(row=row + 1, column=8, value=subjek)
                    ws.cell(row=row + 1, column=9, value=note)
                    row += 1  # lewati baris baru yang baru disisipkan
        row += 1


def last_data_row(ws):
    last = 1
    seen_first_row = False
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        tanggal, ket, kategori, debit, kredit = (c.value for c in row[:5])
        if not any(v is not None for v in (tanggal, ket, kategori, debit, kredit)):
            continue  # baris kosong, jangan dianggap baris data pertama
        is_first = not seen_first_row
        seen_first_row = True
        if is_closing_summary_row(tanggal, kategori, ket, is_first_row=is_first):
            break  # blok rekap penutup (Saldo Akhir/Total Debit/Kredit) -
            # berhenti di sini, jangan ikut dihitung sebagai baris transaksi
        last = row[0].row
    return last


def resolve_account_sheet(hint, sheet_names):
    """Cocokkan teks bebas di kolom Subjek/Objek (mis. 'BRI-567(Biz)',
    'Rekening Jago + Admin Rp200') ke nama sheet rekening sebenarnya."""
    if not hint:
        return None
    h = hint.lower()
    best = None
    for name in sheet_names:
        n = name.lower()
        # ambil token pembeda dari nama sheet, contoh nomor rekening / "jago" / "kasir"
        tokens = [t for t in n.replace("(", " ").replace(")", " ").split() if len(t) >= 3]
        for tok in tokens:
            if tok in h:
                best = name
                break
        if best:
            break
    return best


# ---------------------------------------------------------------------------
# Pencocokan transfer antar rekening ("Rekonsiliasi" inti)
# ---------------------------------------------------------------------------

@dataclass
class Match:
    src: Txn
    dst: Txn = None
    confidence: str = "Needs manual verification"
    reasoning: str = ""
    date_diff: int = None
    nominal_diff: float = None
    is_fliptech: bool = False


def days_between(a, b):
    """Selisih hari antara dua tanggal, menerima datetime.datetime,
    datetime.date, ATAU teks tanggal format umum (lihat coerce_date) - kalau
    cuma cek isinstance(datetime.date), file dengan kolom tanggal bertipe
    teks (mis. hasil parser 'preformatted' tertentu) akan selalu dianggap
    selisih 9999 hari (di luar toleransi), bikin SEMUA transfer di file itu
    gagal matched walau tanggal & nominalnya persis sama."""
    a = coerce_date(a)
    b = coerce_date(b)
    if a is None or b is None:
        return 9999
    return abs((a - b).days)


def _find_fliptech_loan_companion(src, all_txns):
    """Cari baris pendamping nol-nominal 'Bagian dari transaksi Fliptech: ...'
    di sheet & tanggal yang sama dengan src, yang keterangannya menyebut
    kata kunci pinjaman/cicilan/utang - sinyal bahwa transaksi src ini
    sebenarnya CICILAN PINJAMAN ke pihak luar (Fliptech Lentera Inspirasi
    sebagai penyedia pembiayaan), bukan transfer antar rekening sendiri,
    sehingga TIDAK AKAN PERNAH ketemu pasangannya di rekening manapun -
    beda dengan transfer internal biasa yang genuinely belum ketemu.
    Return teks catatan itu kalau ketemu, None kalau tidak."""
    loan_keywords = ["cicilan pinjaman", "cicilan", "pinjaman", "pembayaran utang", "angsuran"]
    src_date = coerce_date(src.date)
    for t in all_txns:
        if t is src or t.sheet != src.sheet or t.nominal != 0:
            continue
        if coerce_date(t.date) != src_date:
            continue
        text = (t.ket or "").lower()
        if not text.startswith("bagian dari transaksi fliptech"):
            continue
        for kw in loan_keywords:
            if kw in text:
                return t.ket
    return None


def find_matches(all_txns, sheet_names):
    """Untuk setiap transaksi bertanda transfer internal, cari pasangan di
    rekening tujuan (berdasarkan Subjek/Objek) dalam jendela +-30 hari,
    dengan toleransi selisih nominal untuk biaya admin/pembulatan.
    Tidak pernah menyimpulkan 'tidak ditemukan' tanpa mencoba seluruh
    kandidat di rekening tujuan terlebih dahulu."""
    # nominal 0 dikecualikan dari pencocokan: baris seperti ini biasanya
    # penanda/referensi dari parser sumber (mis. "Bagian dari transaksi
    # Fliptech: ...") bukan perpindahan uang sungguhan - tidak ada nominal
    # untuk dicocokkan, jadi kalau ikut diproses selalu nyangkut sebagai
    # "Needs manual verification" tanpa nilai informasi apapun
    transfers = [t for t in all_txns if t.is_transfer and t.nominal != 0]
    matched_dst_ids = set()
    consumed_ids = set()  # baik src maupun dst yang sudah punya pasangan
    results = []

    for src in transfers:
        if id(src) in consumed_ids:
            # sudah tercatat sebagai pasangan (dst) dari transaksi lain,
            # tidak perlu dilaporkan dua kali dari sisi yang berlawanan
            continue
        # tentukan rekening lawan dari Subjek/Objek (siapa pun yang BUKAN
        # rekening sheet sumber itu sendiri)
        counterpart_hint = None
        for hint in (src.objek, src.subjek):
            resolved = resolve_account_sheet(hint, sheet_names)
            if resolved and resolved != src.sheet:
                counterpart_hint = resolved
                break

        candidates = [
            t
            for t in all_txns
            if t is not src
            and t.is_transfer
            and t.nominal != 0
            and id(t) not in consumed_ids
            and (counterpart_hint is None or t.sheet == counterpart_hint)
            and (
                (t.nominal > 0) != (src.nominal > 0)  # tanda berlawanan (normal)
                or counterpart_hint is not None  # atau rekening tujuan sudah
                # pasti dari Subjek/Objek -> longgarkan syarat tanda, karena
                # sebagian pencatatan "Pindah Rekening Internal" tidak
                # konsisten memakai tanda negatif untuk uang keluar
            )
        ]
        if not candidates and counterpart_hint is None:
            # tidak ada petunjuk rekening tujuan -> perluas ke semua sheet lain
            candidates = [
                t
                for t in all_txns
                if t is not src
                and t.is_transfer
                and t.nominal != 0
                and id(t) not in consumed_ids
                and t.sheet != src.sheet
                and (t.nominal > 0) != (src.nominal > 0)
            ]

        # skor tiap kandidat: utamakan yang tandanya berlawanan (pencatatan
        # normal), baru selisih tanggal, baru selisih nominal
        scored = []
        for c in candidates:
            nominal_diff = abs(abs(src.nominal) - abs(c.nominal))
            date_diff = days_between(src.date, c.date)
            if date_diff > TOLERANCI_HARI:
                continue
            toleransi = max(TOLERANSI_NOMINAL_ABS, abs(src.nominal) * TOLERANSI_NOMINAL_PERSEN)
            if nominal_diff > toleransi:
                continue
            same_sign = (c.nominal > 0) == (src.nominal > 0)
            scored.append((1 if same_sign else 0, date_diff, nominal_diff, c, same_sign))

        scored.sort(key=lambda x: (x[0], x[1], x[2]))

        if scored:
            _, date_diff, nominal_diff, dst, same_sign = scored[0]
            matched_dst_ids.add(id(dst))
            consumed_ids.add(id(src))
            consumed_ids.add(id(dst))
            is_fliptech = (
                "fliptech" in (src.desc or "").lower()
                or "fliptech" in (dst.desc or "").lower()
                or "fliptech" in (src.ket or "").lower()
                or "fliptech" in (dst.ket or "").lower()
            )
            if nominal_diff == 0 and date_diff == 0:
                conf = "High"
                reason = "Nominal dan tanggal sama persis di kedua rekening."
            elif nominal_diff == 0 and date_diff <= 3:
                conf = "High"
                reason = f"Nominal sama persis, selisih tanggal {date_diff} hari (wajar untuk settlement bank)."
            elif is_fliptech and nominal_diff <= FLIPTECH_FEE_THRESHOLD:
                conf = "High"
                reason = (
                    f"Transaksi via Fliptech, selisih nominal Rp{nominal_diff:,.0f} "
                    f"dipastikan biaya admin transfer (bukan keraguan), selisih tanggal "
                    f"{date_diff} hari."
                ).replace(",", ".")
            elif nominal_diff > 0 and nominal_diff <= TOLERANSI_NOMINAL_ABS:
                conf = "Medium"
                reason = (
                    f"Selisih nominal Rp{nominal_diff:,.0f} kemungkinan biaya admin/transfer, "
                    f"selisih tanggal {date_diff} hari."
                ).replace(",", ".")
            elif date_diff > 3:
                conf = "Medium"
                reason = f"Nominal cocok (toleransi Rp{nominal_diff:,.0f}) tapi tanggal berbeda {date_diff} hari, kemungkinan delayed posting/settlement.".replace(",", ".")
            else:
                conf = "Low"
                reason = f"Kecocokan hanya berdasarkan toleransi umum, perlu verifikasi manual (selisih Rp{nominal_diff:,.0f}, {date_diff} hari).".replace(",", ".")
            if same_sign:
                # tanda sama-sama positif/negatif di kedua rekening -
                # dicocokkan lewat Subjek/Objek + nominal, bukan lewat tanda,
                # karena penulisan tanda di salah satu sisi kemungkinan keliru
                reason += (
                    " Catatan: kedua sisi tercatat dengan tanda yang sama "
                    "(bukan berlawanan) - kemungkinan salah input tanda "
                    "debit/kredit di salah satu rekening, cek manual."
                )
                if conf == "High":
                    conf = "Medium"
            results.append(
                Match(src=src, dst=dst, confidence=conf, reasoning=reason,
                      date_diff=date_diff, nominal_diff=nominal_diff, is_fliptech=is_fliptech)
            )
        else:
            loan_note = _find_fliptech_loan_companion(src, all_txns)
            if loan_note is not None:
                results.append(
                    Match(
                        src=src,
                        dst=None,
                        confidence="Not applicable (bukan transfer internal)",
                        reasoning=(
                            "Baris pendamping Fliptech nol-nominal di baris yang sama "
                            f"eksplisit menyebut \"{loan_note}\" - ini kemungkinan besar "
                            "CICILAN PINJAMAN/PEMBIAYAAN ke pihak luar (Fliptech Lentera "
                            "Inspirasi bertindak sebagai penyedia pembiayaan, bukan sesama "
                            "rekening internal yang kita lacak), BUKAN transfer antar "
                            "rekening sendiri. Transaksi seperti ini TIDAK AKAN PERNAH "
                            "punya pasangan di rekening manapun - pertimbangkan minta "
                            "bank-statement-bot mengkategorikan ulang jadi kategori "
                            "cicilan/utang tersendiri (bukan Transaksi Internal) supaya "
                            "tidak terus muncul di sini."
                        ),
                    )
                )
                continue
            results.append(
                Match(
                    src=src,
                    dst=None,
                    confidence="Needs manual verification",
                    reasoning=(
                        "Tidak ada kandidat dengan nominal berlawanan dalam jendela "
                        f"±{TOLERANCI_HARI} hari di rekening tujuan yang terindikasi "
                        f"({counterpart_hint or 'tidak teridentifikasi dari Subjek/Objek'}). "
                        "Kemungkinan: dana masih dalam perjalanan (in-transit), tercatat di "
                        "bulan berikutnya, atau salah kategori."
                    ),
                )
            )
    combo_results = find_split_merge_matches(results)
    return results, combo_results


def find_split_merge_matches(results):
    """Prinsip audit #4/#6: transaksi bisa terpecah (1 keluar -> 2 masuk)
    atau tergabung (2 keluar -> 1 masuk). Cari di antara sisa transfer yang
    belum ketemu pasangannya (Needs manual verification), apakah kombinasi
    2 transaksi lain menjumlah ke nominal yang cocok, dalam jendela waktu
    yang wajar."""
    import itertools

    unmatched = [m for m in results if m.dst is None]
    pool = [m.src for m in unmatched]
    used = set()
    combos = []

    # urutkan berdasarkan nominal terbesar dulu supaya transaksi induk
    # (yang paling mungkin "dipecah") diproses lebih dulu
    unmatched_sorted = sorted(unmatched, key=lambda m: -abs(m.src.nominal))

    for m in unmatched_sorted:
        src = m.src
        if id(src) in used:
            continue
        candidates = [
            t for t in pool
            if id(t) not in used
            and t is not src
            and (t.nominal > 0) != (src.nominal > 0)
            and days_between(src.date, t.date) <= TOLERANCI_HARI
        ]
        best = None
        for a, b in itertools.combinations(candidates, 2):
            total = abs(a.nominal) + abs(b.nominal)
            diff = abs(total - abs(src.nominal))
            toleransi = max(TOLERANSI_NOMINAL_ABS, abs(src.nominal) * TOLERANSI_NOMINAL_PERSEN)
            if diff <= toleransi:
                if best is None or diff < best[0]:
                    best = (diff, a, b)
        if best:
            diff, a, b = best
            used.add(id(src))
            used.add(id(a))
            used.add(id(b))
            conf = "High" if diff == 0 else "Medium"
            reason = (
                f"Kemungkinan transaksi terpecah/tergabung: {src.sheet} Rp{abs(src.nominal):,.0f} "
                f"~= {a.sheet} Rp{abs(a.nominal):,.0f} + {b.sheet} Rp{abs(b.nominal):,.0f} "
                f"(selisih Rp{diff:,.0f})."
            ).replace(",", ".")
            if isinstance(src.ket, str) and src.ket.strip():
                reason += f" Catatan asal: \"{src.ket}\""
            combos.append({"src": src, "parts": [a, b], "diff": diff, "confidence": conf, "reasoning": reason})

    return combos


FLIPTECH_FEE_THRESHOLD = shared_rules.get("fliptech_fee_threshold", 2000)
TIP_MINUS_THRESHOLD = shared_rules.get("tip_minus_threshold", 100000)


def compute_balance_status(all_txns_by_sheet):
    """Hitung ulang di Python (bukan tunggu Excel/LibreOffice recalculate
    rumus Neraca) apakah tiap rekening bakal balanced atau masih ada
    selisih - dipakai untuk menulis status ini LANGSUNG ke sheet
    Rekonsiliasi supaya user tidak perlu buka sheet Neraca terpisah untuk
    tahu ada masalah atau tidak; begitu laporan digenerate, statusnya
    sudah kelihatan di satu tempat. Mirror persis logika Excel formula di
    write_income_statement/write_balance_sheet (basis kas penuh, TIDAK
    ada penyusutan Aset Tetap - beda dengan quarterly.py/annual.py yang
    mengkapitalisasi 'Belanja Assets')."""
    def sum_exact(txns, category):
        cat = category.strip().lower()
        return sum(t.nominal for t in txns if (t.effective_kategori or "").strip().lower() == cat)

    def sum_multi(txns, categories):
        cats = {c.strip().lower() for c in categories}
        return sum(t.nominal for t in txns if (t.effective_kategori or "").strip().lower() in cats)

    def sum_gaji(txns):
        return sum(t.nominal for t in txns if (t.effective_kategori or "").strip().lower().startswith("gaji"))

    def sum_tip_minus(txns):
        return sum(t.nominal for t in txns if t.is_tip_minus_variant)

    def sum_modal(txns):
        return sum(t.nominal for t in txns if t.is_capital)

    results = {}
    for sheet, txns in all_txns_by_sheet.items():
        if not txns:
            continue
        # rekonstruksi K persis seperti rumus Excel (K2=F2, K(n)=K(n-1)+J(n))
        opening_f = txns[0].saldo if isinstance(txns[0].saldo, (int, float)) else (txns[0].nominal or 0)
        k_values = [opening_f]
        for t in txns[1:]:
            k_values.append(k_values[-1] + t.nominal)
        total_aset = k_values[-1]

        # "Saldo Awal Bulan" di Neraca MERUJUK ke K pada baris yang eksplisit
        # bertanda is_opening (bisa jadi BUKAN baris pertama - lihat kasus
        # nyata: sheet dengan transaksi biasa tercatat sebelum baris "Saldo
        # Awal"-nya sendiri) - kalau langsung pakai txns[0].saldo mentah
        # tanpa rekonstruksi, hasilnya bisa beda dari yang sungguhan dipakai
        # rumus Excel, dan prediksi Selisih ini jadi tidak akurat.
        opening_idx = next((i for i, t in enumerate(txns) if t.is_opening), 0)
        saldo_awal = k_values[opening_idx]

        revenue = sum_multi(txns, INCOME_CATEGORIES_REVENUE)
        expense = sum(sum_exact(txns, cat) for cat in INCOME_CATEGORIES_EXPENSE)
        expense += sum_multi(txns, MARKETING_RND_CATEGORY_TEXTS)
        expense += sum_gaji(txns)
        expense += sum_multi(txns, BANK_FEE_CATEGORY_TEXTS)
        other = sum(
            sum_tip_minus(txns) if cat == "Tip/Minus/Lebih" else sum_exact(txns, cat)
            for cat in OTHER_CATEGORIES
        )
        laba_bersih = revenue + expense + other
        modal = sum_modal(txns)
        ekuitas = saldo_awal + modal + laba_bersih
        transfer_bersih = sum_multi(txns, TRANSFER_CATEGORY_TEXTS)
        selisih = round((total_aset - ekuitas) - transfer_bersih, 2)
        results[sheet] = {
            "total_aset": round(total_aset, 2),
            "ekuitas": round(ekuitas, 2),
            "transfer_bersih": round(transfer_bersih, 2),
            "selisih": selisih,
        }
    return results


def find_minus_flags(all_txns_by_sheet):
    """Kumpulkan indikasi 'minus' yang perlu verifikasi manual: kategori
    Tip/Minus/Lebih dengan nominal > Rp100.000, atau baris berpenanda flag
    (⚑) di keterangan.

    Catatan: Tip/Minus/Lebih dengan nominal <= Rp100.000 dianggap valid
    (wajar terjadi dari pembulatan/kembalian kasir sehari-hari), tidak perlu
    ditandai untuk verifikasi manual.

    Saldo kumulatif negatif di tengah data juga TIDAK dianggap indikasi
    minus - itu cuma efek sementara dari urutan penulisan transaksi (baris
    tertulis belum tentu urut kronologis sempurna), bukan minus riil.
    Kesehatan saldo yang sebenarnya dicek di level akhir bulan lewat kolom
    K (Saldo Kumulatif Rekonstruksi) tiap sheet dan sheet Neraca.

    Cross-check baris-per-baris terhadap saldo hasil rekonstruksi (opening +
    akumulasi nominal) juga sengaja TIDAK dipakai di sini karena pada
    sebagian sheet sumber (mis. BCA) kolom Saldo Kumulatif tidak selalu
    diisi berurutan per baris (beberapa transaksi bertanggal sama dikelompokkan
    dulu), sehingga cross-check per baris menghasilkan banyak false positive."""
    flags = []
    for sheet, txns in all_txns_by_sheet.items():
        for t in txns:
            reasons = []
            if t.is_tip_minus_variant and abs(t.nominal) > TIP_MINUS_THRESHOLD:
                reasons.append(
                    f"Kategori/keterangan menyebut varian Tip/Minus/Lebih (mis. minus/lebih/tip/"
                    f"uang cust) dengan nominal Rp{abs(t.nominal):,.0f} "
                    f"(di atas ambang wajar Rp{TIP_MINUS_THRESHOLD:,.0f})."
                    .replace(",", ".")
                )
            if isinstance(t.ket, str) and "⚑" in t.ket:
                reasons.append(f"Ditandai perlu verifikasi manual: {t.ket}")
            if reasons:
                flags.append((t, reasons))
    return flags


# ---------------------------------------------------------------------------
# Penulisan sheet Rekonsiliasi
# ---------------------------------------------------------------------------

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def conf_fill(conf):
    return {"High": HIGH_FILL, "Medium": MED_FILL, "Low": LOW_FILL}.get(conf, LOW_FILL)


def write_rekonsiliasi_sheet(wb, matches, combo_matches, minus_flags, balance_status=None):
    if "Rekonsiliasi" in wb.sheetnames:
        del wb["Rekonsiliasi"]
    ws = wb.create_sheet("Rekonsiliasi")

    # Urutkan tiap bagian berdasarkan tanggal transaksi (bukan urutan sheet
    # rekening lalu baris) - memudahkan audit karena bisa ditelusuri
    # kronologis lintas semua rekening sekaligus, bukan per-rekening dulu
    # baru pindah ke rekening berikutnya. coerce_date dipakai karena
    # sebagian file punya kolom tanggal bertipe teks; None/tak terbaca
    # ditaruh paling akhir (bukan dianggap "paling awal") biar tidak
    # menyembunyikan baris bermasalah di atas.
    def _sort_key(d):
        parsed = coerce_date(d)
        return (parsed is None, parsed or datetime.date.max)

    matches = sorted(matches, key=lambda m: _sort_key(m.src.date))
    combo_matches = sorted(combo_matches, key=lambda cm: _sort_key(cm["src"].date))
    minus_flags = sorted(minus_flags, key=lambda tf: _sort_key(tf[0].date))

    ws["A1"] = "REKONSILIASI ANTAR REKENING"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Dibuat otomatis. Setiap baris merujuk langsung (link rumus) ke sel asal di sheet "
                "rekening. Tiap bagian diurutkan berdasarkan tanggal (lintas semua rekening) "
                "supaya gampang ditelusuri kronologis saat audit.")
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    # --- Bagian 1: pencocokan transfer antar rekening ---
    r = 4
    ws.cell(row=r, column=1, value="1. PENCOCOKAN TRANSFER ANTAR REKENING")
    ws.cell(row=r, column=1).font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    r += 1

    headers = [
        "Rekening Asal", "Tanggal Asal", "Keterangan Asal", "Nominal Asal",
        "Rekening Tujuan", "Tanggal Tujuan", "Keterangan Tujuan", "Nominal Tujuan",
        "Selisih Tanggal (hari)", "Selisih Nominal (Rp)", "Confidence", "Alasan Audit",
        "Via Fliptech?", "Rekening Penanggung Biaya", "Biaya Admin Teridentifikasi (Rp)",
    ]
    hdr_row = r
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hdr_row, column=i, value=h)
    style_header(ws, hdr_row, len(headers))
    r += 1

    for m in matches:
        ws.cell(row=r, column=1, value=m.src.sheet)
        ws.cell(row=r, column=2, value=f"='{m.src.sheet}'!$A${m.src.row}")
        ws.cell(row=r, column=2).number_format = DATE_FORMAT
        ws.cell(row=r, column=3, value=f"='{m.src.sheet}'!$B${m.src.row}")
        ws.cell(row=r, column=4, value=f"='{m.src.sheet}'!${'D' if m.src.debit else 'E'}${m.src.row}")
        ws.cell(row=r, column=4).number_format = NUMBER_FORMAT
        if m.dst:
            ws.cell(row=r, column=5, value=m.dst.sheet)
            ws.cell(row=r, column=6, value=f"='{m.dst.sheet}'!$A${m.dst.row}")
            ws.cell(row=r, column=6).number_format = DATE_FORMAT
            ws.cell(row=r, column=7, value=f"='{m.dst.sheet}'!$B${m.dst.row}")
            ws.cell(row=r, column=8, value=f"='{m.dst.sheet}'!${'D' if m.dst.debit else 'E'}${m.dst.row}")
            ws.cell(row=r, column=8).number_format = NUMBER_FORMAT
            ws.cell(row=r, column=9, value=m.date_diff)
            # Selisih Nominal dibuat rumus (bukan angka mati) supaya tetap
            # akurat kalau nominal di sheet rekening diedit ulang
            ws.cell(row=r, column=10, value=f"=ABS(ABS($D{r})-ABS($H{r}))")
            ws.cell(row=r, column=10).number_format = NUMBER_FORMAT
        else:
            ws.cell(row=r, column=5, value="(belum ditemukan)")
        ws.cell(row=r, column=11, value=m.confidence)
        ws.cell(row=r, column=12, value=m.reasoning)
        # kolom biaya admin (auto): dipakai Laporan Laba Rugi untuk membukukan
        # selisih transfer via Fliptech sebagai beban riil, bukan dibiarkan
        # menghilang sebagai selisih Neraca yang perlu intervensi manual
        ws.cell(row=r, column=13, value=m.is_fliptech if m.dst else False)
        if m.dst:
            fee_payer = m.src.sheet if m.src.debit else m.dst.sheet
            ws.cell(row=r, column=14, value=fee_payer)
            ws.cell(row=r, column=15,
                    value=f'=IF(AND($M{r}=TRUE,$H{r}<>""),ABS(ABS($D{r})-ABS($H{r})),0)')
        else:
            ws.cell(row=r, column=14, value="-")
            ws.cell(row=r, column=15, value=0)
        ws.cell(row=r, column=15).number_format = NUMBER_FORMAT
        fill = conf_fill(m.confidence)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 12))
        ws.cell(row=r, column=11).fill = fill
        r += 1

    section1_data_start = hdr_row + 1
    section1_last_row = max(r - 1, section1_data_start)  # dipakai Laporan Laba Rugi (SUMIFS biaya admin)
    r += 1
    # --- Bagian 1b: transfer terpecah / tergabung (split & merge) ---
    ws.cell(row=r, column=1, value="1b. TRANSFER TERPECAH / TERGABUNG (SPLIT & MERGE)")
    ws.cell(row=r, column=1).font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    r += 1
    headers1b = [
        "Rekening Asal", "Tanggal Asal", "Nominal Asal (Total)",
        "Pasangan 1", "Nominal 1", "Pasangan 2", "Nominal 2",
        "Selisih (Rp)", "Confidence", "Alasan Audit",
    ]
    hdr_row1b = r
    for i, h in enumerate(headers1b, start=1):
        ws.cell(row=hdr_row1b, column=i, value=h)
    style_header(ws, hdr_row1b, len(headers1b))
    r += 1
    for cm in combo_matches:
        src, a, b = cm["src"], cm["parts"][0], cm["parts"][1]
        ws.cell(row=r, column=1, value=src.sheet)
        ws.cell(row=r, column=2, value=f"='{src.sheet}'!$A${src.row}")
        ws.cell(row=r, column=2).number_format = DATE_FORMAT
        ws.cell(row=r, column=3, value=f"='{src.sheet}'!${'D' if src.debit else 'E'}${src.row}")
        ws.cell(row=r, column=3).number_format = NUMBER_FORMAT
        ws.cell(row=r, column=4, value=f"{a.sheet} (brs {a.row})")
        ws.cell(row=r, column=5, value=f"='{a.sheet}'!${'D' if a.debit else 'E'}${a.row}")
        ws.cell(row=r, column=5).number_format = NUMBER_FORMAT
        ws.cell(row=r, column=6, value=f"{b.sheet} (brs {b.row})")
        ws.cell(row=r, column=7, value=f"='{b.sheet}'!${'D' if b.debit else 'E'}${b.row}")
        ws.cell(row=r, column=7).number_format = NUMBER_FORMAT
        ws.cell(row=r, column=8, value=round(cm["diff"], 2))
        ws.cell(row=r, column=8).number_format = NUMBER_FORMAT
        ws.cell(row=r, column=9, value=cm["confidence"])
        ws.cell(row=r, column=10, value=cm["reasoning"])
        for c in range(1, len(headers1b) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 10))
        ws.cell(row=r, column=9).fill = conf_fill(cm["confidence"])
        r += 1
    if not combo_matches:
        ws.cell(row=r, column=1, value="(tidak ada indikasi transfer terpecah/tergabung)")
        ws.cell(row=r, column=1).font = Font(italic=True, color="6B7280")
        r += 1

    r += 1
    # --- Bagian 2: minus / selisih kas perlu verifikasi ---
    ws.cell(row=r, column=1, value="2. INDIKASI MINUS / SELISIH KAS")
    ws.cell(row=r, column=1).font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    r += 1
    headers2 = ["Rekening", "Tanggal", "Keterangan", "Nominal", "Saldo Kumulatif", "Alasan Perlu Verifikasi"]
    hdr_row2 = r
    for i, h in enumerate(headers2, start=1):
        ws.cell(row=hdr_row2, column=i, value=h)
    style_header(ws, hdr_row2, len(headers2))
    r += 1
    for t, reasons in minus_flags:
        ws.cell(row=r, column=1, value=t.sheet)
        ws.cell(row=r, column=2, value=f"='{t.sheet}'!$A${t.row}")
        ws.cell(row=r, column=2).number_format = DATE_FORMAT
        ws.cell(row=r, column=3, value=f"='{t.sheet}'!$B${t.row}")
        ws.cell(row=r, column=4, value=f"='{t.sheet}'!${'D' if t.debit else 'E'}${t.row}")
        ws.cell(row=r, column=4).number_format = NUMBER_FORMAT
        ws.cell(row=r, column=5, value=f"='{t.sheet}'!$F${t.row}")
        ws.cell(row=r, column=5).number_format = NUMBER_FORMAT
        ws.cell(row=r, column=6, value=" | ".join(reasons))
        for c in range(1, len(headers2) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 6))
        ws.cell(row=r, column=6).fill = LOW_FILL
        r += 1

    r += 1
    # --- Bagian 3: status keseimbangan Neraca (dihitung Python, prediksi
    # sebelum file dibuka/di-recalculate Excel) - supaya begitu laporan
    # digenerate, status "sudah selesai/masih ada selisih" langsung
    # kelihatan di sini tanpa perlu buka sheet Neraca terpisah ---
    ws.cell(row=r, column=1, value="3. STATUS KESEIMBANGAN NERACA")
    ws.cell(row=r, column=1).font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    r += 1
    if balance_status:
        headers3 = ["Rekening", "Total Aset", "Total Ekuitas", "Transfer Bersih", "Selisih", "Status"]
        hdr_row3 = r
        for i, h in enumerate(headers3, start=1):
            ws.cell(row=hdr_row3, column=i, value=h)
        style_header(ws, hdr_row3, len(headers3))
        r += 1
        any_selisih = False
        total_selisih = 0.0
        for sheet, s in balance_status.items():
            balanced = abs(s["selisih"]) < 1  # toleransi Rp1 (noise pembulatan)
            if not balanced:
                any_selisih = True
            total_selisih += s["selisih"]
            ws.cell(row=r, column=1, value=sheet)
            ws.cell(row=r, column=2, value=s["total_aset"])
            ws.cell(row=r, column=3, value=s["ekuitas"])
            ws.cell(row=r, column=4, value=s["transfer_bersih"])
            ws.cell(row=r, column=5, value=s["selisih"])
            ws.cell(row=r, column=6, value="Balanced" if balanced else f"ADA SELISIH Rp{abs(s['selisih']):,.0f}".replace(",", "."))
            for c in (2, 3, 4, 5):
                ws.cell(row=r, column=c).number_format = NUMBER_FORMAT
            for c in range(1, len(headers3) + 1):
                ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=6).fill = HIGH_FILL if balanced else LOW_FILL
            ws.cell(row=r, column=6).font = Font(bold=True)
            r += 1
        r += 1
        if any_selisih:
            ws.cell(row=r, column=1,
                    value=(f"BELUM SELESAI: masih ada selisih total Rp{abs(total_selisih):,.0f} yang belum "
                           "terjelaskan (lihat kolom Selisih per rekening di atas). Cek kembali bagian 1 & 2 "
                           "di atas dan sheet rekening masing-masing (kolom L) sebelum laporan ini dianggap final."
                           .replace(",", ".")))
            ws.cell(row=r, column=1).font = Font(bold=True, color="B91C1C")
        else:
            ws.cell(row=r, column=1, value="SELESAI: semua rekening balanced, tidak ada selisih yang perlu ditelusuri lebih lanjut.")
            ws.cell(row=r, column=1).font = Font(bold=True, color="15803D")
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers3))
        ws.row_dimensions[r].height = 28
        r += 1
    else:
        ws.cell(row=r, column=1, value="(status keseimbangan tidak dihitung)")
        ws.cell(row=r, column=1).font = Font(italic=True, color="6B7280")
        r += 1

    widths = [22, 12, 26, 14, 22, 12, 26, 14, 10, 14, 14, 40, 12, 22, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A6"
    return ws, {"data_start": section1_data_start, "data_end": section1_last_row}


# ---------------------------------------------------------------------------
# Penambahan kolom bantu "Nominal Bersih" di tiap sheet rekening
# (rumus, dipakai sebagai basis SUMIF di laporan keuangan)
# ---------------------------------------------------------------------------

def add_helper_column(ws, last_row):
    """Tambah 3 kolom bantu berbasis rumus (bukan nilai mati):
    J = Nominal Bersih (dipakai basis SUMIF laporan keuangan)
    K = Saldo Kumulatif Rekonstruksi (dihitung ulang dari saldo awal +
        akumulasi nominal, sehingga selalu terisi walau kolom F/Saldo
        Kumulatif aslinya bolong-bolong di sebagian baris)
    L = Selisih vs Saldo Tercatat (alat bantu telusur/audit: harus 0 setiap
        kali kolom F terisi; kalau tidak 0, baris-baris sebelumnya di sheet
        ini kemungkinan TIDAK berurutan secara kronologis terhadap kolom
        Saldo Kumulatif aslinya - baris dengan selisih besar ditandai warna
        kuning otomatis, dan dirangkum di sheet Diagnostik Keseimbangan)
    """
    for col, title in ((10, "Nominal Bersih (Debit atau Kredit)"),
                       (11, "Saldo Kumulatif (Rekonstruksi)"),
                       (12, "Selisih vs Saldo Tercatat")):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row in range(2, last_row + 1):
        # N() memperlakukan sel kosong sebagai 0 tanpa salah mengira nilai
        # 0 eksplisit sebagai "kosong" (beda dengan tes teks <>"")
        ws.cell(row=row, column=10, value=f"=N($D{row})+N($E{row})")
        if row == 2:
            ws.cell(row=row, column=11, value="=$F$2")
        else:
            ws.cell(row=row, column=11, value=f"=$K{row - 1}+$J{row}")
        ws.cell(row=row, column=12,
                value=f'=IF($F{row}<>"",$K{row}-$F{row},0)')

    for col, width in ((10, 26), (11, 28), (12, 22)):
        ws.column_dimensions[get_column_letter(col)].width = width

    # highlight visual: baris dengan penyimpangan signifikan (>Rp1.000)
    # antara saldo rekonstruksi dan saldo tercatat - penanda cepat untuk
    # menelusuri baris mana yang bikin data tidak berurutan/tidak konsisten
    from openpyxl.formatting.rule import CellIsRule
    rng = f"L2:L{last_row}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="greaterThan", formula=["1000"], fill=MED_FILL)
    )


def add_effective_category_column(ws, txns):
    """Kolom bantu M = Kategori Efektif - Kategori ASLI (kolom C), KECUALI
    ada aturan override yang cocok (lihat Txn.category_override/
    CATEGORY_OVERRIDE_RULES) - ditulis sebagai NILAI (bukan rumus, karena
    logikanya melibatkan pencarian kata kunci yang jauh lebih mudah
    dilakukan di Python daripada rumus Excel murni). SEMUA rumus SUMIF/
    SUMIFS kategori di laporan keuangan (Laba Rugi, Neraca, Arus Kas, dst)
    merujuk ke kolom INI, bukan langsung ke kolom C, supaya override
    konsisten di semua laporan tanpa perlu duplikasi logika di tiap
    rumus. Kolom C tetap dibiarkan apa adanya (data asli, untuk audit)."""
    cell = ws.cell(row=1, column=13, value="Kategori Efektif (setelah override kata kunci - dipakai semua rumus)")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    for t in txns:
        ws.cell(row=t.row, column=13, value=t.effective_kategori)
    ws.column_dimensions[get_column_letter(13)].width = 34


# ---------------------------------------------------------------------------
# Laporan Laba Rugi
# ---------------------------------------------------------------------------

INCOME_CATEGORIES_REVENUE = ["Penjualan"]

# Kategori beban yang dicocokkan persis apa adanya (SUMIF biasa)
INCOME_CATEGORIES_EXPENSE = [
    "Belanja Bahan",
    "Belanja Operasional",
    "Belanja Konsumsi",
    "Reparasi",
    "Belanja Assets",
]

# Marketing dan Riset dan Pengembangan (RnD) digabung jadi satu baris -
# dulu "Riset dan Pengembangan" (kategori baru, dipicu keyword "Pelatihan")
# tidak terdaftar sama sekali di INCOME_CATEGORIES_EXPENSE, jadi uangnya
# hilang dari Laba Rugi (sumber selisih Neraca di BCA).
MARKETING_RND_CATEGORY_TEXTS = ["Marketing", "Riset dan Pengembangan"]

# Gaji: dulu satu baris per "Gaji <Bulan> <Tahun>" (mis. "Gaji Desember 2024")
# Gaji: dulu satu baris per "Gaji <Bulan> <Tahun>" (mis. "Gaji Desember 2024")
# yang berarti daftar kategori harus terus ditambah tiap tahun, dan asumsi
# lama (info bulan ada di kolom Kategori) ternyata tidak berlaku di semua
# parser - versi bank ("preformatted") pakai Kategori tetap "Gaji Pegawai"
# untuk SEMUA gaji, info bulannya cuma ada di Keterangan (mis. "Gaji
# Latifatul Husna Januari"). Makanya deteksi bulan ini/lalu sekarang
# dicocokkan ke kolom KETERANGAN (bukan Kategori), dengan Kategori cuma
# dipakai untuk memastikan barisnya memang tentang gaji (wildcard "Gaji*").
def gaji_category_patterns(period_month):
    """Return (nama_bulan_ini, nama_bulan_lalu) berdasarkan nomor bulan
    periode (1-12). Kalau bulan tidak terdeteksi, return (None, None)."""
    if not period_month:
        return None, None
    prev_month = 12 if period_month == 1 else period_month - 1
    return MONTHS_ID[period_month], MONTHS_ID[prev_month]


def sumif_gaji_bulan_formula(sheet, last_row, nama_bulan):
    """SUMIFS: baris berkategori 'Gaji*' DAN keterangannya menyebut nama
    bulan tertentu (mis. '*Januari*') - menangani baik gaya lama (bulan ada
    di Kategori) maupun gaya baru/bank (Kategori tetap 'Gaji Pegawai', bulan
    cuma disebut di Keterangan)."""
    rng_c = f"'{sheet}'!$M$2:$M${last_row}"
    rng_b = f"'{sheet}'!$B$2:$B${last_row}"
    rng_j = f"'{sheet}'!$J$2:$J${last_row}"
    if not nama_bulan:
        return f"=SUMIF({rng_c},\"Gaji*\",{rng_j})"
    return f"=SUMIFS({rng_j},{rng_c},\"Gaji*\",{rng_b},\"*{nama_bulan}*\")"


# Biaya admin, biaya admin transfer (Fliptech, auto dari Rekonsiliasi), bunga,
# dan pajak bank digabung jadi SATU baris "Biaya Admin Bank" - dulu terpecah
# karena bank/versi parser beda pakai istilah berbeda (Biaya Admin & Pajak
# Bank / Biaya Admin dan Bunga Bank / Bunga dan Admin Bank / kini disatukan
# jadi "Biaya Admin Bank" di parser terbaru), padahal secara ekonomi
# sama-sama biaya jasa perbankan. Daftar lama tetap disertakan supaya file
# historis yang masih pakai istilah lama tetap tertangkap.
BANK_FEE_CATEGORY_TEXTS = [
    "Biaya Admin Bank",
    "Biaya Admin & Pajak Bank",
    "Biaya Admin dan Bunga Bank",
    "Bunga dan Admin Bank",
]

OTHER_CATEGORIES = ["Tip/Minus/Lebih", "Penarikan", "Penerimaan", "Pembayaran Hutang"]


# ---------------------------------------------------------------------------
# Helper pivot: setiap laporan keuangan ditulis per-rekening (kolom),
# dengan kolom paling kanan = TOTAL keseluruhan. Ini supaya selisih di
# Neraca/Diagnostik bisa langsung ditelusuri ke rekening mana penyebabnya,
# tanpa harus buka satu-satu.
# ---------------------------------------------------------------------------

def col_letter(i):
    return get_column_letter(i)


def pivot_total_col(sheets):
    return 2 + len(sheets)


def write_pivot_header(ws, row, sheets, label=""):
    ws.cell(row=row, column=1, value=label)
    for i, sheet in enumerate(sheets):
        c = 2 + i
        cell = ws.cell(row=row, column=c, value=sheet)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    total_col = pivot_total_col(sheets)
    ws.cell(row=row, column=total_col, value="TOTAL")
    style_header(ws, row, total_col)
    ws.row_dimensions[row].height = 32


def write_pivot_section(ws, row, label, sheets):
    """Baris judul section (mis. 'PENDAPATAN'), fill/bold di seluruh lebar tabel."""
    total_col = pivot_total_col(sheets)
    ws.cell(row=row, column=1, value=label)
    for c in range(1, total_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL


def write_pivot_data_row(ws, row, label, sheets, formula_fn, bold=False):
    """formula_fn(sheet) -> formula string (SUMIF per sheet, dll). Kolom
    TOTAL diisi SUM dari sel-sel per rekening di baris yang sama (bukan
    dihitung ulang terpisah), supaya konsisten dan gampang dicek manual."""
    ws.cell(row=row, column=1, value=label)
    n = len(sheets)
    for i, sheet in enumerate(sheets):
        c = 2 + i
        cell = ws.cell(row=row, column=c, value=formula_fn(sheet))
        cell.number_format = NUMBER_FORMAT
    total_col = 2 + n
    ws.cell(row=row, column=total_col,
            value=f"=SUM({col_letter(2)}{row}:{col_letter(1 + n)}{row})")
    ws.cell(row=row, column=total_col).number_format = NUMBER_FORMAT
    if bold:
        for c in range(1, total_col + 1):
            ws.cell(row=row, column=c).font = Font(bold=True)


def write_pivot_subtotal_row(ws, row, label, sheets, ref_rows, bold=True):
    """Subtotal per kolom = SUM baris-baris ref_rows di kolom yang sama.
    Kolom TOTAL = SUM sel-sel per rekening di baris subtotal itu sendiri."""
    ws.cell(row=row, column=1, value=label)
    n = len(sheets)
    for i in range(n):
        c = 2 + i
        cl = col_letter(c)
        ws.cell(row=row, column=c, value=f"=SUM({cl}{ref_rows[0]}:{cl}{ref_rows[-1]})")
        ws.cell(row=row, column=c).number_format = NUMBER_FORMAT
    total_col = 2 + n
    ws.cell(row=row, column=total_col,
            value=f"=SUM({col_letter(2)}{row}:{col_letter(1 + n)}{row})")
    ws.cell(row=row, column=total_col).number_format = NUMBER_FORMAT
    if bold:
        for c in range(1, total_col + 1):
            ws.cell(row=row, column=c).font = Font(bold=True)


def write_pivot_formula_row(ws, row, label, sheets, per_col_formula_fn, bold=False):
    """Baris hasil kombinasi rumus antar-baris (mis. Laba = Pendapatan+Beban),
    per_col_formula_fn(col_letter) -> formula string, dipakai sama persis
    untuk tiap kolom rekening MAUPUN kolom TOTAL (referensi sel berbeda,
    logika sama)."""
    ws.cell(row=row, column=1, value=label)
    n = len(sheets)
    for i in range(n):
        c = 2 + i
        cl = col_letter(c)
        ws.cell(row=row, column=c, value=per_col_formula_fn(cl))
        ws.cell(row=row, column=c).number_format = NUMBER_FORMAT
    total_col = 2 + n
    ws.cell(row=row, column=total_col, value=per_col_formula_fn(col_letter(total_col)))
    ws.cell(row=row, column=total_col).number_format = NUMBER_FORMAT
    if bold:
        for c in range(1, total_col + 1):
            ws.cell(row=row, column=c).font = Font(bold=True)


def sumif_one_sheet(sheet, last_row, category):
    return (f"=SUMIF('{sheet}'!$M$2:$M${last_row},\"{category}\","
            f"'{sheet}'!$J$2:$J${last_row})")


def sumif_modal_one_sheet(sheet, last_row):
    """Modal & Setoran Pemilik + kasus 'Transfer Masuk ... dari rekening
    sendiri' (uang milik owner sendiri dipindah antar rekening, mis.
    pencairan investasi pribadi) - lihat catatan di CAPITAL_SELF_TRANSFER_KEYWORDS.
    Ditambah 'Laba Ditahan Bulanan' (laba bulan berjalan yang disimpan,
    biasanya dimasukkan sebagai modal baru bulan berikutnya atau dana
    darurat) - user menegaskan ini dianggap kategori modal dari owner."""
    rng_c = f"'{sheet}'!$M$2:$M${last_row}"
    rng_b = f"'{sheet}'!$B$2:$B${last_row}"
    rng_j = f"'{sheet}'!$J$2:$J${last_row}"
    return (f"=SUMIF({rng_c},\"Modal*\",{rng_j})"
            f"+SUMIF({rng_c},\"Laba Ditahan Bulanan\",{rng_j})"
            f"+SUMIFS({rng_j},{rng_c},\"Transfer Masuk\",{rng_b},\"*rekening sendiri*\")")


def sumif_tip_minus_one_sheet(sheet, last_row):
    rng_c = f"'{sheet}'!$M$2:$M${last_row}"
    rng_j = f"'{sheet}'!$J$2:$J${last_row}"
    return f"=SUMIF({rng_c},\"Tip/Minus/Lebih\",{rng_j})"


def sumif_gaji_lainnya_formula(sheet, last_row, nama_bulan_list):
    """Jaring pengaman: tangkap semua baris berkategori 'Gaji*' TAPI
    keterangannya tidak menyebut bulan ini/lalu (mis. gaji utuh dari bulan
    lain, atau baris gaji tanpa nama bulan sama sekali) - supaya tidak ada
    beban gaji yang diam-diam hilang dari Laba Rugi."""
    rng_c = f"'{sheet}'!$M$2:$M${last_row}"
    rng_b = f"'{sheet}'!$B$2:$B${last_row}"
    rng_j = f"'{sheet}'!$J$2:$J${last_row}"
    parts = [f"SUMIF({rng_c},\"Gaji*\",{rng_j})"]
    for nama_bulan in nama_bulan_list:
        if nama_bulan:
            parts.append(f"SUMIFS({rng_j},{rng_c},\"Gaji*\",{rng_b},\"*{nama_bulan}*\")")
    return "=" + parts[0] + "".join(f"-{p}" for p in parts[1:])


def write_income_statement(wb, sheets_last_row, period_label, period_month, recon_range):
    name = "Laporan Laba Rugi"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    sheets = list(sheets_last_row.keys())
    ws["A1"] = f"LAPORAN LABA RUGI - {period_label.upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Pivot per rekening (rumus SUMIF beralamat absolut), kolom TOTAL paling kanan "
                "= jumlah keseluruhan. Bandingkan antar kolom untuk menelusuri selisih per rekening.")
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    r = 4
    write_pivot_header(ws, r, sheets)
    r += 1

    write_pivot_section(ws, r, "PENDAPATAN", sheets)
    r += 1
    rev_rows = []
    for cat in INCOME_CATEGORIES_REVENUE:
        write_pivot_data_row(ws, r, cat, sheets,
                              lambda sheet, cat=cat: sumif_one_sheet(sheet, sheets_last_row[sheet], cat))
        rev_rows.append(r)
        r += 1
    total_rev_row = r
    write_pivot_subtotal_row(ws, r, "Total Pendapatan", sheets, rev_rows)
    r += 2

    write_pivot_section(ws, r, "BEBAN", sheets)
    r += 1
    exp_rows = []
    for cat in INCOME_CATEGORIES_EXPENSE:
        write_pivot_data_row(ws, r, cat, sheets,
                              lambda sheet, cat=cat: sumif_one_sheet(sheet, sheets_last_row[sheet], cat))
        exp_rows.append(r)
        r += 1
    write_pivot_data_row(
        ws, r, "Marketing & RnD", sheets,
        lambda sheet: sumif_multi_one_sheet(sheet, sheets_last_row[sheet], MARKETING_RND_CATEGORY_TEXTS),
    )
    exp_rows.append(r)
    r += 1
    # Gaji: dicocokkan lewat KETERANGAN (bukan Kategori) supaya konsisten
    # baik format lama (bulan ada di Kategori) maupun format bank/preformatted
    # (Kategori tetap "Gaji Pegawai", bulan cuma disebut di Keterangan, mis.
    # "Gaji Latifatul Husna Januari"). "Gaji <bulan ini>" = beban berjalan,
    # "Gaji <bulan lalu>" = accrual, "Gaji Lainnya" = jaring pengaman.
    nama_ini, nama_lalu = gaji_category_patterns(period_month)
    write_pivot_data_row(ws, r, f"Gaji Bulan Ini (Gaji {nama_ini or 'Bulan Ini'})", sheets,
                          lambda sheet: sumif_gaji_bulan_formula(sheet, sheets_last_row[sheet], nama_ini))
    exp_rows.append(r)
    r += 1
    write_pivot_data_row(ws, r, f"Gaji Accrual (Gaji {nama_lalu or 'Accrual'})", sheets,
                          lambda sheet: sumif_gaji_bulan_formula(sheet, sheets_last_row[sheet], nama_lalu))
    exp_rows.append(r)
    r += 1
    write_pivot_data_row(
        ws, r, "Gaji Lainnya (bulan lain/tidak disebutkan)", sheets,
        lambda sheet: sumif_gaji_lainnya_formula(sheet, sheets_last_row[sheet], [nama_ini, nama_lalu]),
    )
    exp_rows.append(r)
    r += 1
    # Biaya Admin Bank: gabungan biaya admin bank, biaya admin
    # transfer (mis. via Fliptech, auto-terdeteksi dari Rekonsiliasi kolom
    # N/O), bunga, dan pajak bank - dulu terpecah jadi beberapa baris karena
    # tiap bank/versi parser pakai istilah beda, sekarang satu baris saja
    fee_row = r
    write_pivot_data_row(
        ws, r, "Biaya Admin Bank (termasuk biaya transfer Fliptech)", sheets,
        lambda sheet: (
            f"={sumif_multi_one_sheet(sheet, sheets_last_row[sheet], BANK_FEE_CATEGORY_TEXTS)[1:]}"
            f"-SUMIFS('Rekonsiliasi'!$O${recon_range['data_start']}:$O${recon_range['data_end']},"
            f"'Rekonsiliasi'!$N${recon_range['data_start']}:$N${recon_range['data_end']},\"{sheet}\")"
        ),
    )
    exp_rows.append(r)
    r += 1
    total_exp_row = r
    write_pivot_subtotal_row(ws, r, "Total Beban", sheets, exp_rows)
    r += 2

    write_pivot_section(ws, r, "LAIN-LAIN (perlu verifikasi manual)", sheets)
    r += 1
    other_rows = []
    for cat in OTHER_CATEGORIES:
        if cat == "Tip/Minus/Lebih":
            write_pivot_data_row(ws, r, cat, sheets,
                                  lambda sheet: sumif_tip_minus_one_sheet(sheet, sheets_last_row[sheet]))
        else:
            write_pivot_data_row(ws, r, cat, sheets,
                                  lambda sheet, cat=cat: sumif_one_sheet(sheet, sheets_last_row[sheet], cat))
        other_rows.append(r)
        r += 1
    total_other_row = r
    write_pivot_subtotal_row(ws, r, "Total Lain-lain", sheets, other_rows)
    r += 2

    net_row = r
    write_pivot_formula_row(
        ws, r, "LABA / RUGI BERSIH", sheets,
        lambda cl: f"={cl}{total_rev_row}+{cl}{total_exp_row}+{cl}{total_other_row}",
        bold=True,
    )
    for c in range(1, pivot_total_col(sheets) + 1):
        ws.cell(row=r, column=c).font = Font(bold=True, size=12)

    ws.column_dimensions["A"].width = 34
    for i in range(len(sheets)):
        ws.column_dimensions[col_letter(2 + i)].width = 16
    ws.column_dimensions[col_letter(pivot_total_col(sheets))].width = 18
    ws.freeze_panes = "B5"
    return ws, {"total_rev": total_rev_row, "total_exp": total_exp_row,
                "total_other": total_other_row, "net": net_row, "sheet": name,
                "sheets": sheets, "total_col": pivot_total_col(sheets)}


# ---------------------------------------------------------------------------
# Neraca (Balance Sheet)
# ---------------------------------------------------------------------------

TRANSFER_CATEGORY_TEXTS = [
    "Pindah Rekening Internal",
    "Pindang Rekening Internal",
    "Transfer Internal",
    "Transfer Lainnya",
    "Transaksi Internal",
]


def _validate_category_override_targets():
    """Cegah kelas bug yang pernah kejadian: kategori tujuan di
    CATEGORY_OVERRIDE_RULES harus PERSIS salah satu string yang benar-benar
    dicek rumus SUMIF/SUMIFS laporan keuangan (INCOME_CATEGORIES_EXPENSE/
    REVENUE, MARKETING_RND_CATEGORY_TEXTS, BANK_FEE_CATEGORY_TEXTS,
    OTHER_CATEGORIES, atau 'Modal & Setoran Pemilik') - BUKAN cuma label
    baris yang enak dibaca (mis. 'Marketing & RnD' pernah kepakai padahal
    yang benar-benar dicek SUMIF adalah 'Marketing' saja, uangnya jadi
    tidak ketangkap di manapun dan bikin Neraca selisih diam-diam)."""
    known = set(
        INCOME_CATEGORIES_EXPENSE + INCOME_CATEGORIES_REVENUE +
        MARKETING_RND_CATEGORY_TEXTS + BANK_FEE_CATEGORY_TEXTS +
        OTHER_CATEGORIES + TRANSFER_CATEGORY_TEXTS + ["Modal & Setoran Pemilik"]
    )
    for rule in CATEGORY_OVERRIDE_RULES:
        target = rule["category"]
        if target not in known:
            raise AssertionError(
                f"CATEGORY_OVERRIDE_RULES: kategori tujuan {target!r} bukan salah satu kategori "
                "yang dikenali rumus SUMIF/SUMIFS laporan keuangan - uang yang di-override ke sini "
                "tidak akan ketangkap di manapun. Perbaiki jadi salah satu dari: "
                f"{sorted(known)}"
            )


_validate_category_override_targets()


def sumif_multi_one_sheet(sheet, last_row, categories):
    parts = [
        f"SUMIF('{sheet}'!$M$2:$M${last_row},\"{cat}\",'{sheet}'!$J$2:$J${last_row})"
        for cat in categories
    ]
    return "=" + "+".join(parts)


def write_balance_sheet(wb, sheets_last_row, opening_rows, income_ref, period_end_label, recon_range):
    name = "Neraca"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    sheets = list(sheets_last_row.keys())
    ws["A1"] = f"NERACA - PER {period_end_label.upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Pivot per rekening, kolom TOTAL paling kanan = keseluruhan. Baris 'CEK KESEIMBANGAN' "
                "dan 'Selisih Belum Terjelaskan' per kolom langsung menunjukkan rekening mana yang "
                "selisih - lihat juga sheet Diagnostik Keseimbangan.")
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    r = 4
    write_pivot_header(ws, r, sheets)
    r += 1

    write_pivot_section(ws, r, "ASET (KAS & SETARA KAS)", sheets)
    r += 1
    kas_row = r
    write_pivot_data_row(ws, r, "Kas & Setara Kas (Saldo Akhir)", sheets,
                          lambda sheet: f"='{sheet}'!$K${sheets_last_row[sheet]}", bold=True)
    total_asset_row = kas_row
    r += 2

    write_pivot_section(ws, r, "EKUITAS", sheets)
    r += 1
    saldo_awal_row = r
    write_pivot_data_row(ws, r, "Saldo Awal Bulan", sheets,
                          lambda sheet: f"='{sheet}'!$K${opening_rows[sheet]}")
    r += 1
    modal_row = r
    write_pivot_data_row(ws, r, "Modal & Setoran Pemilik (+ Laba Ditahan Bulanan) (bulan ini)", sheets,
                          lambda sheet: sumif_modal_one_sheet(sheet, sheets_last_row[sheet]))
    r += 1
    laba_row = r
    # kolom rekening di Neraca urutannya sama dengan di Laporan Laba Rugi
    # (keduanya dari sheets_last_row.keys() yang sama), jadi tinggal pakai
    # huruf kolom yang sama untuk menautkan baris LABA/RUGI BERSIH per rekening
    write_pivot_formula_row(
        ws, r, "Laba Bersih Bulan Ini", sheets,
        lambda cl: f"='{income_ref['sheet']}'!{cl}{income_ref['net']}",
    )
    r += 1
    total_equity_row = r
    write_pivot_subtotal_row(ws, r, "Total Ekuitas", sheets, [saldo_awal_row, laba_row])
    r += 2

    balance_check_row = r
    write_pivot_formula_row(
        ws, r, "CEK KESEIMBANGAN (Aset - Ekuitas)", sheets,
        lambda cl: f"={cl}{total_asset_row}-{cl}{total_equity_row}",
        bold=True,
    )
    r += 1
    transfer_row = r
    # Transfer Bersih murni (SUMIF kategori transfer) DITAMBAH biaya admin
    # yang sudah "dipisahkan" jadi beban riil di Laba Rugi (baris Biaya Admin
    # Transfer) - supaya baris ini hanya berisi porsi transfer yang benar
    # dua sisinya matched, bukan lagi bercampur dengan biaya admin
    write_pivot_data_row(
        ws, r, "Transfer Bersih (rekening ini, info)", sheets,
        lambda sheet: (
            f"={sumif_multi_one_sheet(sheet, sheets_last_row[sheet], TRANSFER_CATEGORY_TEXTS)[1:]}"
            f"+SUMIFS('Rekonsiliasi'!$O${recon_range['data_start']}:$O${recon_range['data_end']},"
            f"'Rekonsiliasi'!$N${recon_range['data_start']}:$N${recon_range['data_end']},\"{sheet}\")"
        ),
    )
    r += 1
    residual_row = r
    write_pivot_formula_row(
        ws, r, "Selisih Belum Terjelaskan (Selisih - Transfer Bersih)", sheets,
        lambda cl: f"={cl}{balance_check_row}-{cl}{transfer_row}",
        bold=True,
    )
    r += 1
    ws.cell(row=r, column=1,
            value=("Transfer Bersih seharusnya saling menutup ~0 di kolom TOTAL (lihat sheet "
                   "Rekonsiliasi kalau tidak). Per rekening wajar tidak 0 (rekening itu bisa jadi "
                   "pengirim/penerima bersih bulan ini). Yang perlu ditelusuri adalah 'Selisih Belum "
                   "Terjelaskan' - kalau besar di satu rekening, itu tandanya data di sheet rekening "
                   "itu (lihat kolom L) yang bermasalah, bukan soal transfer."))
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="6B7280")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=pivot_total_col(sheets))
    ws.row_dimensions[r].height = 48

    ws.column_dimensions["A"].width = 38
    for i in range(len(sheets)):
        ws.column_dimensions[col_letter(2 + i)].width = 16
    ws.column_dimensions[col_letter(pivot_total_col(sheets))].width = 18
    ws.freeze_panes = "B5"
    return ws, {"total_asset": total_asset_row, "total_equity": total_equity_row,
                "saldo_awal": saldo_awal_row, "balance_check": balance_check_row,
                "transfer_row": transfer_row, "residual_row": residual_row,
                "sheet": name, "sheets": sheets, "total_col": pivot_total_col(sheets)}


# ---------------------------------------------------------------------------
# Laporan Arus Kas (Cash Flow Statement)
# ---------------------------------------------------------------------------

def write_cash_flow(wb, sheets_last_row, income_ref, balance_ref, period_label):
    name = "Laporan Arus Kas"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    sheets = list(sheets_last_row.keys())
    ws["A1"] = f"LAPORAN ARUS KAS - {period_label.upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Metode langsung, pivot per rekening. Transfer antar rekening sendiri sengaja tidak "
                "dimasukkan karena saling menutup nol (lihat sheet Rekonsiliasi).")
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    r = 4
    write_pivot_header(ws, r, sheets)
    r += 1

    write_pivot_section(ws, r, "ARUS KAS DARI AKTIVITAS OPERASI", sheets)
    r += 1
    op_row = r
    write_pivot_formula_row(
        ws, r, "Laba Bersih Bulan Ini (basis kas)", sheets,
        lambda cl: f"='{income_ref['sheet']}'!{cl}{income_ref['net']}",
    )
    r += 1
    total_op_row = r
    write_pivot_subtotal_row(ws, r, "Kas Bersih dari Operasi", sheets, [op_row, op_row])
    r += 2

    write_pivot_section(ws, r, "ARUS KAS DARI AKTIVITAS PENDANAAN", sheets)
    r += 1
    fin_row = r
    write_pivot_data_row(ws, r, "Modal & Setoran Pemilik (+ Laba Ditahan Bulanan)", sheets,
                          lambda sheet: sumif_modal_one_sheet(sheet, sheets_last_row[sheet]))
    r += 1
    total_fin_row = r
    write_pivot_subtotal_row(ws, r, "Kas Bersih dari Pendanaan", sheets, [fin_row, fin_row])
    r += 2

    net_change_row = r
    write_pivot_formula_row(
        ws, r, "KENAIKAN (PENURUNAN) KAS BERSIH", sheets,
        lambda cl: f"={cl}{total_op_row}+{cl}{total_fin_row}",
        bold=True,
    )
    r += 1
    saldo_awal_row = r
    write_pivot_formula_row(
        ws, r, "Saldo Kas Awal Bulan", sheets,
        lambda cl: f"='{balance_ref['sheet']}'!{cl}{balance_ref['saldo_awal']}",
    )
    r += 1
    saldo_akhir_row = r
    write_pivot_formula_row(
        ws, r, "Saldo Kas Akhir Bulan", sheets,
        lambda cl: f"={cl}{net_change_row}+{cl}{saldo_awal_row}",
        bold=True,
    )
    r += 1
    write_pivot_formula_row(
        ws, r, "Cek vs Total Aset di Neraca", sheets,
        lambda cl: f"={cl}{saldo_akhir_row}-'{balance_ref['sheet']}'!{cl}{balance_ref['total_asset']}",
    )

    ws.column_dimensions["A"].width = 38
    for i in range(len(sheets)):
        ws.column_dimensions[col_letter(2 + i)].width = 16
    ws.column_dimensions[col_letter(pivot_total_col(sheets))].width = 18
    ws.freeze_panes = "B5"
    return ws


# ---------------------------------------------------------------------------
# Diagnostik Keseimbangan - alat telusur kalau Neraca/Arus Kas selisih
# ---------------------------------------------------------------------------

def write_diagnostic_sheet(wb, sheets_last_row, balance_ref, closing_info_by_sheet):
    """Sheet khusus buat menelusuri KENAPA CEK KESEIMBANGAN di Neraca tidak
    nol. Beberapa kemungkinan penyebab yang paling sering terjadi:
    1. Transfer antar rekening yang belum matched (lihat sheet Rekonsiliasi
       bagian 1) - nilainya tidak ikut dihitung di Laba Rugi/Ekuitas, tapi
       tetap mempengaruhi saldo kas riil.
    2. Baris-baris di sheet rekening TIDAK berurutan kronologis terhadap
       kolom Saldo Kumulatif aslinya (kolom F), sehingga saldo hasil
       rekonstruksi (kolom K) menyimpang dari saldo tercatat. Bagian ini
       menunjukkan tepat di rekening mana dan seberapa besar penyimpangan
       itu terjadi, lewat kolom L (Selisih vs Saldo Tercatat) di tiap sheet
       rekening.
    3. Kalau sheet rekening punya blok rekap penutup (Saldo Akhir/Total
       Debit/Total Kredit di baris-baris akhir), nilai itu SUDAH dikeluarkan
       dari rekonstruksi kolom J/K/L (supaya tidak dobel/merusak saldo
       berjalan) dan dipakai di sini sebagai acuan pembanding independen."""
    name = "Diagnostik Keseimbangan"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws["A1"] = "DIAGNOSTIK KESEIMBANGAN - ALAT TELUSUR SELISIH"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        "Dipakai kalau baris 'CEK KESEIMBANGAN' di Neraca tidak nol. "
        "Cek bagian di bawah: transfer yang belum matched (sheet "
        "Rekonsiliasi bagian 1), penyimpangan urutan data per rekening, "
        "dan cross-check terhadap blok rekap penutup sheet (kalau ada)."
    )
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    r = 4
    ws.cell(row=r, column=1, value="1. RINGKASAN KESEIMBANGAN")
    ws.cell(row=r, column=1).font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    r += 1
    total_col_letter = col_letter(balance_ref["total_col"])
    ws.cell(row=r, column=1, value="Total Aset (Neraca, kolom TOTAL)")
    ws.cell(row=r, column=2, value=f"='{balance_ref['sheet']}'!{total_col_letter}{balance_ref['total_asset']}")
    r += 1
    ws.cell(row=r, column=1, value="Total Ekuitas (Neraca, kolom TOTAL)")
    ws.cell(row=r, column=2, value=f"='{balance_ref['sheet']}'!{total_col_letter}{balance_ref['total_equity']}")
    r += 1
    selisih_row = r
    ws.cell(row=r, column=1, value="Selisih (Aset - Ekuitas)")
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=2, value=f"='{balance_ref['sheet']}'!{total_col_letter}{balance_ref['balance_check']}")
    ws.cell(row=r, column=2).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="Kemungkinan sumber #1: transfer belum matched (lihat sheet Rekonsiliasi bagian 1)")
    ws.cell(row=r, column=2, value="=COUNTIF(Rekonsiliasi!$K$6:$K$300,\"Needs manual verification\")")
    ws.cell(row=r, column=3, value="baris - buka sheet Rekonsiliasi, cari warna merah")
    r += 2

    ws.cell(row=r, column=1,
            value="1b. SELISIH BELUM TERJELASKAN PER REKENING (dari sheet Neraca)")
    ws.cell(row=r, column=1).font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    r += 1
    headers1b = ["Rekening", "Selisih Belum Terjelaskan (Rp)"]
    hdr_row1b = r
    for i, h in enumerate(headers1b, start=1):
        ws.cell(row=hdr_row1b, column=i, value=h)
    style_header(ws, hdr_row1b, len(headers1b))
    r += 1
    sheets_list = list(sheets_last_row.keys())
    for i, sheet in enumerate(sheets_list):
        cl = col_letter(2 + i)
        ws.cell(row=r, column=1, value=sheet)
        ws.cell(row=r, column=2, value=f"='{balance_ref['sheet']}'!{cl}{balance_ref['residual_row']}")
        ws.cell(row=r, column=2).number_format = NUMBER_FORMAT
        r += 1
    ws.cell(row=r, column=1, value="TOTAL")
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=2, value=f"='{balance_ref['sheet']}'!{total_col_letter}{balance_ref['residual_row']}")
    ws.cell(row=r, column=2).font = Font(bold=True)
    ws.cell(row=r, column=2).number_format = NUMBER_FORMAT
    r += 2

    ws.cell(row=r, column=1,
            value="1c. CROSS-CHECK SALDO AKHIR RESMI (dari blok rekap penutup sheet, kalau ada)")
    ws.cell(row=r, column=1).font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    r += 1
    headers1c = ["Rekening", "Saldo Akhir Resmi (dari blok penutup)", "Saldo Akhir Rekonstruksi (Kolom K)", "Selisih"]
    hdr_row1c = r
    for i, h in enumerate(headers1c, start=1):
        ws.cell(row=hdr_row1c, column=i, value=h)
    style_header(ws, hdr_row1c, len(headers1c))
    r += 1
    any_closing_found = False
    for sheet, last_row in sheets_last_row.items():
        info = closing_info_by_sheet.get(sheet, {})
        saldo_akhir_resmi = info.get("saldo_akhir")
        ws.cell(row=r, column=1, value=sheet)
        if saldo_akhir_resmi is not None:
            any_closing_found = True
            ws.cell(row=r, column=2, value=saldo_akhir_resmi)
            ws.cell(row=r, column=2).number_format = NUMBER_FORMAT
            ws.cell(row=r, column=3, value=f"='{sheet}'!$K${last_row}")
            ws.cell(row=r, column=3).number_format = NUMBER_FORMAT
            ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
            ws.cell(row=r, column=4).number_format = NUMBER_FORMAT
        else:
            ws.cell(row=r, column=2, value="(tidak ada blok penutup di sheet ini)")
            ws.cell(row=r, column=2).font = Font(italic=True, color="6B7280")
        r += 1
    if not any_closing_found:
        ws.cell(row=r, column=1,
                value="Tidak ada sheet dengan blok rekap penutup (Saldo Akhir/Total Debit/Kredit) terdeteksi.")
        ws.cell(row=r, column=1).font = Font(italic=True, color="6B7280")
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="2. PENYIMPANGAN URUTAN DATA PER REKENING (Kolom K vs Kolom F)")
    ws.cell(row=r, column=1).font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    r += 1
    headers = [
        "Rekening", "Jumlah Baris Menyimpang (>Rp1rb)", "Selisih Maksimum (Rp)",
        "Selisih di Baris Terakhir (Rp)", "Baris Pertama Menyimpang", "Keterangan Baris Itu",
    ]
    hdr_row = r
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hdr_row, column=i, value=h)
    style_header(ws, hdr_row, len(headers))
    r += 1
    for sheet, last_row in sheets_last_row.items():
        rng_l = f"'{sheet}'!$L$2:$L${last_row}"
        rng_a = f"'{sheet}'!$A$2:$A${last_row}"
        rng_c = f"'{sheet}'!$B$2:$B${last_row}"
        ws.cell(row=r, column=1, value=sheet)
        ws.cell(row=r, column=2, value=f"=COUNTIF({rng_l},\">1000\")+COUNTIF({rng_l},\"<-1000\")")
        ws.cell(row=r, column=3, value=f"=SUMPRODUCT(MAX(ABS({rng_l})))")
        ws.cell(row=r, column=3).number_format = NUMBER_FORMAT
        ws.cell(row=r, column=4, value=f"='{sheet}'!$L${last_row}")
        ws.cell(row=r, column=4).number_format = NUMBER_FORMAT
        ws.cell(row=r, column=5,
                value=(f'=IFERROR(INDEX({rng_a},MATCH(TRUE,INDEX(ABS({rng_l})>1000,0),0)),'
                       f'"Tidak ada penyimpangan signifikan")'))
        ws.cell(row=r, column=5).number_format = DATE_FORMAT
        ws.cell(row=r, column=6,
                value=(f'=IFERROR(INDEX({rng_c},MATCH(TRUE,INDEX(ABS({rng_l})>1000,0),0)),"-")'))
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 6))
        r += 1

    r += 1
    ws.cell(row=r, column=1,
            value=("Catatan: penyimpangan berarti baris-baris SEBELUM titik itu di sheet rekening "
                   "tidak tersusun berurutan sesuai kolom Saldo Kumulatif aslinya (kemungkinan input "
                   "manual tidak kronologis, atau digabung per sesi/hari alih-alih per transaksi). "
                   "Saldo akhir bulan (Total Aset di Neraca) tetap dihitung dari hasil rekonstruksi "
                   "(kolom K), bukan dari kolom F yang bolong urutannya - tapi 'Saldo Awal' yang "
                   "dipakai di Neraca mengasumsikan baris pertama tiap sheet adalah titik awal yang "
                   "valid. Kalau baris pertama BUKAN baris Saldo Awal (lihat sheet Rekonsiliasi atau "
                   "cek manual), kemungkinan itu sumber selisihnya.")
            )
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="6B7280")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 60

    widths = [30, 24, 20, 22, 20, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


MONTHS_ID = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]


def coerce_date(value):
    """Ubah nilai tanggal dari sumber manapun (datetime.datetime,
    datetime.date, atau teks format umum) jadi datetime.date. Dipakai
    supaya file lama/preformatted yang kolom tanggalnya bukan objek
    datetime asli (mis. tersimpan sebagai teks) tetap bisa dideteksi
    periodenya. Return None kalau tidak bisa diparse."""
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y",
                    "%d-%b-%y", "%d-%b-%Y", "%d %B %Y", "%m/%d/%Y"):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def detect_period(all_txns):
    """Tebak (tahun, bulan) laporan dari tanggal transaksi yang paling
    sering muncul. Return (None, None) kalau tidak ada tanggal valid."""
    from collections import Counter

    counts = Counter()
    for t in all_txns:
        d = coerce_date(t.date)
        if d is not None:
            counts[(d.year, d.month)] += 1
    if not counts:
        return None, None
    (year, month), _ = counts.most_common(1)[0]
    return year, month


def detect_period_label(all_txns):
    """Label teks 'Bulan Tahun' dari (tahun, bulan) yang terdeteksi, dipakai
    di judul-judul laporan."""
    year, month = detect_period(all_txns)
    if year is None:
        return "PERIODE TIDAK TERDETEKSI"
    return f"{MONTHS_ID[month]} {year}"


def detect_period_end_date(all_txns, year_month_label):
    """Tanggal terakhir transaksi pada bulan yang terdeteksi, dipakai untuk
    judul Neraca ('per tanggal X')."""
    dated = [coerce_date(t.date) for t in all_txns]
    dated = [d for d in dated if d is not None]
    if not dated:
        return ""
    last = max(dated)
    return f"{last.day} {MONTHS_ID[last.month]} {last.year}"


# ---------------------------------------------------------------------------
# Orkestrasi utama
# ---------------------------------------------------------------------------

def reload_shared_rules():
    """Baca ulang shared_rules (Postgres/JSON) dan timpa variabel modul
    yang relevan - dipanggil di awal run_reconciliation() supaya bot yang
    sudah lama jalan (proses long-running di Railway) tetap pakai aturan
    TERBARU dari Postgres tiap file baru diproses, bukan cuma versi yang
    kebetulan aktif saat bot pertama kali start."""
    global CATEGORY_OVERRIDE_RULES, TRANSFER_KEYWORDS, CAPITAL_KEYWORDS
    global DESC_TRANSFER_KEYWORDS, CAPITAL_SELF_TRANSFER_KEYWORDS
    global _PROTECTED_FROM_CATEGORY_OVERRIDE, TIP_MINUS_THRESHOLD, FLIPTECH_FEE_THRESHOLD
    shared_rules._cache = None
    CATEGORY_OVERRIDE_RULES = shared_rules.get("category_override_rules", _DEFAULT_CATEGORY_OVERRIDE_RULES)
    TRANSFER_KEYWORDS = shared_rules.get("transfer_keywords", TRANSFER_KEYWORDS)
    CAPITAL_KEYWORDS = shared_rules.get("capital_keywords", CAPITAL_KEYWORDS)
    DESC_TRANSFER_KEYWORDS = shared_rules.get("desc_transfer_keywords", DESC_TRANSFER_KEYWORDS)
    CAPITAL_SELF_TRANSFER_KEYWORDS = shared_rules.get("capital_self_transfer_keywords", CAPITAL_SELF_TRANSFER_KEYWORDS)
    _PROTECTED_FROM_CATEGORY_OVERRIDE = set(shared_rules.get(
        "protected_from_category_override", sorted(_PROTECTED_FROM_CATEGORY_OVERRIDE)
    ))
    TIP_MINUS_THRESHOLD = shared_rules.get("tip_minus_threshold", TIP_MINUS_THRESHOLD)
    FLIPTECH_FEE_THRESHOLD = shared_rules.get("fliptech_fee_threshold", FLIPTECH_FEE_THRESHOLD)
    _validate_category_override_targets()


def run_reconciliation(input_path, output_path, with_statements=False):
    """Fungsi utama: rekonsiliasi antar rekening (pencocokan transfer,
    deteksi split/merge, indikasi minus/selisih). Ini yang jalan secara
    default setiap ada file masuk.

    with_statements=True akan menambahkan 3 sheet laporan keuangan
    (Laba Rugi, Neraca, Arus Kas) - dibuat opsional supaya proses default
    tetap ringan dan fokus ke rekonsiliasi saja, sesuai kebutuhan.
    """
    reload_shared_rules()
    wb = openpyxl.load_workbook(input_path)
    REPORT_SHEET_NAMES = {
        "Rekonsiliasi", "Laporan Laba Rugi", "Neraca", "Laporan Arus Kas",
        "Diagnostik Keseimbangan",
    }
    account_sheets = [s for s in wb.sheetnames if s not in REPORT_SHEET_NAMES]

    all_txns = []
    all_txns_by_sheet = {}
    sheets_last_row = {}
    opening_rows = {}
    closing_info_by_sheet = {}
    for sname in account_sheets:
        ws = wb[sname]
        split_fliptech_combined_rows(ws)
        lr = last_data_row(ws)
        sheets_last_row[sname] = lr
        add_helper_column(ws, lr)
        txns, closing_info = read_account_sheet(ws)
        add_effective_category_column(ws, txns)
        all_txns.extend(txns)
        all_txns_by_sheet[sname] = txns
        closing_info_by_sheet[sname] = closing_info
        opening = next((t for t in txns if t.is_opening), None)
        opening_rows[sname] = opening.row if opening else 2

    matches, combo_matches = find_matches(all_txns, account_sheets)
    minus_flags = find_minus_flags(all_txns_by_sheet)
    balance_status = compute_balance_status(all_txns_by_sheet)

    # transaksi yang sudah terjelaskan lewat split/merge tidak perlu lagi
    # tampil sebagai "Needs manual verification" biasa di bagian 1
    combo_covered_ids = set()
    for cm in combo_matches:
        combo_covered_ids.add(id(cm["src"]))
        combo_covered_ids.add(id(cm["parts"][0]))
        combo_covered_ids.add(id(cm["parts"][1]))
    matches_section1 = [
        m for m in matches
        if m.dst is not None or id(m.src) not in combo_covered_ids
    ]

    ws_recon, recon_range = write_rekonsiliasi_sheet(wb, matches_section1, combo_matches, minus_flags, balance_status)

    order = list(account_sheets) + ["Rekonsiliasi"]

    period_year, period_month = detect_period(all_txns)
    period_label = detect_period_label(all_txns)
    period_end_label = detect_period_end_date(all_txns, period_label)

    if with_statements:
        income_ws, income_ref = write_income_statement(wb, sheets_last_row, period_label, period_month, recon_range)
        balance_ws, balance_ref = write_balance_sheet(wb, sheets_last_row, opening_rows, income_ref, period_end_label, recon_range)
        write_cash_flow(wb, sheets_last_row, income_ref, balance_ref, period_label)
        write_diagnostic_sheet(wb, sheets_last_row, balance_ref, closing_info_by_sheet)
        order += [income_ref["sheet"], balance_ref["sheet"], "Laporan Arus Kas", "Diagnostik Keseimbangan"]

    # urutan sheet: rekening dulu, lalu laporan
    wb._sheets = [wb[s] for s in order]

    wb.save(output_path)

    # match dengan confidence "Not applicable" (teridentifikasi sebagai
    # cicilan pinjaman via Fliptech, BUKAN transfer internal yang genuinely
    # belum ketemu pasangannya) tidak dihitung sebagai unmatched - sudah
    # ada penjelasannya, tidak perlu verifikasi manual lagi
    n_transfer_unmatched = sum(
        1 for m in matches_section1 if m.dst is None and m.confidence != "Not applicable (bukan transfer internal)"
    )
    n_transfer_not_applicable = sum(
        1 for m in matches_section1 if m.confidence == "Not applicable (bukan transfer internal)"
    )
    n_balance_issues = sum(1 for s in balance_status.values() if abs(s["selisih"]) >= 1)
    no_issues = n_transfer_unmatched == 0 and len(minus_flags) == 0 and n_balance_issues == 0

    summary = {
        "n_transfer_high": sum(1 for m in matches if m.confidence == "High"),
        "n_transfer_medium": sum(1 for m in matches if m.confidence == "Medium"),
        "n_transfer_low": sum(1 for m in matches if m.confidence == "Low"),
        "n_transfer_split_merge": len(combo_matches),
        "n_transfer_unmatched": n_transfer_unmatched,
        "n_transfer_not_applicable": n_transfer_not_applicable,
        "n_minus_flags": len(minus_flags),
        "n_balance_issues": n_balance_issues,
        "with_statements": with_statements,
        "period_label": period_label,
        "no_issues": no_issues,
    }
    return summary


if __name__ == "__main__":
    import sys
    inp = sys.argv[1] if len(sys.argv) > 1 else "Recon_Januari_2025.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "Recon_Januari_2025_HASIL.xlsx"
    with_stmt = "--laporan" in sys.argv
    s = run_reconciliation(inp, out, with_statements=with_stmt)
    print(s)
