"""Audit silang mesin kasir (POS) vs rekap keuangan bank/kas - FITUR
MANUAL, tidak jalan otomatis, harus di-trigger via /auditkasir (bot.py)
atau dipanggil langsung sebagai skrip.

KENAPA AGREGASI BULANAN/HARIAN GABUNGAN, BUKAN PENCOCOKAN PER TRANSAKSI
ATAU DIPISAH PER REKENING BANK:
1. Data bank (Rekap) tidak menyimpan referensi transaksi POS (No
   Transaksi kasir) - cuma "Jam ...; Teller/User ID" dan kadang kode
   settlement (QRISOnUs/QRISOffUs/MID). Tidak ada kunci unik untuk
   mencocokkan satu transaksi kasir ke satu baris bank secara presisi.
2. QRIS/kartu settle ke rekening H+1 atau lebih lambat, dan TIDAK SELALU
   HARIAN (bisa beberapa hari terkumpul jadi satu setoran) - dikonfirmasi
   user.
3. PENTING (dikonfirmasi user): SEMUA transaksi berkategori 'Penjualan'
   yang masuk ke rekening BANK MANAPUN (bukan Kas-Buku) otomatis
   dianggap penjualan via QRIS - TIDAK dipisah/dicocokkan per rekening
   bank tertentu (mis. BRI-507 khusus utk QRIS BRI, BCA-887 khusus utk
   QRIS BCA), karena settlement bisa masuk ke rekening manapun, tidak
   selalu sesuai provider QRIS yang dipakai pembeli. Jadi audit ini
   membandingkan 2 pool saja: Cash (POS) vs Kas-Buku (bank), dan
   Non-Tunai gabungan (QRIS BRI+QRIS BCA+Kartu dari POS) vs GABUNGAN
   'Penjualan' di SEMUA rekening bank selain Kas-Buku.
4. 'Penjualan Langsung' - transaksi bank berkategori 'Penjualan' yang
   ditandai user (lewat kata 'Penjualan Langsung' di Keterangan/Objek/
   dst) sebagai penjualan yang diverifikasi manual, BUKAN dari POS -
   dikecualikan dari pool Non-Tunai (tidak akan pernah ada padanannya di
   data POS), ditampilkan terpisah sebagai informasi saja.

KETERBATASAN yang harus disadari user:
- Beberapa hari di awal/akhir bulan WAJAR meleset ke bulan
  sebelum/sesudahnya karena jeda settlement - laporan ini menandai hari
  pertama/terakhir tiap bulan sebagai konteks, bukan tuduhan kesalahan.
"""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

import reconcile as rc

HEADER_FILL = "1F2937"
ALT_FILL = "F3F4F6"
GOOD_FILL = "DCFCE7"
WARN_FILL = "FEF3C7"
BAD_FILL = "FEE2E2"
BORDER = Border(*[Side(style="thin", color="D1D5DB")] * 4)

METODE_KATEGORI = {
    "cash": "Cash",
    "bank transfer - qris bri": "QRIS BRI",
    "bank transfer - qris bca": "QRIS BCA",
    "kartu debit/kredit - bri": "Kartu BRI",
}


class KasirAuditError(Exception):
    """Dilempar kalau file input tidak sesuai format yang diharapkan -
    pesannya dirancang untuk langsung ditunjukkan ke user di bot."""


def normalize_metode(metode_raw):
    """Petakan teks 'Metode Pembayaran' dari POS ke kategori standar.
    Metode yang tidak dikenali (mis. baru dari POS) dikembalikan apa
    adanya dengan prefix 'Lainnya:' supaya tetap kelihatan di laporan,
    bukan diam-diam diabaikan."""
    key = str(metode_raw or "").strip().lower()
    if key in ("", "-"):
        return "Belum Bayar"
    return METODE_KATEGORI.get(key, f"Lainnya: {metode_raw}")


def _parse_pos_datetime(s):
    s = str(s or "").strip()
    if not s:
        return None
    for fmt in ("%d-%m-%Y %H:%M:%S", "%d-%m-%Y", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


_INTERPRETASI_KATEGORI_PREFIXES = ["Cash", "QRIS BRI", "QRIS BCA", "Kartu BRI", "Kartu"]
_ESTIMASI_SETTLE_RE = re.compile(r"estimasi settle\s*:\s*(\d{1,2}/\d{1,2}/\d{4})", re.IGNORECASE)


def _kategori_dari_nama_sheet_interpretasi(sheet_title):
    """'Cash Desember 2024' -> 'Cash'; 'QRIS BRI Desember 2024' -> 'QRIS BRI'.
    Return None kalau nama sheet tidak diawali salah satu kategori yang
    dikenal (bukan sheet Interpretasi Penjualan). Beda dari
    normalize_metode() - di sini nama sheet SUDAH JADI nama kategori
    tujuan langsung, tidak perlu dipetakan dari teks metode POS mentah."""
    for prefix in _INTERPRETASI_KATEGORI_PREFIXES:
        if sheet_title.lower().startswith(prefix.lower() + " "):
            return "Kartu BRI" if prefix == "Kartu" else prefix
    return None


def looks_like_interpretasi_file(path):
    """True kalau file ini format 'Interpretasi Penjualan' (per-transaksi,
    struktur standar rekonsiliasi, sheet dinamai per kategori metode
    bayar) - BEDA dari export POS mentah (parse_pos_sales) dan BEDA dari
    file Rekap (banyak sheet rekening bank)."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception:
        return False
    for sn in wb.sheetnames:
        if _kategori_dari_nama_sheet_interpretasi(sn) is not None:
            header = [wb[sn].cell(row=1, column=c).value for c in range(1, 9)]
            if header == _EXPECTED_HEADER:
                return True
    return False


def parse_interpretasi_penjualan(path):
    """Baca file 'Interpretasi Penjualan' - beda dari export POS mentah:
    per-transaksi SUDAH dalam format standar rekonsiliasi (Tanggal/
    Keterangan/Kategori/dst), tiap sheet mewakili SATU kategori metode
    bayar ('Cash Desember 2024', 'QRIS BRI Desember 2024', dst), dan
    untuk metode non-tunai ada 'Estimasi settle: DD/MM/YYYY' eksplisit
    di Keterangan Tambahan - TIDAK PERLU ASUMSI H+1, tanggal settle
    sudah dihitung langsung per transaksi.

    Refund muncul sebagai baris TERPISAH (Keterangan='Refund', Debit
    negatif) - bukan kolom refund seperti di export POS mentah.

    Return list of dict kompatibel dengan struktur parse_pos_sales()
    (bisa dipakai bareng di aggregate_pos_by_month/aggregate_pos_by_day),
    PLUS field 'tanggal_settle' untuk rekonsiliasi presisi harian."""
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    found_any = False
    for sn in wb.sheetnames:
        metode = _kategori_dari_nama_sheet_interpretasi(sn)
        if metode is None:
            continue
        ws = wb[sn]
        header = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        if header != _EXPECTED_HEADER:
            continue
        found_any = True
        txns, _ = rc.read_account_sheet(ws)
        for t in txns:
            if t.is_opening:
                continue
            tgl = rc.coerce_date(t.date)
            if tgl is None:
                continue
            tgl_dt = datetime(tgl.year, tgl.month, tgl.day)
            m = _ESTIMASI_SETTLE_RE.search(str(t.ket or ""))
            if m:
                tgl_settle = _parse_pos_datetime(m.group(1))
            else:
                tgl_settle = tgl_dt  # Cash: settle di hari yang sama
            is_refund = str(t.desc or "").strip().lower() == "refund"
            rows.append({
                "no_transaksi": t.ket,
                "waktu_order": None,
                "waktu_bayar": None,
                "tanggal": tgl_dt,
                "tanggal_settle": tgl_settle,
                "total": 0 if is_refund else t.nominal,
                "metode_raw": sn,
                "metode": metode,
                "status": "Lunas",
                "kasir": None,
                "refund_tanggal": tgl_dt if is_refund else None,
                "refund_jumlah": abs(t.nominal) if is_refund else 0,
                "refund_metode": metode if is_refund else None,
                "refund_alasan": None,
            })
    if not found_any:
        raise KasirAuditError(
            f"File '{path}' tidak dikenali sebagai format Interpretasi Penjualan - tidak ada sheet "
            "dengan nama kategori metode bayar (Cash/QRIS BRI/QRIS BCA/Kartu BRI) berstruktur standar."
        )
    return rows


def aggregate_interp_by_settle_date(interp_txns):
    """Kelompokkan transaksi Interpretasi Penjualan per (tanggal SETTLE
    aktual, kategori metode) - dipakai rekonsiliasi presisi (bukan
    asumsi H+1 seragam, tapi tanggal settle yang sudah dihitung per
    transaksi)."""
    agg = {}
    for t in interp_txns:
        tgl_settle = t["tanggal_settle"]
        if tgl_settle is None:
            continue
        key = (tgl_settle.date(), t["metode"])
        a = agg.setdefault(key, {"kotor": 0, "refund": 0, "n": 0})
        a["kotor"] += t["total"]
        a["refund"] += t["refund_jumlah"]
        a["n"] += 1
    return agg


def parse_pos_sales(path):
    """Baca file 'Detail Penjualan' (export mesin kasir/POS). Header
    kolom ada di baris 13 (baris 1-12 metadata ringkasan). Return list of
    dict per transaksi. Melempar KasirAuditError kalau header tidak
    cocok format yang diharapkan (supaya gagal jelas, bukan salah baca
    diam-diam)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    for r in range(1, min(ws.max_row, 20) + 1):
        if str(ws.cell(row=r, column=1).value or "").strip() == "No Transaksi":
            header_row = r
            break
    if header_row is None:
        raise KasirAuditError(
            f"File '{path}' tidak dikenali sebagai export Detail Penjualan POS - "
            "tidak ketemu baris header 'No Transaksi' di 20 baris pertama."
        )
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        no_transaksi = ws.cell(row=r, column=1).value
        if not no_transaksi:
            continue
        waktu_bayar = ws.cell(row=r, column=3).value
        tgl = _parse_pos_datetime(waktu_bayar) or _parse_pos_datetime(ws.cell(row=r, column=2).value)
        rows.append({
            "no_transaksi": no_transaksi,
            "waktu_order": ws.cell(row=r, column=2).value,
            "waktu_bayar": waktu_bayar,
            "tanggal": tgl,
            "total": ws.cell(row=r, column=6).value or 0,
            "metode_raw": ws.cell(row=r, column=7).value,
            "metode": normalize_metode(ws.cell(row=r, column=7).value),
            "status": ws.cell(row=r, column=9).value,
            "kasir": ws.cell(row=r, column=11).value,
            "refund_tanggal": ws.cell(row=r, column=12).value,
            "refund_jumlah": ws.cell(row=r, column=13).value or 0,
            "refund_metode": ws.cell(row=r, column=14).value,
            "refund_alasan": ws.cell(row=r, column=15).value,
        })
    if not rows:
        raise KasirAuditError(f"File '{path}': tidak ada baris transaksi ditemukan setelah header.")
    return rows


def _bulan_label_dari_tanggal(tgl):
    if tgl is None:
        return None
    return f"{rc.MONTHS_ID[tgl.month]} {tgl.year}"


def aggregate_pos_by_month(all_pos_txns):
    """Kelompokkan transaksi POS per (bulan, kategori metode) - hitung
    total kotor, total refund, bersih (kotor - refund), jumlah transaksi.
    Transaksi 'Belum Bayar' (status Belum Lunas) DIKECUALIKAN dari total
    (uangnya belum benar-benar berpindah tangan)."""
    agg = {}
    for t in all_pos_txns:
        bulan = _bulan_label_dari_tanggal(t["tanggal"])
        if bulan is None or t["status"] == "Belum Lunas":
            continue
        key = (bulan, t["metode"])
        a = agg.setdefault(key, {"kotor": 0, "refund": 0, "n": 0})
        a["kotor"] += t["total"]
        a["refund"] += t["refund_jumlah"]
        a["n"] += 1
    return agg


def aggregate_pos_by_day(all_pos_txns):
    """Kelompokkan transaksi POS per (tanggal, kategori metode) - dipakai
    untuk rekonsiliasi harian dengan asumsi H+1 (QRIS/kartu baru masuk
    rekening hari berikutnya). Refund diperlakukan sama seperti agregasi
    bulanan: dikurangi dari total kotor tanggal transaksi ASLI (bukan
    tanggal refund) - refund adalah koreksi atas penjualan hari itu."""
    agg = {}
    for t in all_pos_txns:
        tgl = t["tanggal"]
        if tgl is None or t["status"] == "Belum Lunas":
            continue
        key = (tgl.date(), t["metode"])
        a = agg.setdefault(key, {"kotor": 0, "refund": 0, "n": 0})
        a["kotor"] += t["total"]
        a["refund"] += t["refund_jumlah"]
        a["n"] += 1
    return agg


_EXPECTED_HEADER = ["Tanggal", "Keterangan Transaksi", "Kategori Transaksi", "Debit", "Kredit",
                    "Saldo Kumulatif", "Subjek Transaksi", "Objek Transaksi"]


def _looks_like_account_sheet(ws):
    header = [ws.cell(row=1, column=c).value for c in range(1, len(_EXPECTED_HEADER) + 1)]
    return header == _EXPECTED_HEADER


def aggregate_bank_penjualan(rekap_paths):
    """Untuk tiap file Rekap (rekonsiliasi bulanan biasa), jumlahkan
    transaksi berkategori efektif 'Penjualan' per (bulan, kelompok).
    Pakai reconcile.py langsung supaya konsisten dengan logika
    kategorisasi yang sama dipakai rekonsiliasi normal.

    PENTING (dikonfirmasi user): SEMUA 'Penjualan' yang masuk ke rekening
    BANK manapun (bukan Kas-Buku) otomatis dianggap penjualan via QRIS -
    TIDAK dipisah per rekening bank (mis. BRI-507 vs BCA-887), karena
    settlement QRIS bisa masuk ke rekening manapun, tidak selalu sesuai
    provider QRIS yang dipakai pembeli. Jadi dikelompokkan jadi 2 pool
    saja: 'Kas-Buku' (penjualan tunai) dan 'Non-Tunai (Bank)' (gabungan
    SEMUA rekening bank lain, mewakili seluruh penjualan QRIS/kartu).

    'Penjualan Langsung' (kode di Keterangan/Objek/dst yang ditandai user
    saat verifikasi manual) DIKECUALIKAN dari kedua pool di atas -
    ini penjualan yang TIDAK lewat POS sama sekali (makanya kategorinya
    tetap 'Penjualan' apa adanya, tidak diubah), jadi tidak akan pernah
    match dengan data POS manapun. Dihitung terpisah supaya kelihatan di
    laporan, bukan dianggap sebagai selisih yang mencurigakan."""
    agg = {}
    daily = {}  # (kelompok, tanggal) -> total, buat cek settlement lintas bulan
    for path in rekap_paths:
        wb = openpyxl.load_workbook(path)
        account_sheets = [s for s in wb.sheetnames if _looks_like_account_sheet(wb[s])]
        if not account_sheets:
            raise KasirAuditError(
                f"File '{path}' tidak punya sheet rekening yang dikenali - pastikan ini file "
                "rekonsiliasi bulanan (hasil /start atau upload biasa ke bot), bukan laporan "
                "kuartalan/tahunan gabungan."
            )
        for sn in account_sheets:
            rc.split_fliptech_combined_rows(wb[sn])
            txns, _ = rc.read_account_sheet(wb[sn])
            base = sn.rsplit(" ", 2)[0]
            for t in txns:
                if t.effective_kategori != "Penjualan":
                    continue
                tgl = rc.coerce_date(t.date)
                bulan = _bulan_label_dari_tanggal(tgl) if tgl else None
                if bulan is None:
                    continue
                text = f"{t.desc or ''} {t.ket or ''} {t.objek or ''} {t.subjek or ''}".lower()
                if "penjualan langsung" in text:
                    kelompok = "Penjualan Langsung (manual, bukan dari POS)"
                elif base == "Kas-Buku":
                    kelompok = "Kas-Buku"
                else:
                    kelompok = "Non-Tunai (Bank)"
                key = (bulan, kelompok)
                agg[key] = agg.get(key, 0) + t.nominal
                dkey = (kelompok, tgl)
                daily[dkey] = daily.get(dkey, 0) + t.nominal
    return agg, daily

def _write_harian_sheet(wb, title, pos_daily, bank_daily, bank_rekening, all_dates, shift_hari,
                          catatan_header):
    """Satu sheet rekonsiliasi harian untuk SATU kategori metode -
    bandingkan total POS tanggal T dengan total Bank tanggal T+shift_hari
    (shift_hari=0 untuk Cash/tanpa penundaan, 1 untuk asumsi H+1 QRIS/
    kartu). Baris dengan selisih signifikan (>Rp5.000 ATAU >5% dari nilai
    POS, mana yang lebih longgar) ditandai merah - ini yang perlu
    ditelusuri (kemungkinan settlement tidak terjadi hari itu/terlambat
    lebih dari H+1, refund, atau salah metode)."""
    from datetime import timedelta
    ws = wb.create_sheet(title)
    ws["A1"] = title.upper()
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = catatan_header
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 40

    r = 4
    headers = ["Tanggal POS", "POS - Kotor", "POS - Refund", "POS - Bersih",
               f"Tanggal Bank (+{shift_hari} hari)", f"Bank ({bank_rekening})", "Selisih", "Status"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.border = BORDER
    r += 1

    running_pos = 0
    running_bank = 0
    n_selisih = 0
    for tgl in all_dates:
        pos = pos_daily.get(tgl)
        tgl_bank = tgl + timedelta(days=shift_hari)
        bank_val = bank_daily.get((bank_rekening, tgl_bank))
        if pos is None and bank_val is None:
            continue
        kotor = pos["kotor"] if pos else 0
        refund = pos["refund"] if pos else 0
        bersih = kotor - refund
        bank_val = bank_val or 0
        selisih = bersih - bank_val
        running_pos += bersih
        running_bank += bank_val
        toleransi = max(5000, bersih * 0.05)
        if abs(selisih) <= toleransi:
            status, fill = "OK", GOOD_FILL
        else:
            status, fill = "SELISIH - cek manual", BAD_FILL
            n_selisih += 1
        ws.cell(row=r, column=1, value=tgl.strftime("%d/%m/%Y"))
        ws.cell(row=r, column=2, value=kotor)
        ws.cell(row=r, column=3, value=refund)
        ws.cell(row=r, column=4, value=bersih)
        ws.cell(row=r, column=5, value=tgl_bank.strftime("%d/%m/%Y"))
        ws.cell(row=r, column=6, value=bank_val)
        ws.cell(row=r, column=7, value=selisih)
        ws.cell(row=r, column=8, value=status)
        for c in (2, 3, 4, 6, 7):
            ws.cell(row=r, column=c).number_format = rc.NUMBER_FORMAT
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=8).fill = PatternFill("solid", fgColor=fill)
        ws.cell(row=r, column=8).font = Font(bold=True)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="TOTAL PERIODE")
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=4, value=running_pos)
    ws.cell(row=r, column=6, value=running_bank)
    ws.cell(row=r, column=7, value=running_pos - running_bank)
    for c in (4, 6, 7):
        ws.cell(row=r, column=c).number_format = rc.NUMBER_FORMAT
        ws.cell(row=r, column=c).font = Font(bold=True)
    r += 1
    selisih_total = running_pos - running_bank
    if abs(selisih_total) <= max(10000, running_pos * 0.02):
        ws.cell(row=r, column=1, value=(
            "Total periode NYAMBUNG (selisih kecil) - artinya hari-hari SELISIH di atas kemungkinan "
            "besar cuma jeda settlement yang lebih lambat dari H+1 (menumpuk lalu masuk sekaligus di "
            "hari lain), bukan uang yang benar-benar hilang."
        ))
        ws.cell(row=r, column=1).font = Font(italic=True, color="15803D")
    else:
        rp_selisih = f"{abs(selisih_total):,.0f}".replace(",", ".")
        ws.cell(row=r, column=1, value=(
            f"Total periode TIDAK NYAMBUNG (selisih Rp{rp_selisih}) - ini BUKAN cuma "
            "soal jeda settlement, ada uang yang genuinely tidak tercatat di salah satu sisi. Perlu "
            "ditelusuri manual."
        ))
        ws.cell(row=r, column=1).font = Font(italic=True, bold=True, color="B91C1C")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 30

    widths = [14, 15, 13, 15, 18, 16, 14, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    return n_selisih


def run_kasir_audit(pos_paths, rekap_paths, output_path, interp_paths=None):
    interp_paths = interp_paths or []
    all_pos = []
    for p in pos_paths:
        all_pos.extend(parse_pos_sales(p))
    all_interp = []
    for p in interp_paths:
        all_interp.extend(parse_interpretasi_penjualan(p))
    # Interpretasi Penjualan (kalau ada) dipakai UNTUK Ringkasan Bulanan
    # juga - datanya per-transaksi persis sama, cuma sudah diinterpretasi
    # dengan tanggal settle eksplisit, jadi lebih presisi daripada
    # export POS mentah kalau keduanya tersedia untuk bulan yang sama.
    all_pos_for_monthly = all_pos + all_interp
    pos_agg = aggregate_pos_by_month(all_pos_for_monthly)
    pos_daily_raw = aggregate_pos_by_day(all_pos_for_monthly)
    bank_agg, bank_daily = aggregate_bank_penjualan(rekap_paths)

    bulan_list = sorted({b for b, _ in pos_agg} | {b for b, _ in bank_agg},
                         key=lambda lbl: (lbl.split()[1], rc.MONTHS_ID.index(lbl.split()[0])))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Ringkasan Audit Kasir")
    ws["A1"] = "AUDIT SILANG MESIN KASIR vs REKAP KEUANGAN"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        "Perbandingan TOTAL BULANAN per metode bayar (bukan per transaksi/harian) - QRIS/kartu settle "
        "ke rekening beberapa hari kemudian dan tidak selalu harian, jadi pencocokan bulanan lebih andal "
        "daripada harian. Selisih kecil relatif terhadap volume bulan itu WAJAR (jeda settlement lintas "
        "bulan) - selisih besar perlu ditelusuri manual (kemungkinan salah kategori/metode, refund yang "
        "tidak konsisten, atau setoran yang belum masuk)."
    )
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 55

    r = 4
    headers = ["Bulan", "Metode", "POS - Kotor", "POS - Refund", "POS - Bersih",
               "Bank Tercatat (Penjualan)", "Selisih", "Selisih (%)", "Catatan"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

    kelompok_order = ["Kas-Buku", "Non-Tunai (Bank)"]
    kelompok_pos_kategori = {
        "Kas-Buku": ["Cash"],
        "Non-Tunai (Bank)": ["QRIS BRI", "QRIS BCA", "Kartu BRI"],
    }
    kelompok_label = {"Kas-Buku": "Cash", "Non-Tunai (Bank)": "Non-Tunai (Bank)"}
    row_i = 0
    for bulan in bulan_list:
        for kelompok in kelompok_order:
            kotor = refund = n = 0
            for m in kelompok_pos_kategori[kelompok]:
                pos = pos_agg.get((bulan, m))
                if pos:
                    kotor += pos["kotor"]
                    refund += pos["refund"]
                    n += pos["n"]
            bersih = kotor - refund
            bank_val = bank_agg.get((bulan, kelompok))
            if n == 0 and bank_val is None:
                continue
            selisih = (bersih - bank_val) if bank_val is not None else None
            pct = (abs(selisih) / bersih * 100) if (selisih is not None and bersih) else None
            if bank_val is None:
                catatan = f"Tidak ada data Rekap untuk kelompok '{kelompok}' bulan ini."
                fill = WARN_FILL
            elif pct is None:
                catatan = "-"
                fill = ALT_FILL if row_i % 2 else None
            elif pct < 5:
                catatan = "Selisih kecil, kemungkinan besar jeda settlement normal."
                fill = GOOD_FILL
            elif pct < 15:
                catatan = "Selisih cukup besar - cek transaksi awal/akhir bulan (kemungkinan settlement lintas bulan) atau refund."
                fill = WARN_FILL
            else:
                catatan = "SELISIH BESAR - kemungkinan salah kategori/metode, refund tidak tercatat rapi, atau setoran belum masuk. Perlu ditelusuri manual."
                fill = BAD_FILL
            ws.cell(row=r, column=1, value=bulan)
            ws.cell(row=r, column=2, value=f"{kelompok_label[kelompok]} ({n} transaksi)")
            ws.cell(row=r, column=3, value=kotor)
            ws.cell(row=r, column=4, value=refund)
            ws.cell(row=r, column=5, value=bersih)
            ws.cell(row=r, column=6, value=bank_val if bank_val is not None else "-")
            ws.cell(row=r, column=7, value=selisih if selisih is not None else "-")
            ws.cell(row=r, column=8, value=f"{pct:.1f}%" if pct is not None else "-")
            ws.cell(row=r, column=9, value=catatan)
            for c in (3, 4, 5, 6, 7):
                ws.cell(row=r, column=c).number_format = rc.NUMBER_FORMAT
            for c in range(1, 10):
                ws.cell(row=r, column=c).border = BORDER
                if fill:
                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=fill)
            row_i += 1
            r += 1

    # Penjualan Langsung - informational saja, TIDAK dibandingkan ke POS
    # (memang tidak akan pernah ada di data POS, ini penjualan di luar
    # POS yang diverifikasi manual dan sengaja tetap berkategori
    # 'Penjualan' apa adanya)
    ada_langsung = any(k == "Penjualan Langsung (manual, bukan dari POS)" for _, k in bank_agg)
    if ada_langsung:
        r += 1
        ws.cell(row=r, column=1, value="Penjualan Langsung (manual, verifikasi terpisah, bukan dari POS - informasi saja):")
        ws.cell(row=r, column=1).font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        r += 1
        for bulan in bulan_list:
            val = bank_agg.get((bulan, "Penjualan Langsung (manual, bukan dari POS)"))
            if val is None:
                continue
            ws.cell(row=r, column=1, value=bulan)
            ws.cell(row=r, column=2, value="Penjualan Langsung")
            ws.cell(row=r, column=6, value=val)
            ws.cell(row=r, column=9, value="Tidak dibandingkan ke POS - memang bukan transaksi dari POS.")
            ws.cell(row=r, column=6).number_format = rc.NUMBER_FORMAT
            for c in range(1, 10):
                ws.cell(row=r, column=c).border = BORDER
            r += 1

    widths = [16, 26, 15, 14, 15, 18, 14, 12, 45]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # Sheet detail refund
    ws2 = wb.create_sheet("Detail Refund")
    ws2["A1"] = "DETAIL TRANSAKSI REFUND"
    ws2["A1"].font = Font(bold=True, size=13)
    r2 = 3
    h2 = ["No Transaksi", "Tanggal Bayar", "Metode", "Total", "Tanggal Refund", "Jumlah Refund", "Metode Refund", "Alasan", "Kasir"]
    for i, h in enumerate(h2, start=1):
        c = ws2.cell(row=r2, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.border = BORDER
    r2 += 1
    refund_txns = sorted(
        (t for t in all_pos_for_monthly if t["refund_jumlah"] and t["refund_jumlah"] != 0),
        key=lambda t: t["tanggal"] or datetime.min,
    )
    for t in refund_txns:
        ws2.cell(row=r2, column=1, value=t["no_transaksi"])
        ws2.cell(row=r2, column=2, value=t["waktu_bayar"])
        ws2.cell(row=r2, column=3, value=t["metode"])
        ws2.cell(row=r2, column=4, value=t["total"])
        ws2.cell(row=r2, column=5, value=t["refund_tanggal"])
        ws2.cell(row=r2, column=6, value=t["refund_jumlah"])
        ws2.cell(row=r2, column=7, value=t["refund_metode"])
        ws2.cell(row=r2, column=8, value=t["refund_alasan"])
        ws2.cell(row=r2, column=9, value=t["kasir"])
        for c in (4, 6):
            ws2.cell(row=r2, column=c).number_format = rc.NUMBER_FORMAT
        for c in range(1, 10):
            ws2.cell(row=r2, column=c).border = BORDER
        r2 += 1
    if not refund_txns:
        ws2.cell(row=r2, column=1, value="(tidak ada transaksi refund di periode ini)")
        ws2.cell(row=r2, column=1).font = Font(italic=True, color="6B7280")
    widths2 = [20, 18, 22, 14, 18, 14, 16, 18, 14]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Sheet transaksi belum lunas (informational)
    ws3 = wb.create_sheet("Belum Lunas (dikecualikan)")
    ws3["A1"] = "TRANSAKSI 'BELUM LUNAS' - DIKECUALIKAN DARI AUDIT (uang belum berpindah tangan)"
    ws3["A1"].font = Font(bold=True, size=13)
    r3 = 3
    for i, h in enumerate(["No Transaksi", "Waktu Order", "Total", "Kasir"], start=1):
        c = ws3.cell(row=r3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.border = BORDER
    r3 += 1
    belum_lunas = [t for t in all_pos_for_monthly if t["status"] == "Belum Lunas"]
    for t in belum_lunas:
        ws3.cell(row=r3, column=1, value=t["no_transaksi"])
        ws3.cell(row=r3, column=2, value=t["waktu_order"])
        ws3.cell(row=r3, column=3, value=t["total"])
        ws3.cell(row=r3, column=4, value=t["kasir"])
        ws3.cell(row=r3, column=3).number_format = rc.NUMBER_FORMAT
        for c in range(1, 5):
            ws3.cell(row=r3, column=c).border = BORDER
        r3 += 1
    if not belum_lunas:
        ws3.cell(row=r3, column=1, value="(tidak ada transaksi belum lunas di periode ini)")
        ws3.cell(row=r3, column=1).font = Font(italic=True, color="6B7280")
    for i, w in enumerate([20, 20, 14, 14], start=1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Rekonsiliasi harian - Cash (tanpa penundaan) vs Non-Tunai gabungan
    # (asumsi H+1, semua rekening bank selain Kas-Buku digabung jadi
    # satu pool "Non-Tunai (Bank)" - lihat catatan di aggregate_bank_penjualan)
    all_dates_set = {tgl for (tgl, _) in pos_daily_raw} | {tgl for (_, tgl) in bank_daily}
    all_dates_sorted = sorted(all_dates_set)

    cash_daily = {tgl: v for (tgl, m), v in pos_daily_raw.items() if m == "Cash"}
    non_tunai_daily = {}
    for (tgl, m), v in pos_daily_raw.items():
        if m in ("QRIS BRI", "QRIS BCA", "Kartu BRI"):
            a = non_tunai_daily.setdefault(tgl, {"kotor": 0, "refund": 0, "n": 0})
            a["kotor"] += v["kotor"]
            a["refund"] += v["refund"]
            a["n"] += v["n"]

    n_selisih_cash = _write_harian_sheet(
        wb, "Harian - Cash", cash_daily, bank_daily, "Kas-Buku", all_dates_sorted, 0,
        "Cash tidak ada penundaan settlement - dibandingkan tanggal yang SAMA persis."
    )
    n_selisih_non_tunai = _write_harian_sheet(
        wb, "Harian - Non-Tunai (H+1)", non_tunai_daily, bank_daily, "Non-Tunai (Bank)", all_dates_sorted, 1,
        "Asumsi QRIS/kartu baru masuk rekening bank SEHARI SETELAH tanggal transaksi POS (H+1). Semua "
        "rekening bank selain Kas-Buku digabung jadi satu pool 'Non-Tunai (Bank)' - penjualan via QRIS/"
        "kartu bisa settle ke rekening manapun, tidak selalu sesuai provider yang dipakai pembeli."
    )

    # Kalau ada file Interpretasi Penjualan (tanggal settle eksplisit per
    # transaksi, bukan asumsi H+1 seragam) - bangun sheet presisi
    # TAMBAHAN yang lebih akurat, memakai tanggal settle yang sudah
    # dihitung langsung, bukan tebakan H+1.
    n_selisih_presisi = {}
    if all_interp:
        interp_settle_agg = aggregate_interp_by_settle_date(all_interp)
        settle_dates_sorted = sorted({d for (d, _) in interp_settle_agg} | {d for (_, d) in bank_daily})
        presisi_cash = {d: v for (d, m), v in interp_settle_agg.items() if m == "Cash"}
        presisi_non_tunai = {}
        for (d, m), v in interp_settle_agg.items():
            if m in ("QRIS BRI", "QRIS BCA", "Kartu BRI"):
                a = presisi_non_tunai.setdefault(d, {"kotor": 0, "refund": 0, "n": 0})
                a["kotor"] += v["kotor"]
                a["refund"] += v["refund"]
                a["n"] += v["n"]
        if presisi_cash:
            n_selisih_presisi["Cash"] = _write_harian_sheet(
                wb, "Presisi - Cash", presisi_cash, bank_daily, "Kas-Buku", settle_dates_sorted, 0,
                "Dari file Interpretasi Penjualan - tanggal settle SUDAH dihitung eksplisit per transaksi."
            )
        if presisi_non_tunai:
            n_selisih_presisi["Non-Tunai"] = _write_harian_sheet(
                wb, "Presisi - Non-Tunai", presisi_non_tunai, bank_daily, "Non-Tunai (Bank)", settle_dates_sorted, 0,
                "Dari file Interpretasi Penjualan - tanggal settle SUDAH dihitung eksplisit per transaksi "
                "(bukan asumsi H+1 seragam), jadi 'Tanggal Bank' di sini SAMA dengan tanggal settle yang "
                "tertera, tidak ada pergeseran tambahan. Semua rekening bank selain Kas-Buku digabung jadi "
                "satu pool 'Non-Tunai (Bank)'."
            )

    wb.save(output_path)
    n_refund_total = sum(t["refund_jumlah"] for t in all_pos_for_monthly if t["refund_jumlah"])
    return {
        "n_bulan": len(bulan_list),
        "n_transaksi_pos": len(all_pos_for_monthly),
        "n_refund": len(refund_txns),
        "total_refund": n_refund_total,
        "n_belum_lunas": len(belum_lunas),
        "n_hari_selisih_cash": n_selisih_cash,
        "n_hari_selisih_non_tunai": n_selisih_non_tunai,
        "n_selisih_presisi": n_selisih_presisi,
    }


if __name__ == "__main__":
    import sys
    args = sys.argv[1:]
    pos_paths = [a for a in args if "detail" in a.lower() or "sales" in a.lower()]
    rekap_paths = [a for a in args if a not in pos_paths and not a.endswith("_out.xlsx")]
    out = "kasir_audit_output.xlsx"
    print(run_kasir_audit(pos_paths, rekap_paths, out))
