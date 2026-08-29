"""
quarterly.py
Laporan keuangan kuartalan: Laba Rugi, Neraca, Arus Kas, Roster Gaji 3
Bulan, dan Analisis & Tren - dibangun dari 3 file hasil rekonsiliasi
bulanan yang SUDAH completed (bulan harus berurutan).

Output HANYA 5 sheet ini - sheet rekening mentah dari file bulanan TIDAK
disalin ke workbook kuartalan. Karena itu, angka per kategori/pegawai
dihitung langsung di Python dari data sumber (bukan rumus SUMIF/SUMIFS ke
sheet rekening, karena sheet itu tidak ada di file output). Subtotal dan
baris yang murni menggabungkan sel lain DALAM sheet output yang sama
(Total Pendapatan, Total Beban, Laba Bersih kumulatif, dll) tetap rumus,
karena sel-sel itu memang saling merujuk di dalam satu workbook yang sama.

Beda dengan reconcile.py (yang MELAKUKAN rekonsiliasi), modul ini murni
MENGGABUNGKAN & MERINGKAS 3 bulan yang datanya sudah bersih. Tidak ada
pencocokan transfer baru di sini - itu tanggung jawab reconcile.py di
masing-masing file bulanan sebelum digabung.

Catatan penting soal Roster Gaji: pengelompokan gaji per pegawai per bulan
memakai bulan yang DISEBUT DI KETERANGAN transaksi (mis. "Gaji Eva Januari
2025"), dicari lintas SEMUA bulan kuartal - bukan berdasarkan di sheet
bulan mana transaksinya fisik tercatat. Gaji sering dibayar telat (gaji
Januari baru dibayar/tercatat Februari), jadi kalau dikelompokkan per
sheet, itu akan salah masuk kolom bulan pembayaran padahal seharusnya
kolom bulan yang digaji. Baris "Gaji Pegawai" di Laba Rugi Kuartal TIDAK
memakai logika ini - itu tetap dihitung per bulan pembayaran (basis kas),
karena Laba Rugi/Neraca memang laporan basis kas bulan berjalan.

Catatan soal Aset Tetap & Penyusutan: transaksi berkategori "Belanja
Assets" TIDAK dibebankan penuh di bulan pembeliannya (beda dengan
reconcile.py versi bulanan, yang tidak punya visibilitas lintas bulan
untuk melakukan ini) - di sini dikapitalisasi jadi Aset Tetap lalu
disusutkan garis lurus per bulan. Aturan masa manfaat: nilai aset dibagi
Rp15.000.000 dibulatkan ke atas = jumlah tahun masa manfaat (berlaku
kelipatan - Rp20jt jadi 2 tahun, Rp45jt jadi 3 tahun, dst), minimal 1
tahun. KETERBATASAN: hanya aset yang transaksi pembeliannya ADA di salah
satu bulan yang di-input laporan ini yang terdeteksi - aset yang dibeli
SEBELUM periode laporan tidak ikut disusutkan karena datanya di luar
jangkauan file yang diupload.
"""

import math
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference

import reconcile as rc

REPORT_SHEET_NAMES = {
    "Rekonsiliasi", "Laporan Laba Rugi", "Neraca", "Laporan Arus Kas",
    "Diagnostik Keseimbangan",
}


class QuarterlyInputError(Exception):
    """Dilempar kalau input tidak valid (bulan tidak berurutan, file rusak,
    dll) - pesannya dirancang untuk langsung ditunjukkan ke user di bot."""


# ---------------------------------------------------------------------------
# Baca & validasi 3 file bulanan
# ---------------------------------------------------------------------------

def compute_sheet_k(txns):
    """Replikasi rumus K2=$F$2, K(n)=K(n-1)+J(n) yang dipakai reconcile.py
    (kolom bantu Saldo Kumulatif Rekonstruksi), tapi dihitung di Python
    supaya tidak perlu sheet rekening tetap ada di workbook output.
    Return (saldo_awal, saldo_akhir) untuk satu sheet.

    Dibulatkan ke 2 desimal (presisi sen rupiah) supaya noise floating-point
    dari penjumlahan beruntun banyak transaksi tidak menumpuk jadi selisih
    kecil (mis. Rp0,64) waktu dibandingkan dengan total yang dihitung lewat
    jalur penjumlahan lain (per kategori) - dua jalur sama-sama benar tapi
    urutan penjumlahannya beda, jadi tanpa pembulatan bisa sedikit meleset.

    PENTING: baris Saldo Awal yang di-inject bank-statement-bot kadang
    kolom Saldo Kumulatifnya berisi RUMUS ('=E2', merujuk ke kolom Kredit
    di baris yang sama), bukan angka literal. openpyxl tidak mengevaluasi
    rumus (beda dengan reconcile.py yang semuanya rumus Excel, dihitung
    Excel sendiri saat dibuka) - kalau dibaca apa adanya, nilainya None dan
    opening jadi salah dianggap 0. Makanya kalau txns[0].saldo tidak bisa
    dibaca sebagai angka, fallback ke nominal baris itu sendiri (Debit/
    Kredit) - untuk baris Saldo Awal ini persis sama nilainya dengan yang
    seharusnya dihasilkan rumus '=E2' tadi."""
    if not txns:
        return 0.0, 0.0
    opening = txns[0].saldo if txns[0].saldo is not None else txns[0].nominal
    running = opening
    for t in txns[1:]:
        running += t.nominal
    return round(opening, 2), round(running, 2)


def load_month(path):
    wb = openpyxl.load_workbook(path)
    account_sheets = [s for s in wb.sheetnames if s not in REPORT_SHEET_NAMES]
    if not account_sheets:
        raise QuarterlyInputError(
            f"File '{path}' tidak punya sheet rekening yang dikenali (cuma ada: {wb.sheetnames})."
        )
    all_txns = []
    total_saldo_awal = 0.0
    total_aset = 0.0
    for sname in account_sheets:
        ws = wb[sname]
        txns, _ = rc.read_account_sheet(ws)
        all_txns.extend(txns)
        opening, closing = compute_sheet_k(txns)
        total_saldo_awal += opening
        total_aset += closing
    total_saldo_awal = round(total_saldo_awal, 2)
    total_aset = round(total_aset, 2)
    year, month = rc.detect_period(all_txns)
    if year is None:
        raise QuarterlyInputError(f"Periode tidak terdeteksi di file '{path}'.")
    return {
        "year": year,
        "month": month,
        "label": f"{rc.MONTHS_ID[month]} {year}",
        "all_txns": all_txns,
        "total_saldo_awal": total_saldo_awal,
        "total_aset": total_aset,
    }


def validate_consecutive(months):
    """months: list of dict dari load_month(), URUTAN SESUAI UPLOAD (belum
    tentu urut kronologis). Return list yang sama tapi sudah disortir
    kronologis. Raise QuarterlyInputError kalau bulannya bukan 3 bulan
    berturut-turut."""
    if len(months) != 3:
        raise QuarterlyInputError(f"Butuh tepat 3 file, yang diterima {len(months)}.")
    ordered = sorted(months, key=lambda m: (m["year"], m["month"]))
    labels = [m["label"] for m in ordered]
    for i in range(1, 3):
        y0, m0 = ordered[i - 1]["year"], ordered[i - 1]["month"]
        y1, m1 = ordered[i]["year"], ordered[i]["month"]
        exp_m = 1 if m0 == 12 else m0 + 1
        exp_y = y0 + 1 if m0 == 12 else y0
        if (y1, m1) != (exp_y, exp_m):
            raise QuarterlyInputError(
                "Bulan yang diupload tidak berurutan: " + ", ".join(labels) +
                f". Setelah {rc.MONTHS_ID[m0]} {y0} seharusnya "
                f"{rc.MONTHS_ID[exp_m]} {exp_y}, bukan {rc.MONTHS_ID[m1]} {y1}."
            )
    return ordered


# ---------------------------------------------------------------------------
# Agregasi kategori (Python, bukan rumus - sheet rekening sumber tidak ikut
# disalin ke workbook output)
# ---------------------------------------------------------------------------

def sum_category(txns, category):
    cat = category.lower()
    return round(sum(t.nominal for t in txns if (t.kategori or "").lower() == cat), 2)


def sum_category_multi(txns, categories):
    cats = {c.lower() for c in categories}
    return round(sum(t.nominal for t in txns if (t.kategori or "").lower() in cats), 2)


def sum_category_prefix(txns, prefix):
    p = prefix.lower()
    return round(sum(t.nominal for t in txns if (t.kategori or "").lower().startswith(p)), 2)


def sum_modal(txns):
    """Modal & Setoran Pemilik + 'Transfer Masuk ... dari rekening sendiri'
    (uang milik owner sendiri dipindah antar rekening) - konsisten dengan
    Txn.is_capital di reconcile.py."""
    return round(sum(t.nominal for t in txns if t.is_capital), 2)


# ---------------------------------------------------------------------------
# Aset Tetap & Penyusutan - kapitalisasi "Belanja Assets" lalu disusutkan
# garis lurus, bukan dibebankan penuh di bulan pembelian. Lihat catatan di
# docstring atas file ini untuk aturan masa manfaat & keterbatasannya.
# ---------------------------------------------------------------------------

ASSET_CATEGORY_TEXT = "belanja assets"
DEPRECIATION_THRESHOLD = 15_000_000  # per Rp15jt kelipatan = +1 tahun masa manfaat


def scan_assets_from_months(months):
    """Scan semua bulan buat cari transaksi 'Belanja Assets', bangun daftar
    aset tetap dengan jadwal penyusutan garis lurus. Tanggal beli disimpan
    ABSOLUT (acquired_year/acquired_month), bukan index relatif ke laporan
    ini - supaya bisa digabung dengan aset dari laporan lain (kontinuitas
    lintas kuartal/tahun) lewat merge_asset_lists()."""
    assets = []
    for m in months:
        for t in m["all_txns"]:
            if (t.kategori or "").strip().lower() != ASSET_CATEGORY_TEXT:
                continue
            cost = abs(t.nominal)
            if cost <= 0:
                continue
            n_tahun = max(math.ceil(cost / DEPRECIATION_THRESHOLD), 1)
            useful_life_months = n_tahun * 12
            monthly_dep = round(cost / useful_life_months, 2)
            assets.append({
                "acquired_year": m["year"], "acquired_month": m["month"],
                "label": m["label"], "desc": t.desc, "cost": cost,
                "n_tahun": n_tahun, "useful_life_months": useful_life_months,
                "monthly_dep": monthly_dep,
            })
    return assets


def _asset_key(a):
    """Kunci alami buat deduplikasi - kalau bulan yang sama ikut ter-scan
    ulang di laporan lain (mis. kuartal Q1 dan laporan tahunan yang sama-sama
    mencakup Januari), aset yang sama tidak didaftarkan dobel."""
    return (a["acquired_year"], a["acquired_month"], a["desc"], round(a["cost"], 2))


def merge_asset_lists(*asset_lists):
    """Gabung beberapa daftar aset (mis. dari buku aset laporan sebelumnya +
    hasil scan bulan-bulan baru), dedup berdasarkan _asset_key. Kalau kunci
    sama muncul di lebih dari satu daftar, yang dipakai adalah yang pertama
    ditemukan (urutan argumen menentukan prioritas)."""
    merged = {}
    for lst in asset_lists:
        for a in lst:
            key = _asset_key(a)
            if key not in merged:
                merged[key] = a
    return list(merged.values())


def assets_with_relative_idx(assets, months):
    """Hitung month_idx RELATIF terhadap bulan pertama laporan ini (months[0])
    untuk tiap aset - bisa NEGATIF kalau aset itu dibeli sebelum periode
    laporan ini (dari laporan sebelumnya yang di-carry-forward), itu wajar
    dan tetap dihitung benar oleh depreciation_for_month_idx/
    book_value_at_month_idx karena keduanya cuma bandingkan rentang."""
    base_year, base_month = months[0]["year"], months[0]["month"]
    result = []
    for a in assets:
        rel_idx = (a["acquired_year"] - base_year) * 12 + (a["acquired_month"] - base_month)
        result.append({**a, "month_idx": rel_idx})
    return result


def load_asset_ledger(path):
    """Baca sheet 'Buku Aset Tetap' dari file laporan kuartalan/tahunan
    SEBELUMNYA (kalau ada) - dipakai untuk menyambung penyusutan aset lama
    yang dibeli SEBELUM periode laporan yang sedang dibuat. Return list
    kosong kalau sheet itu tidak ada (mis. laporan lama belum punya fitur
    ini, atau memang belum ada aset tetap sama sekali)."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return []
    name = "Buku Aset Tetap"
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    assets = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            acquired_year, acquired_month, label, desc, cost, n_tahun = (
                int(row[0]), int(row[1]), row[2], row[3], float(row[4]), int(row[5])
            )
        except (TypeError, ValueError):
            continue
        useful_life_months = n_tahun * 12
        monthly_dep = round(cost / useful_life_months, 2)
        assets.append({
            "acquired_year": acquired_year, "acquired_month": acquired_month,
            "label": label, "desc": desc, "cost": cost, "n_tahun": n_tahun,
            "useful_life_months": useful_life_months, "monthly_dep": monthly_dep,
        })
    return assets


def write_buku_aset_tetap(wb, assets, months):
    """Sheet 'Buku Aset Tetap' - daftar LENGKAP semua aset yang diketahui
    (dari laporan ini + yang di-carry-forward dari laporan sebelumnya),
    dengan 2 kolom pertama (Tahun Beli, Bulan Beli) tersembunyi dari mata
    tapi tetap ada buat dibaca ulang otomatis oleh load_asset_ledger() saat
    laporan BERIKUTNYA dibuat - itu yang menjaga kontinuitas lintas
    kuartal/tahun. Selalu dibuat (bahkan kalau assets kosong) supaya
    strukturnya konsisten buat laporan berikutnya membaca."""
    name = "Buku Aset Tetap"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    labels = [m["label"] for m in months]
    last_year, last_month = months[-1]["year"], months[-1]["month"]
    last_idx = len(months) - 1

    ws["A1"] = "BUKU ASET TETAP (INDUK) - UNTUK KONTINUITAS ANTAR LAPORAN"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("JANGAN HAPUS/UBAH kolom Tahun Beli & Bulan Beli - laporan kuartalan/tahunan "
                "BERIKUTNYA akan membaca sheet ini otomatis kalau file ini dipakai sebagai "
                "'carry-forward' untuk menyambung penyusutan aset yang dibeli sebelum periode "
                "laporan baru itu.")
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    r = 4
    headers = ["Tahun Beli", "Bulan Beli", "Label Bulan Beli", "Deskripsi", "Nilai Perolehan",
               "Masa Manfaat (Tahun)", "Penyusutan per Bulan", "Akumulasi Penyusutan s.d. Akhir Laporan Ini",
               "Nilai Buku Akhir Laporan Ini", "Status"]
    hdr_row = r
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hdr_row, column=i, value=h)
    rc.style_header(ws, hdr_row, len(headers))
    r += 1

    rel_assets = assets_with_relative_idx(assets, months)
    for a, rel in zip(assets, rel_assets):
        month_idx = rel["month_idx"]
        bulan_berjalan = min(max(last_idx - month_idx + 1, 0), a["useful_life_months"])
        akumulasi = round(bulan_berjalan * a["monthly_dep"], 2)
        nilai_buku = max(round(a["cost"] - akumulasi, 2), 0.0)
        lunas = bulan_berjalan >= a["useful_life_months"]
        ws.cell(row=r, column=1, value=a["acquired_year"])
        ws.cell(row=r, column=2, value=a["acquired_month"])
        ws.cell(row=r, column=3, value=a["label"])
        ws.cell(row=r, column=4, value=a["desc"])
        ws.cell(row=r, column=5, value=a["cost"])
        ws.cell(row=r, column=6, value=a["n_tahun"])
        ws.cell(row=r, column=7, value=a["monthly_dep"])
        ws.cell(row=r, column=8, value=akumulasi)
        ws.cell(row=r, column=9, value=nilai_buku)
        ws.cell(row=r, column=10, value="Lunas (nilai buku 0)" if lunas else "Masih disusutkan")
        for c in (5, 7, 8, 9):
            ws.cell(row=r, column=c).number_format = rc.NUMBER_FORMAT
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = rc.BORDER
        r += 1

    if not assets:
        ws.cell(row=r, column=1, value="(belum ada transaksi 'Belanja Assets' yang terdeteksi)")
        ws.cell(row=r, column=1).font = Font(italic=True, color="6B7280")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 30
    for col in "EFGHI":
        ws.column_dimensions[col].width = 20
    ws.column_dimensions["J"].width = 20
    ws.freeze_panes = "A5"
    return ws


def depreciation_for_month_idx(assets, month_idx):
    """Total beban penyusutan SEMUA aset terdaftar untuk bulan ke month_idx
    (0-based) - 0 kalau bulan itu sebelum aset dibeli atau sudah lewat masa
    manfaatnya."""
    total = 0.0
    for a in assets:
        if a["month_idx"] <= month_idx < a["month_idx"] + a["useful_life_months"]:
            total += a["monthly_dep"]
    return round(total, 2)


def book_value_at_month_idx(assets, month_idx):
    """Total nilai buku bersih (nilai perolehan - akumulasi penyusutan)
    SEMUA aset terdaftar, per AKHIR bulan ke month_idx (0-based)."""
    total = 0.0
    for a in assets:
        if a["month_idx"] > month_idx:
            continue  # belum dibeli di bulan ini
        bulan_berjalan = min(month_idx - a["month_idx"] + 1, a["useful_life_months"])
        akumulasi = round(bulan_berjalan * a["monthly_dep"], 2)
        total += max(a["cost"] - akumulasi, 0.0)
    return round(total, 2)


def _txns_for_label(months, label):
    for m in months:
        if m["label"] == label:
            return m["all_txns"]
    return []


def _nama_bulan_short(label):
    """'Januari 2025' -> 'Januari'."""
    return label.rsplit(" ", 1)[0]


# ---------------------------------------------------------------------------
# Laba Rugi Kuartal (kolom = 3 bulan + TOTAL). Basis kas per bulan
# pembayaran - Gaji TIDAK direalokasi ke bulan yang digaji (beda dengan
# Roster Gaji), sesuai kebutuhan laporan ini apa adanya.
# ---------------------------------------------------------------------------

def write_quarterly_income_statement(wb, months, assets, period_word="Kuartal"):
    name = f"Laba Rugi {period_word}"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    labels = [m["label"] for m in months]
    period_text = f"{labels[0]} - {labels[-1]}"
    ws["A1"] = f"LAPORAN LABA RUGI {period_word.upper()} - {period_text.upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"Kolom = tiap bulan (basis kas, dihitung dari data sumber), kolom TOTAL = jumlah "
                f"{len(labels)} bulan. Roster gaji per pegawai (basis accrual) ada di sheet terpisah. "
                "Pembelian aset tetap TIDAK dibebankan penuh di sini - disusutkan garis lurus "
                "(baris 'Beban Penyusutan'), sisanya jadi Aset Tetap di Neraca.")
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    r = 4
    rc.write_pivot_header(ws, r, labels)
    r += 1

    rc.write_pivot_section(ws, r, "PENDAPATAN", labels)
    r += 1
    rev_rows = []
    for cat in rc.INCOME_CATEGORIES_REVENUE:
        rc.write_pivot_data_row(
            ws, r, cat, labels,
            lambda label, cat=cat: sum_category(_txns_for_label(months, label), cat),
        )
        rev_rows.append(r)
        r += 1
    total_rev_row = r
    rc.write_pivot_subtotal_row(ws, r, "Total Pendapatan", labels, rev_rows)
    r += 2

    rc.write_pivot_section(ws, r, "BEBAN", labels)
    r += 1
    exp_rows = []
    for cat in rc.INCOME_CATEGORIES_EXPENSE:
        if cat.strip().lower() == ASSET_CATEGORY_TEXT:
            continue  # dikapitalisasi jadi Aset Tetap, bukan beban tunai penuh - lihat baris Beban Penyusutan
        rc.write_pivot_data_row(
            ws, r, cat, labels,
            lambda label, cat=cat: sum_category(_txns_for_label(months, label), cat),
        )
        exp_rows.append(r)
        r += 1
    rc.write_pivot_data_row(
        ws, r, "Marketing & RnD", labels,
        lambda label: sum_category_multi(_txns_for_label(months, label), rc.MARKETING_RND_CATEGORY_TEXTS),
    )
    exp_rows.append(r)
    r += 1
    rc.write_pivot_data_row(
        ws, r, "Gaji Pegawai (basis kas - rincian accrual per orang: sheet Roster Gaji)", labels,
        lambda label: sum_category_prefix(_txns_for_label(months, label), "gaji"),
    )
    exp_rows.append(r)
    r += 1
    rc.write_pivot_data_row(
        ws, r, "Biaya Admin Bank (termasuk biaya transfer Fliptech)", labels,
        lambda label: sum_category_multi(_txns_for_label(months, label), rc.BANK_FEE_CATEGORY_TEXTS),
    )
    exp_rows.append(r)
    r += 1
    if assets:
        rc.write_pivot_data_row(
            ws, r, "Beban Penyusutan (Aset Tetap, garis lurus)", labels,
            lambda label: -depreciation_for_month_idx(assets, labels.index(label)),
        )
        exp_rows.append(r)
        r += 1
    total_exp_row = r
    rc.write_pivot_subtotal_row(ws, r, "Total Beban", labels, exp_rows)
    r += 2

    rc.write_pivot_section(ws, r, "LAIN-LAIN (perlu verifikasi manual)", labels)
    r += 1
    other_rows = []
    for cat in rc.OTHER_CATEGORIES:
        rc.write_pivot_data_row(
            ws, r, cat, labels,
            lambda label, cat=cat: sum_category(_txns_for_label(months, label), cat),
        )
        other_rows.append(r)
        r += 1
    total_other_row = r
    rc.write_pivot_subtotal_row(ws, r, "Total Lain-lain", labels, other_rows)
    r += 2

    net_row = r
    rc.write_pivot_formula_row(
        ws, r, "LABA / RUGI BERSIH", labels,
        lambda cl: f"={cl}{total_rev_row}+{cl}{total_exp_row}+{cl}{total_other_row}",
        bold=True,
    )
    for c in range(1, rc.pivot_total_col(labels) + 1):
        ws.cell(row=r, column=c).font = Font(bold=True, size=12)

    ws.column_dimensions["A"].width = 46
    for i in range(len(labels)):
        ws.column_dimensions[rc.col_letter(2 + i)].width = 18
    ws.column_dimensions[rc.col_letter(rc.pivot_total_col(labels))].width = 18
    ws.freeze_panes = "B5"

    return {"sheet": name, "net": net_row, "total_rev": total_rev_row,
            "total_exp": total_exp_row, "total_other": total_other_row,
            "labels": labels, "total_col": rc.pivot_total_col(labels)}


# ---------------------------------------------------------------------------
# Helper pivot "snapshot": khusus buat Neraca Kuartal, di mana kolom kanan
# HARUS berarti "posisi di akhir kuartal" (= kolom bulan terakhir), BUKAN
# penjumlahan 3 bulan - beda dengan Laba Rugi/Arus Kas yang memang laporan
# arus (flow) sehingga penjumlahan masuk akal. Neraca itu snapshot (stok),
# menjumlah 3 snapshot saldo akan salah secara akuntansi.
# ---------------------------------------------------------------------------

def write_snapshot_header(ws, row, labels, period_word="Kuartal"):
    ws.cell(row=row, column=1, value="")
    for i, label in enumerate(labels):
        c = 2 + i
        cell = ws.cell(row=row, column=c, value=label)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    total_col = 2 + len(labels)
    ws.cell(row=row, column=total_col, value=f"AKHIR {period_word.upper()} ({labels[-1]})")
    rc.style_header(ws, row, total_col)
    ws.row_dimensions[row].height = 32


def write_snapshot_data_row(ws, row, label, labels, formula_fn, bold=False):
    ws.cell(row=row, column=1, value=label)
    n = len(labels)
    for i, lb in enumerate(labels):
        c = 2 + i
        ws.cell(row=row, column=c, value=formula_fn(lb))
        ws.cell(row=row, column=c).number_format = rc.NUMBER_FORMAT
    total_col = 2 + n
    last_col_letter = rc.col_letter(total_col - 1)
    ws.cell(row=row, column=total_col, value=f"={last_col_letter}{row}")
    ws.cell(row=row, column=total_col).number_format = rc.NUMBER_FORMAT
    if bold:
        for c in range(1, total_col + 1):
            ws.cell(row=row, column=c).font = Font(bold=True)


def write_snapshot_subtotal_row(ws, row, label, labels, ref_rows, bold=True):
    n = len(labels)
    ws.cell(row=row, column=1, value=label)
    for i in range(n):
        c = 2 + i
        cl = rc.col_letter(c)
        ws.cell(row=row, column=c, value=f"=SUM({cl}{ref_rows[0]}:{cl}{ref_rows[-1]})")
        ws.cell(row=row, column=c).number_format = rc.NUMBER_FORMAT
    total_col = 2 + n
    last_col_letter = rc.col_letter(total_col - 1)
    ws.cell(row=row, column=total_col, value=f"={last_col_letter}{row}")
    ws.cell(row=row, column=total_col).number_format = rc.NUMBER_FORMAT
    if bold:
        for c in range(1, total_col + 1):
            ws.cell(row=row, column=c).font = Font(bold=True)


def write_snapshot_formula_row(ws, row, label, labels, per_col_formula_fn, bold=False):
    n = len(labels)
    ws.cell(row=row, column=1, value=label)
    for i in range(n):
        c = 2 + i
        cl = rc.col_letter(c)
        ws.cell(row=row, column=c, value=per_col_formula_fn(cl))
        ws.cell(row=row, column=c).number_format = rc.NUMBER_FORMAT
    total_col = 2 + n
    last_col_letter = rc.col_letter(total_col - 1)
    ws.cell(row=row, column=total_col, value=f"={last_col_letter}{row}")
    ws.cell(row=row, column=total_col).number_format = rc.NUMBER_FORMAT
    if bold:
        for c in range(1, total_col + 1):
            ws.cell(row=row, column=c).font = Font(bold=True)


# ---------------------------------------------------------------------------
# Neraca Kuartal
# ---------------------------------------------------------------------------

def write_quarterly_balance_sheet(wb, months, income_ref, assets, period_word="Kuartal"):
    name = f"Neraca {period_word}"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    labels = [m["label"] for m in months]
    ws["A1"] = f"NERACA {period_word.upper()} - PER AKHIR {labels[0].upper()} S.D. {labels[-1].upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Tiap kolom = posisi akhir bulan itu (snapshot). Sisi Ekuitas dihitung kumulatif "
                f"dari bulan pertama {period_word.lower()} supaya nyambung dengan sisi Aset.")
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    r = 4
    write_snapshot_header(ws, r, labels, period_word)
    r += 1

    rc.write_pivot_section(ws, r, "ASET", labels)
    r += 1
    kas_row = r
    write_snapshot_data_row(
        ws, r, "Kas & Setara Kas (Saldo Akhir Bulan)", labels,
        lambda label: next(m["total_aset"] for m in months if m["label"] == label),
    )
    r += 1
    if assets:
        aset_tetap_row = r
        write_snapshot_data_row(
            ws, r, "Aset Tetap (Nilai Buku Bersih, setelah penyusutan)", labels,
            lambda label: book_value_at_month_idx(assets, labels.index(label)),
        )
        r += 1
        total_asset_row = r
        write_snapshot_subtotal_row(ws, r, "Total Aset", labels, [kas_row, aset_tetap_row])
        r += 1
    else:
        total_asset_row = kas_row
        ws.cell(row=kas_row, column=1).font = Font(bold=True)
        for i in range(len(labels)):
            ws.cell(row=kas_row, column=2 + i).font = Font(bold=True)
        ws.cell(row=kas_row, column=rc.pivot_total_col(labels)).font = Font(bold=True)
    r += 1

    rc.write_pivot_section(ws, r, f"EKUITAS (kumulatif sejak awal {period_word.lower()})", labels)
    r += 1
    saldo_awal_kuartal = months[0]["total_saldo_awal"]
    saldo_awal_row = r
    write_snapshot_data_row(
        ws, r, f"Saldo Awal {period_word} (tetap, dari bulan pertama)", labels,
        lambda label, val=saldo_awal_kuartal: (
            val if label == labels[0] else f"={rc.col_letter(2)}{saldo_awal_row}"
        ),
    )
    r += 1
    modal_row = r
    modal_per_month = [sum_modal(m["all_txns"]) for m in months]
    modal_kumulatif = []
    running = 0.0
    for v in modal_per_month:
        running = round(running + v, 2)
        modal_kumulatif.append(running)
    write_snapshot_data_row(
        ws, r, "Modal & Setoran Pemilik (kumulatif s.d. bulan ini)", labels,
        lambda label: modal_kumulatif[labels.index(label)],
    )
    r += 1
    laba_row = r
    write_snapshot_formula_row(
        ws, r, "Laba Bersih (kumulatif s.d. bulan ini)", labels,
        lambda cl: laba_kumulatif_formula(income_ref, cl),
    )
    r += 1
    total_equity_row = r
    write_snapshot_subtotal_row(ws, r, "Total Ekuitas", labels, [saldo_awal_row, laba_row])
    r += 2

    balance_check_row = r
    write_snapshot_formula_row(
        ws, r, "CEK KESEIMBANGAN (Aset - Ekuitas)", labels,
        lambda cl: f"={cl}{total_asset_row}-{cl}{total_equity_row}",
        bold=True,
    )
    r += 1
    transfer_row = r
    transfer_per_month = [sum_category_multi(m["all_txns"], rc.TRANSFER_CATEGORY_TEXTS) for m in months]
    transfer_kumulatif = []
    running = 0.0
    for v in transfer_per_month:
        running = round(running + v, 2)
        transfer_kumulatif.append(running)
    write_snapshot_data_row(
        ws, r, "Transfer Bersih Kumulatif (info)", labels,
        lambda label: transfer_kumulatif[labels.index(label)],
    )
    r += 1
    residual_row = r
    write_snapshot_formula_row(
        ws, r, "Selisih Belum Terjelaskan", labels,
        lambda cl: f"={cl}{balance_check_row}-{cl}{transfer_row}",
        bold=True,
    )
    r += 1
    # Cek kontinuitas antar bulan: apakah Saldo Awal yang DIDEKLARASIKAN di
    # file bulan ini sama dengan Saldo Akhir yang benar-benar DIHITUNG dari
    # transaksi bulan sebelumnya? Kalau tidak, itu sumber selisih yang
    # SUDAH ADA DI DATA SUMBER (mis. pembulatan bank antar bulan) - bukan
    # bug perhitungan, dan tidak bisa diperbaiki dari sisi laporan ini.
    continuity_row = r
    continuity_gaps = [0.0]
    for i in range(1, len(months)):
        continuity_gaps.append(round(months[i]["total_saldo_awal"] - months[i - 1]["total_aset"], 2))
    write_snapshot_data_row(
        ws, r, "Cek Kontinuitas: Saldo Awal Bulan Ini - Saldo Akhir Bulan Lalu", labels,
        lambda label: continuity_gaps[labels.index(label)],
    )
    r += 1
    ws.cell(row=r, column=1,
            value=("Kalau baris di atas tidak 0, itu artinya file bulan ini mendeklarasikan Saldo "
                   "Awal yang beda dari hasil hitung transaksi bulan sebelumnya - selisih ini SUDAH "
                   "ADA di data sumber (mis. pembulatan bank saat laporan berganti bulan), bukan "
                   f"kesalahan perhitungan laporan {period_word.lower()} ini. Kalau 'Selisih Belum "
                   "Terjelaskan' di atas mendekati angka yang sama, itu kemungkinan besar penyebabnya."))
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="6B7280")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=rc.pivot_total_col(labels))
    ws.row_dimensions[r].height = 48
    r += 2

    # ------------------------------------------------------------------
    # RINGKASAN & PROGRES BULANAN - sengaja dipisah dari bagian Ekuitas di
    # atas (yang kumulatif, untuk kebutuhan akuntansi formal). Bagian ini
    # murni angka PER BULAN (bukan akumulasi sampai bulan itu), supaya bisa
    # langsung lihat performa bulan tertentu tanpa harus mengurangi angka
    # kumulatif secara manual. Kolom TOTAL di sini = jumlah 3 bulan (wajar
    # dijumlah karena semuanya angka arus/delta, bukan saldo/snapshot).
    # ------------------------------------------------------------------
    ws.cell(row=r, column=1, value="RINGKASAN & PROGRES BULANAN (non-kumulatif)")
    ws.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    ws.cell(row=r, column=1,
            value=("Angka per bulan berdiri sendiri (bukan akumulasi) - beda dengan bagian EKUITAS "
                   "di atas yang sengaja kumulatif untuk kebutuhan Neraca formal."))
    ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="6B7280")
    r += 2

    rc.write_pivot_header(ws, r, labels)
    r += 1
    modal_bulan_row = r
    rc.write_pivot_data_row(
        ws, r, "Modal Masuk Bulan Ini (bukan kumulatif)", labels,
        lambda label: modal_per_month[labels.index(label)],
    )
    r += 1
    laba_bulan_row = r
    rc.write_pivot_formula_row(
        ws, r, "Laba/Rugi Bulan Ini", labels,
        lambda cl: f"='{income_ref['sheet']}'!{cl}{income_ref['net']}",
    )
    r += 1
    kas_bulan_row = r
    rc.write_pivot_formula_row(
        ws, r, "Perubahan Kas Bersih Bulan Ini (vs bulan sebelumnya)", labels,
        lambda cl: _perubahan_kas_formula(cl, labels, total_asset_row, saldo_awal_row),
        bold=True,
    )

    ws.column_dimensions["A"].width = 42
    for i in range(len(labels)):
        ws.column_dimensions[rc.col_letter(2 + i)].width = 18
    ws.column_dimensions[rc.col_letter(rc.pivot_total_col(labels))].width = 22
    ws.freeze_panes = "B5"

    return {"sheet": name, "total_asset": total_asset_row, "total_equity": total_equity_row,
            "saldo_awal": saldo_awal_row, "balance_check": balance_check_row,
            "labels": labels, "total_col": rc.pivot_total_col(labels)}


def laba_kumulatif_formula(income_ref, current_col_letter):
    """SUM kolom Laba Rugi dari bulan pertama s.d. kolom bulan ini
    (kumulatif) - rujukan lintas sheet (Neraca -> Laba Rugi Kuartal) tetap
    aman karena kedua sheet itu ada di workbook output yang sama."""
    income_sheet = income_ref["sheet"]
    net_row = income_ref["net"]
    first_letter = rc.col_letter(2)
    return f"=SUM('{income_sheet}'!{first_letter}{net_row}:{current_col_letter}{net_row})"


def _perubahan_kas_formula(cl, labels, total_asset_row, saldo_awal_row):
    """Perubahan kas bulan ini = Kas bulan ini - Kas bulan sebelumnya (atau
    Saldo Awal Kuartal untuk bulan pertama). Kolom TOTAL = perubahan
    sepanjang kuartal (kas bulan terakhir - saldo awal kuartal), sama
    hasilnya dengan menjumlah 3 perubahan bulanan (teleskopik)."""
    total_col_letter = rc.col_letter(len(labels) + 2)
    if cl == total_col_letter:
        last_month_letter = rc.col_letter(len(labels) + 1)
        return f"={last_month_letter}{total_asset_row}-{rc.col_letter(2)}{saldo_awal_row}"
    idx = next(i for i in range(len(labels)) if rc.col_letter(2 + i) == cl)
    if idx == 0:
        return f"={cl}{total_asset_row}-{rc.col_letter(2)}{saldo_awal_row}"
    prev_letter = rc.col_letter(2 + idx - 1)
    return f"={cl}{total_asset_row}-{prev_letter}{total_asset_row}"


# ---------------------------------------------------------------------------
# Laporan Arus Kas Kuartal
# ---------------------------------------------------------------------------

def write_quarterly_cash_flow(wb, months, income_ref, balance_ref, period_word="Kuartal"):
    name = f"Arus Kas {period_word}"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    labels = [m["label"] for m in months]
    ws["A1"] = f"LAPORAN ARUS KAS {period_word.upper()} - {labels[0].upper()} S.D. {labels[-1].upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Metode langsung, kolom = tiap bulan + TOTAL {period_word.lower()}."
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    r = 4
    rc.write_pivot_header(ws, r, labels)
    r += 1

    rc.write_pivot_section(ws, r, "ARUS KAS DARI AKTIVITAS OPERASI", labels)
    r += 1
    op_row = r
    rc.write_pivot_formula_row(
        ws, r, "Laba Bersih Bulan Ini (basis kas)", labels,
        lambda cl: f"='{income_ref['sheet']}'!{cl}{income_ref['net']}",
    )
    r += 1
    total_op_row = r
    rc.write_pivot_subtotal_row(ws, r, "Kas Bersih dari Operasi", labels, [op_row, op_row])
    r += 2

    rc.write_pivot_section(ws, r, "ARUS KAS DARI AKTIVITAS PENDANAAN", labels)
    r += 1
    fin_row = r
    rc.write_pivot_data_row(
        ws, r, "Modal & Setoran Pemilik", labels,
        lambda label: sum_modal(_txns_for_label(months, label)),
    )
    r += 1
    total_fin_row = r
    rc.write_pivot_subtotal_row(ws, r, "Kas Bersih dari Pendanaan", labels, [fin_row, fin_row])
    r += 2

    net_change_row = r
    rc.write_pivot_formula_row(
        ws, r, "KENAIKAN (PENURUNAN) KAS BERSIH", labels,
        lambda cl: f"={cl}{total_op_row}+{cl}{total_fin_row}",
        bold=True,
    )
    r += 1
    saldo_awal_row = r
    rc.write_pivot_formula_row(
        ws, r, "Saldo Kas Awal Bulan", labels,
        lambda cl: (
            f"='{balance_ref['sheet']}'!{rc.col_letter(2)}{balance_ref['saldo_awal']}"
            if cl == rc.col_letter(2)
            else _saldo_awal_chain(cl, labels, net_change_row, saldo_awal_row)
        ),
    )
    r += 1
    saldo_akhir_row = r
    rc.write_pivot_formula_row(
        ws, r, "Saldo Kas Akhir Bulan", labels,
        lambda cl: f"={cl}{net_change_row}+{cl}{saldo_awal_row}",
        bold=True,
    )
    r += 1
    rc.write_pivot_formula_row(
        ws, r, "Cek vs Total Aset di Neraca", labels,
        lambda cl: f"={cl}{saldo_akhir_row}-'{balance_ref['sheet']}'!{cl}{balance_ref['total_asset']}",
    )

    ws.column_dimensions["A"].width = 40
    for i in range(len(labels)):
        ws.column_dimensions[rc.col_letter(2 + i)].width = 18
    ws.column_dimensions[rc.col_letter(rc.pivot_total_col(labels))].width = 18
    ws.freeze_panes = "B5"
    return ws


def _saldo_awal_chain(cl, labels, net_change_row, saldo_awal_row):
    """Saldo awal bulan ke-2 & ke-3 = saldo akhir bulan sebelumnya (kolom
    sebelah kiri di sheet yang sama). Kolom TOTAL kuartal = saldo awal
    bulan pertama (titik mula kuartal)."""
    total_col_letter = rc.col_letter(len(labels) + 2)
    if cl == total_col_letter:
        return f"={rc.col_letter(2)}{saldo_awal_row}"
    idx = None
    for i in range(len(labels)):
        if rc.col_letter(2 + i) == cl:
            idx = i
            break
    prev_letter = rc.col_letter(2 + idx - 1)
    saldo_akhir_row = saldo_awal_row + 1
    return f"={prev_letter}{saldo_akhir_row}"


# ---------------------------------------------------------------------------
# Roster Gaji 3 Bulan (matriks: baris = pegawai, kolom = 3 bulan + TOTAL).
# Dikelompokkan berdasarkan bulan yang DISEBUT DI KETERANGAN transaksi
# (basis accrual), dicari lintas semua bulan kuartal - lihat catatan di
# docstring atas file ini.
# ---------------------------------------------------------------------------

def _norm_name(name):
    return " ".join(name.split()).casefold()


# Alias nama pegawai: variasi penulisan yang merujuk ke orang yang sama.
# Kunci harus huruf kecil. Dicocokkan sebagai AWALAN teks Objek yang sudah
# dirapikan (bukan cuma exact match) - supaya kasus seperti "Roziyan Hidayat
# Mei (Sisanya Di Tunai)" (ada catatan tambahan nempel di nama) tetap
# kecocokan ke alias "roziyan hidayat".
EMPLOYEE_ALIASES = {
    "viona winda octavia": "Viona Winda Octavia",
    "viona": "Viona Winda Octavia",
    "ahmad roziyan hidayat": "Ahmad Roziyan Hidayat",
    "ahmad roziyan h.": "Ahmad Roziyan Hidayat",
    "roziyan hidayat": "Ahmad Roziyan Hidayat",
    "roziyan": "Ahmad Roziyan Hidayat",
    "ojan": "Ahmad Roziyan Hidayat",
    "kak ojan": "Ahmad Roziyan Hidayat",
    "owner": "Ahmad Roziyan Hidayat",
}
_EMPLOYEE_ALIAS_KEYS_SORTED = sorted(EMPLOYEE_ALIASES.keys(), key=len, reverse=True)


def resolve_employee_name(raw_objek):
    """Petakan teks Objek mentah ke nama kanonik pegawai. Dicocokkan sebagai
    awalan (bukan cuma exact match persis) diikuti batas kata/akhir teks,
    supaya nama yang ada embel-embel tambahan (mis. 'Roziyan Hidayat Mei
    (Sisanya Di Tunai)') tetap dikenali sebagai orang yang sama. Kalau tidak
    ada alias yang cocok, pakai teks aslinya (dirapikan spasinya) apa
    adanya."""
    norm = _norm_name(raw_objek)
    for key in _EMPLOYEE_ALIAS_KEYS_SORTED:
        if norm == key or norm.startswith(key + " ") or norm.startswith(key + "("):
            return EMPLOYEE_ALIASES[key]
    return " ".join(raw_objek.split())


def write_roster_gaji(wb, months, period_word="kuartal"):
    name = f"Roster Gaji {len(months)} Bulan"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    labels = [m["label"] for m in months]
    ws["A1"] = f"ROSTER GAJI PEGAWAI - {labels[0].upper()} S.D. {labels[-1].upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    nama_bulan_list = [_nama_bulan_short(lb) for lb in labels]
    ws["A2"] = ("Dikelompokkan berdasarkan bulan yang DISEBUT DI KETERANGAN transaksi (basis "
                f"accrual, mis. 'Gaji Eva Januari 2025'), dicari lintas semua bulan {period_word} - "
                f"supaya gaji yang telat dibayar/dicatat tetap masuk kolom bulan yang benar. Kalau "
                f"kolom {labels[-1]} kosong padahal pegawainya muncul di bulan lain, kemungkinan "
                "gajinya belum dibayar - lihat kolom Catatan.")
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    def _parse_bulan_dari_keterangan(ket):
        ket_lower = (ket or "").lower()
        for nama in nama_bulan_list:
            if nama.lower() in ket_lower:
                return nama
        return None

    employee_month_amount = {}  # (kunci-nama, idx-bulan) -> jumlah
    employees_by_month = [set() for _ in labels]
    canonical_name = {}
    all_employee_keys = []
    n_gaji_luar_kuartal = 0
    for m in months:
        for t in m["all_txns"]:
            if (t.kategori or "").lower().startswith("gaji") and t.objek:
                bulan_untuk = _parse_bulan_dari_keterangan(t.desc)
                if bulan_untuk is None:
                    n_gaji_luar_kuartal += 1
                    continue
                idx = nama_bulan_list.index(bulan_untuk)
                canonical = resolve_employee_name(t.objek)
                key = canonical.casefold()
                employees_by_month[idx].add(key)
                canonical_name.setdefault(key, canonical)
                employee_month_amount[(key, idx)] = employee_month_amount.get((key, idx), 0.0) + t.nominal
    for idx in range(len(labels)):
        for key in sorted(employees_by_month[idx]):
            if key not in all_employee_keys:
                all_employee_keys.append(key)
    all_employees = [canonical_name[k] for k in all_employee_keys]

    r = 4
    rc.write_pivot_header(ws, r, labels)
    ws.cell(row=r, column=rc.pivot_total_col(labels) + 1, value="Catatan")
    ws.cell(row=r, column=rc.pivot_total_col(labels) + 1).fill = rc.HEADER_FILL
    ws.cell(row=r, column=rc.pivot_total_col(labels) + 1).font = rc.HEADER_FONT
    r += 1

    n_belum_dibayar_bulan_terakhir = 0
    last_label = labels[-1]
    for key, emp in zip(all_employee_keys, all_employees):
        rc.write_pivot_data_row(
            ws, r, emp, labels,
            lambda label, key=key: employee_month_amount.get((key, labels.index(label)), 0.0),
        )
        paid_last_month = key in employees_by_month[-1]
        if not paid_last_month:
            n_belum_dibayar_bulan_terakhir += 1
            ws.cell(row=r, column=rc.pivot_total_col(labels) + 1,
                    value=(f"Belum ada gaji untuk {last_label} tercatat - kemungkinan "
                           f"dibebankan sebagai accrual di bulan setelah {period_word} ini."))
            ws.cell(row=r, column=rc.pivot_total_col(labels) + 1).font = Font(italic=True, color="B45309")
            ws.cell(row=r, column=rc.pivot_total_col(labels) + 1).alignment = Alignment(wrap_text=True)
        r += 1

    if all_employees:
        first_data_row = 5
        last_data_row_ = r - 1
        rc.write_pivot_subtotal_row(ws, r, "TOTAL", labels, [first_data_row, last_data_row_])
    else:
        ws.cell(row=r, column=1, value=f"(tidak ada transaksi berkategori Gaji untuk bulan-bulan di {period_word} ini)")
        ws.cell(row=r, column=1).font = Font(italic=True, color="6B7280")
    r += 1
    if n_gaji_luar_kuartal:
        ws.cell(row=r, column=1,
                value=(f"Catatan: {n_gaji_luar_kuartal} baris transaksi berkategori Gaji* keterangannya "
                       f"tidak menyebut salah satu dari {len(labels)} bulan {period_word} ini (kemungkinan "
                       f"accrual dari bulan sebelum {period_word} dimulai) - tetap terhitung dalam baris "
                       "'Gaji Pegawai' basis kas di Laba Rugi, tapi tidak muncul di matriks roster ini "
                       "karena tidak jelas masuk kolom bulan yang mana."))
        ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="6B7280")
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=rc.pivot_total_col(labels) + 1)
        ws.row_dimensions[r].height = 40

    ws.column_dimensions["A"].width = 28
    for i in range(len(labels)):
        ws.column_dimensions[rc.col_letter(2 + i)].width = 18
    ws.column_dimensions[rc.col_letter(rc.pivot_total_col(labels))].width = 18
    ws.column_dimensions[rc.col_letter(rc.pivot_total_col(labels) + 1)].width = 46
    ws.freeze_panes = "B5"

    return {"n_pegawai": len(all_employees),
            "n_belum_dibayar_bulan_terakhir": n_belum_dibayar_bulan_terakhir}


# ---------------------------------------------------------------------------
# Analisis & Tren (sheet ke-5): grafik + pembacaan tren naratif
# ---------------------------------------------------------------------------

def compute_month_summary(txns):
    revenue = sum(sum_category(txns, c) for c in rc.INCOME_CATEGORIES_REVENUE)
    expense = sum(sum_category(txns, c) for c in rc.INCOME_CATEGORIES_EXPENSE)
    expense += sum_category_multi(txns, rc.MARKETING_RND_CATEGORY_TEXTS)
    expense += sum_category_prefix(txns, "gaji")
    expense += sum_category_multi(txns, rc.BANK_FEE_CATEGORY_TEXTS)
    other = sum(sum_category(txns, c) for c in rc.OTHER_CATEGORIES)
    net = revenue + expense + other
    return revenue, expense, other, net


def write_analysis_sheet(wb, months, period_word="Kuartal"):
    name = "Analisis & Tren"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    labels = [m["label"] for m in months]
    ws["A1"] = f"ANALISIS & TREN KEUANGAN - {labels[0].upper()} S.D. {labels[-1].upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"Grafik dan pembacaan tren otomatis dari data {len(labels)} bulan di sheet "
                f"Laba Rugi {period_word} & Neraca {period_word}.")
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")

    summaries = [compute_month_summary(m["all_txns"]) for m in months]
    revenues = [s[0] for s in summaries]
    expenses = [s[1] for s in summaries]  # negatif
    nets = [s[3] for s in summaries]
    kas = [m["total_aset"] for m in months]

    # --- tabel data 1: ringkasan per bulan (basis grafik garis) ---
    r = 4
    ws.cell(row=r, column=1, value="RINGKASAN PER BULAN (data grafik)")
    ws.cell(row=r, column=1).font = rc.SECTION_FONT
    ws.cell(row=r, column=1).fill = rc.SECTION_FILL
    r += 1
    hdr1 = r
    headers1 = ["Bulan", "Pendapatan", "Beban", "Laba Bersih", "Kas Akhir"]
    for i, h in enumerate(headers1, start=1):
        ws.cell(row=hdr1, column=i, value=h)
    rc.style_header(ws, hdr1, len(headers1))
    r += 1
    data1_start = r
    for i, lb in enumerate(labels):
        ws.cell(row=r, column=1, value=lb)
        ws.cell(row=r, column=2, value=revenues[i])
        ws.cell(row=r, column=3, value=abs(expenses[i]))
        ws.cell(row=r, column=4, value=nets[i])
        ws.cell(row=r, column=5, value=kas[i])
        for c in range(2, 6):
            ws.cell(row=r, column=c).number_format = rc.NUMBER_FORMAT
        r += 1
    data1_end = r - 1
    r += 1

    # --- tabel data 2: komposisi beban per kategori (basis grafik batang) ---
    ws.cell(row=r, column=1, value=f"KOMPOSISI BEBAN PER KATEGORI (total {period_word.lower()})")
    ws.cell(row=r, column=1).font = rc.SECTION_FONT
    ws.cell(row=r, column=1).fill = rc.SECTION_FILL
    r += 1
    hdr2 = r
    ws.cell(row=hdr2, column=1, value="Kategori")
    ws.cell(row=hdr2, column=2, value="Total")
    rc.style_header(ws, hdr2, 2)
    r += 1
    data2_start = r
    all_txns_quarter = [t for m in months for t in m["all_txns"]]
    expense_cats = list(rc.INCOME_CATEGORIES_EXPENSE) + ["Marketing & RnD", "Gaji Pegawai", "Biaya Admin Bank"]
    expense_vals = []
    for cat in rc.INCOME_CATEGORIES_EXPENSE:
        expense_vals.append(abs(sum_category(all_txns_quarter, cat)))
    expense_vals.append(abs(sum_category_multi(all_txns_quarter, rc.MARKETING_RND_CATEGORY_TEXTS)))
    expense_vals.append(abs(sum_category_prefix(all_txns_quarter, "gaji")))
    expense_vals.append(abs(sum_category_multi(all_txns_quarter, rc.BANK_FEE_CATEGORY_TEXTS)))
    for cat, val in zip(expense_cats, expense_vals):
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=2).number_format = rc.NUMBER_FORMAT
        r += 1
    data2_end = r - 1
    r += 2

    # --- grafik garis: Pendapatan vs Beban vs Laba Bersih ---
    chart1 = LineChart()
    chart1.title = "Pendapatan vs Beban vs Laba Bersih per Bulan"
    chart1.style = 2
    chart1.y_axis.title = "Rupiah"
    chart1.x_axis.title = "Bulan"
    cats = Reference(ws, min_col=1, min_row=data1_start, max_row=data1_end)
    for col, name_seri in ((2, "Pendapatan"), (3, "Beban"), (4, "Laba Bersih")):
        data = Reference(ws, min_col=col, min_row=hdr1, max_row=data1_end)
        chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.width = 18
    chart1.height = 9
    chart_anchor_row = r
    ws.add_chart(chart1, f"A{chart_anchor_row}")

    # --- grafik garis: tren kas ---
    chart2 = LineChart()
    chart2.title = "Tren Kas Akhir Bulan"
    chart2.style = 10
    chart2.y_axis.title = "Rupiah"
    chart2.x_axis.title = "Bulan"
    data_kas = Reference(ws, min_col=5, min_row=hdr1, max_row=data1_end)
    chart2.add_data(data_kas, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.width = 18
    chart2.height = 9
    ws.add_chart(chart2, f"A{chart_anchor_row + 19}")

    # --- grafik batang: komposisi beban ---
    chart3 = BarChart()
    chart3.type = "col"
    chart3.title = f"Komposisi Beban per Kategori (Total {period_word})"
    chart3.y_axis.title = "Rupiah"
    chart3.style = 11
    data_exp = Reference(ws, min_col=2, min_row=hdr2, max_row=data2_end)
    cats_exp = Reference(ws, min_col=1, min_row=data2_start, max_row=data2_end)
    chart3.add_data(data_exp, titles_from_data=True)
    chart3.set_categories(cats_exp)
    chart3.width = 18
    chart3.height = 9
    ws.add_chart(chart3, f"A{chart_anchor_row + 38}")

    # --- narasi tren (dihitung Python, ditulis sebagai teks) ---
    narasi_row = chart_anchor_row + 57
    ws.cell(row=narasi_row, column=1, value="PEMBACAAN TREN")
    ws.cell(row=narasi_row, column=1).font = rc.SECTION_FONT
    ws.cell(row=narasi_row, column=1).fill = rc.SECTION_FILL
    narasi_row += 1

    kalimat = build_narasi_tren(labels, revenues, expenses, nets, kas, expense_cats, expense_vals)
    for line in kalimat:
        ws.cell(row=narasi_row, column=1, value=f"- {line}")
        ws.cell(row=narasi_row, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=narasi_row, start_column=1, end_row=narasi_row, end_column=5)
        ws.row_dimensions[narasi_row].height = 28
        narasi_row += 1

    ws.column_dimensions["A"].width = 30
    for c in "BCDE":
        ws.column_dimensions[c].width = 18

    return ws


def _pct_change(awal, akhir):
    if awal == 0:
        return None
    return (akhir - awal) / abs(awal) * 100


def build_narasi_tren(labels, revenues, expenses, nets, kas, expense_cats, expense_vals):
    kalimat = []

    # tren laba bersih
    growth = _pct_change(nets[0], nets[-1])
    if growth is None:
        kalimat.append(
            f"Laba bersih {labels[0]} tercatat Rp0, jadi persentase perubahan ke {labels[-1]} "
            f"(Rp{nets[-1]:,.0f}) tidak bisa dihitung.".replace(",", ".")
        )
    else:
        arah = "naik" if growth > 0 else ("turun" if growth < 0 else "stabil")
        kalimat.append(
            f"Laba bersih {arah} {abs(growth):,.1f}% dari {labels[0]} (Rp{nets[0]:,.0f}) ke "
            f"{labels[-1]} (Rp{nets[-1]:,.0f}).".replace(",", ".")
        )

    # tren pendapatan
    rev_growth = _pct_change(revenues[0], revenues[-1])
    if rev_growth is not None:
        arah = "naik" if rev_growth > 0 else ("turun" if rev_growth < 0 else "stabil")
        kalimat.append(
            f"Pendapatan {arah} {abs(rev_growth):,.1f}% dari {labels[0]} ke {labels[-1]}."
            .replace(",", ".")
        )

    # bulan rugi
    bulan_rugi = [lb for lb, n in zip(labels, nets) if n < 0]
    if bulan_rugi:
        kalimat.append(f"Rugi bersih terjadi di: {', '.join(bulan_rugi)}.")
    else:
        kalimat.append(f"Semua {len(labels)} bulan dalam periode ini sama-sama mencatat laba bersih positif.")

    # kategori beban terbesar
    if expense_vals and sum(expense_vals) > 0:
        idx_max = max(range(len(expense_vals)), key=lambda i: expense_vals[i])
        total_beban = sum(expense_vals)
        pct = expense_vals[idx_max] / total_beban * 100
        kalimat.append(
            f"Kategori beban terbesar sepanjang periode ini: {expense_cats[idx_max]} "
            f"(Rp{expense_vals[idx_max]:,.0f})".replace(",", ".") +
            f" - {pct:,.1f}% dari total beban.".replace(",", ".")
        )

    # tren kas
    kas_growth = _pct_change(kas[0], kas[-1])
    if kas_growth is not None:
        arah = "naik" if kas_growth > 0 else ("turun" if kas_growth < 0 else "stabil")
        kalimat.append(
            f"Posisi kas {arah} {abs(kas_growth):,.1f}% dari Rp{kas[0]:,.0f} ({labels[0]}) ke "
            f"Rp{kas[-1]:,.0f} ({labels[-1]}).".replace(",", ".")
        )
    kas_negatif = [lb for lb, k in zip(labels, kas) if k < 0]
    if kas_negatif:
        kalimat.append(
            f"PERHATIAN: posisi kas negatif tercatat di akhir bulan: {', '.join(kas_negatif)} - "
            "perlu ditelusuri manual."
        )

    return kalimat


# ---------------------------------------------------------------------------
# Orkestrasi utama
# ---------------------------------------------------------------------------

def run_quarterly_report(paths, output_path, carry_forward_paths=None):
    """paths: list of 3 path file bulanan (urutan upload bebas, akan
    disortir kronologis). carry_forward_paths: opsional, list path file
    laporan kuartalan/tahunan SEBELUMNYA (yang punya sheet 'Buku Aset
    Tetap') - dibaca untuk menyambung penyusutan aset yang dibeli sebelum
    periode laporan ini, supaya kontinuitas aset tetap terjaga lintas
    laporan. Return dict ringkasan. Output 6 sheet: Laba Rugi Kuartal,
    Neraca Kuartal, Arus Kas Kuartal, Roster Gaji 3 Bulan, Analisis & Tren,
    Buku Aset Tetap - tidak ada sheet rekening mentah."""
    months = [load_month(p) for p in paths]
    months = validate_consecutive(months)

    carried = []
    for p in (carry_forward_paths or []):
        carried.extend(load_asset_ledger(p))
    scanned = scan_assets_from_months(months)
    all_assets = merge_asset_lists(carried, scanned)
    rel_assets = assets_with_relative_idx(all_assets, months)

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    income_ref = write_quarterly_income_statement(out_wb, months, rel_assets)
    balance_ref = write_quarterly_balance_sheet(out_wb, months, income_ref, rel_assets)
    write_quarterly_cash_flow(out_wb, months, income_ref, balance_ref)
    roster_summary = write_roster_gaji(out_wb, months)
    write_analysis_sheet(out_wb, months)
    write_buku_aset_tetap(out_wb, all_assets, months)

    order = ["Laba Rugi Kuartal", "Neraca Kuartal", "Arus Kas Kuartal",
             "Roster Gaji 3 Bulan", "Analisis & Tren", "Buku Aset Tetap"]
    out_wb._sheets = [out_wb[s] for s in order]

    out_wb.save(output_path)

    return {
        "periode": [m["label"] for m in months],
        "n_pegawai": roster_summary["n_pegawai"],
        "n_belum_dibayar_bulan_terakhir": roster_summary["n_belum_dibayar_bulan_terakhir"],
        "n_aset_tetap": len(all_assets),
        "n_aset_carry_forward": len(carried),
    }


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 5:
        print("Usage: python quarterly.py bulan1.xlsx bulan2.xlsx bulan3.xlsx output.xlsx "
              "[--carry-forward laporan_lama1.xlsx laporan_lama2.xlsx ...]")
        sys.exit(1)
    paths = sys.argv[1:4]
    out = sys.argv[4]
    carry_forward = []
    if "--carry-forward" in sys.argv:
        idx = sys.argv.index("--carry-forward")
        carry_forward = sys.argv[idx + 1:]
    try:
        s = run_quarterly_report(paths, out, carry_forward_paths=carry_forward)
        print(s)
    except QuarterlyInputError as e:
        print("Error:", e)
        sys.exit(1)
